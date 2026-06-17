"""
build_master_table.py  ->  Buyer_Targeting_Master_Table.xlsx

The definitive single drill-down command table for a patent sale/licensing campaign.
One row per (package x shortlisted buyer). 54 columns in 10 grouped blocks:

  PRIORITY/IDENTITY | OFFERING (package+assets) | BUYER | HOW IDENTIFIED (sourcing
  signal A-G) | BUY-LIKELIHOOD FACTORS | VERDICT | RISK/WHY-NOT | FIT & MATCH |
  APPROACH & ECONOMICS | PIPELINE

Derived columns are formulas (Priority, Motivation, Capacity, Composite, Verdict,
Expected value, Multi-signal). Excel Table + AutoFilter + collapsible column groups
+ heat-map / data-bar / verdict conditional formatting. Arial; zero formula errors.
Sample data illustrative (fictional buyers). Companion 'Legend & Methodology' sheet.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill as PF, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gl
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY='1F3864'; STEEL='44546A'; TEAL='005B7F'; BLUE='2E75B6'; LBLUE='D6E4F0'; LSTEEL='EBF3FB'
WHITE='FFFFFF'; LGRAY='F5F5F5'; MGRAY='D9D9D9'; DGRAY='595959'; INK='1A1A1A'; INPUT='0000CC'
GREEN='2E7D32'; LGREEN='E8F5E9'; AMBER='E69500'; LAMBER='FFF3E0'; RED='C62828'; LRED='FDECEA'
GOLD='B8860B'; LGOLD='FFF7DC'; PURPLE='5B2C83'
THIN=Side(style='thin',color='D0D0D0'); MED=Side(style='medium',color=STEEL)
TB=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def BG(h): return PF('solid',start_color=h,end_color=h)
def AL(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def C(ws,r,c,v=None,bold=False,fg=INK,bg=None,sz=9,align='left',wrap=False,italic=False,border=True,fmt=None):
    cc=ws.cell(r,c,v); cc.font=Font(name='Arial',bold=bold,color=fg,size=sz,italic=italic)
    if bg: cc.fill=BG(bg)
    cc.alignment=AL(align,wrap=wrap)
    if border: cc.border=TB
    if fmt: cc.number_format=fmt
    return cc

# ── package constants (the offerings) ──────────────────────────────────────────
PKG={
 'PKG-01':dict(domain='Wireless › PHY (5G NR)',n=3,anchor='A-01 LDPC channel-coding',ask=4.5,eou='Anchor charted',sep='Potential SEP (3GPP)',title='Clean'),
 'PKG-02':dict(domain='Cybersecurity › Network',n=3,anchor='A-04 DDoS / flood protection',ask=2.2,eou='Charted',sep='No',title='Clean'),
 'PKG-03':dict(domain='Battery › Cathode (NMC/LFP)',n=2,anchor='A-07 High-nickel NMC811',ask=6.0,eou='Partial',sep='No',title='Clean'),
 'PKG-04':dict(domain='Cybersecurity › Data-at-rest',n=2,anchor='A-09 Storage encryption & key dist.',ask=3.0,eou='Charted',sep='No',title='1 assignment gap'),
}

# ── rows: each (package x buyer). f=[FTO,Lit,Cite,Fit,Std,Fin,Appe,Urg] ─────────
import csv as _csv
ROWS=[]
for _r in _csv.DictReader(open('buyer_scores.csv',newline='')):
    _r['rev']=int(float(_r['rev'])); _r['ip']=int(float(_r['ip']))
    _r['nsig']=int(float(_r['nsig'])); _r['sem']=int(float(_r['sem']))
    _r['f']=[int(_r.pop('f%d'%i)) for i in range(8)]
    ROWS.append(_r)

# compute composite + priority in python for sorting
W=[2,1,1,2,1,1.5,1,1.5]
def comp(f): return round(sum(W[i]*f[i] for i in range(8))/sum(W)*20)
def prio(c,n): return 'P1' if ((c>=75 and n>=2) or (c>=65 and n>=3)) else ('P2' if (c>=60 or (c>=55 and n>=2)) else 'P3')
for x in ROWS: x['_c']=comp(x['f']); x['_p']=prio(x['_c'],x['nsig'])
ROWS.sort(key=lambda x:({'P1':0,'P2':1,'P3':2}[x['_p']], -x['_c']))

# ── column schema: (header, key-or-None, width, fmt, wrap) ; None key = formula ──
COLS=[
 ('Priority',None,8,None,False),('Buyer','buyer',20,None,False),('Package','pkg',9,None,False),
 # offering
 ('Pkg domain','domain',22,None,True),('# Assets','n',7,None,False),('Anchor asset','anchor',24,None,True),
 ('Asking ($M)','ask',10,'0.0',False),('EoU','eou',13,None,True),('SEP / std','sep',16,None,True),('Title','title',14,None,True),
 # buyer
 ('Archetype','arch',10,None,False),('Sector','sector',16,None,True),('HQ','hq',6,None,False),
 ('Rev ($M)','rev',9,'#,##0',False),('IP (#)','ip',8,'#,##0',False),('Prior buyer','prior',7,None,False),
 # how identified
 ('Sourcing channel','chan',26,None,True),('Signal category','cat',18,None,True),('# Signals','nsig',8,None,False),
 ('Multi-signal',None,10,None,False),('Trigger event','trig',38,None,True),('Evidence source','evid',20,None,True),('Recency','rec',9,None,False),
 # factors
 ('FTO','f0',6,None,False),('Litig.','f1',6,None,False),('Cite','f2',6,None,False),('Fit','f3',6,None,False),
 ('Std','f4',6,None,False),('Fin','f5',6,None,False),('Appe','f6',6,None,False),('Urg','f7',6,None,False),
 ('Motivation',None,10,None,False),('Capacity',None,10,None,False),
 # verdict
 ('Composite',None,10,None,False),('Verdict',None,10,None,False),('Exp. value ($M)',None,11,'0.0',False),
 # risk
 ('Key headwind','hw',30,None,True),('Severity','sev',8,None,False),('Design-around','da',9,None,False),('DJ/approach risk','dj',9,None,False),
 # fit & match
 ('Bundle match','bundle',34,None,True),('Semantic fit','sem',9,None,False),('Products that read','prod',22,None,True),
 # approach & economics
 ('Deal structure','deal',26,None,True),('Time-to-close','ttc',16,None,True),('Recommended hook','hook',46,None,True),
 ('Outreach channel','out',28,None,True),('Target contact','con',20,None,True),
 # pipeline
 ('Stage','stage',12,None,False),('Owner','owner',8,None,False),('Next action','nx',26,None,True),
 ('Next date','nxd',11,None,False),('Notes','note',34,None,True),
]
NCOL=len(COLS)  # 54
# column-letter lookups by header
L={h:gl(i+1) for i,(h,*_) in enumerate(COLS)}

# group bands: (label, start_header, end_header, color)
GROUPS=[('PRIORITY / IDENTITY','Priority','Package',NAVY),
        ('OFFERING — package & assets','Pkg domain','Title',TEAL),
        ('BUYER','Archetype','Prior buyer',STEEL),
        ('HOW IDENTIFIED — sourcing signal','Sourcing channel','Recency',PURPLE),
        ('BUY-LIKELIHOOD FACTORS (0–5)','FTO','Capacity',BLUE),
        ('VERDICT','Composite','Exp. value ($M)',GREEN),
        ('RISK / WHY-NOT','Key headwind','DJ/approach risk',RED),
        ('FIT & MATCH','Bundle match','Products that read',GOLD),
        ('APPROACH & ECONOMICS','Deal structure','Target contact',STEEL),
        ('PIPELINE','Stage','Notes',TEAL)]

wb=Workbook(); ws=wb.active; ws.title='Master Targeting Table'; ws.sheet_view.showGridLines=False
ws.sheet_properties.tabColor=NAVY
C(ws,1,1,'BUYER-TARGETING MASTER TABLE  ·  patent sale & licensing campaign command sheet',bold=True,fg=WHITE,bg=NAVY,sz=14)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NCOL); ws.row_dimensions[1].height=24
C(ws,2,1,'One row per (package × shortlisted buyer). Sorted P1→P3. Filter any column · collapse column groups with the [-] bars · blue = inputs, formulas auto-compute. Sample data illustrative.',italic=True,fg=DGRAY,sz=9)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=NCOL)
# group band row 3
hidx={h:i+1 for i,(h,*_) in enumerate(COLS)}
for lab,h0,h1,clr in GROUPS:
    c0,c1=hidx[h0],hidx[h1]
    ws.merge_cells(start_row=3,start_column=c0,end_row=3,end_column=c1)
    C(ws,3,c0,lab,bold=True,fg=WHITE,bg=clr,sz=9,align='center')
ws.row_dimensions[3].height=18
# header row 4
HR=4
for i,(h,*_rest) in enumerate(COLS):
    C(ws,HR,i+1,h,bold=True,fg=WHITE,bg=STEEL,sz=8,align='center',wrap=True)
    ws.column_dimensions[gl(i+1)].width=_rest[1]
ws.row_dimensions[HR].height=40

# data rows
def fcol(key,r):  # factor f0..f7 -> letters
    return f'{L[{"f0":"FTO","f1":"Litig.","f2":"Cite","f3":"Fit","f4":"Std","f5":"Fin","f6":"Appe","f7":"Urg"}[key]]}{r}'
for ri,x in enumerate(ROWS):
    r=HR+1+ri; p=PKG[x['pkg']]
    rowvals={}
    for i,(h,key,w,fmt,wrap) in enumerate(COLS):
        col=i+1
        if key is None: continue
        if key in ('domain','n','anchor','ask','eou','sep','title'): v=p[key]
        elif key=='f0': v=x['f'][0]
        elif key.startswith('f') and key[1:].isdigit(): v=x['f'][int(key[1:])]
        else: v=x.get(key,'')
        al='center' if h in ('# Assets','Asking ($M)','HQ','Rev ($M)','IP (#)','Prior buyer','# Signals','Recency','FTO','Litig.','Cite','Fit','Std','Fin','Appe','Urg','Severity','Design-around','DJ/approach risk','Semantic fit','Archetype','Package','Stage','Owner','Next date') else 'left'
        C(ws,r,col,v,sz=8,align=al,wrap=wrap,fmt=fmt)
    # formula columns
    yfac=f'{L["FTO"]}{r}:{L["Urg"]}{r}'
    C(ws,r,hidx['Motivation'],f'=ROUND(AVERAGE({L["FTO"]}{r},{L["Litig."]}{r},{L["Cite"]}{r},{L["Urg"]}{r})*20,0)',sz=8,align='center')
    C(ws,r,hidx['Capacity'],f'=ROUND(AVERAGE({L["Fit"]}{r},{L["Std"]}{r},{L["Fin"]}{r},{L["Appe"]}{r})*20,0)',sz=8,align='center')
    C(ws,r,hidx['Composite'],f'=ROUND(SUMPRODUCT({{2,1,1,2,1,1.5,1,1.5}},{yfac})/11*20,0)',bold=True,sz=9,align='center')
    cc=f'{L["Composite"]}{r}'; ns=f'{L["# Signals"]}{r}'
    C(ws,r,hidx['Verdict'],f'=IF({cc}>=75,"Likely",IF({cc}>=55,"Possible","Unlikely"))',sz=8,align='center')
    C(ws,r,hidx['Exp. value ($M)'],f'=ROUND({cc}/100*{L["Asking ($M)"]}{r},1)',sz=8,align='center',fmt='0.0')
    C(ws,r,hidx['Priority'],f'=IF(OR(AND({cc}>=75,{ns}>=2),AND({cc}>=65,{ns}>=3)),"P1",IF(OR({cc}>=60,AND({cc}>=55,{ns}>=2)),"P2","P3"))',bold=True,sz=9,align='center')
    C(ws,r,hidx['Multi-signal'],f'=IF({ns}>=2,"◆ Multi","Single")',sz=8,align='center')
NROW=HR+len(ROWS)

# Excel Table
tab=Table(displayName='MasterTargeting',ref=f'A{HR}:{gl(NCOL)}{NROW}')
tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False)
ws.add_table(tab)

# conditional formatting
rng=lambda h: f'{L[h]}{HR+1}:{L[h]}{NROW}'
ws.conditional_formatting.add(f'{L["FTO"]}{HR+1}:{L["Urg"]}{NROW}',
    ColorScaleRule(start_type='num',start_value=0,start_color='F8696B',mid_type='num',mid_value=2.5,mid_color='FFEB84',end_type='num',end_value=5,end_color='63BE7B'))
ws.conditional_formatting.add(rng('# Signals'),DataBarRule(start_type='num',start_value=0,end_type='num',end_value=3,color=PURPLE))
ws.conditional_formatting.add(rng('Composite'),DataBarRule(start_type='num',start_value=0,end_type='num',end_value=100,color=BLUE))
ws.conditional_formatting.add(rng('Exp. value ($M)'),DataBarRule(start_type='num',start_value=0,end_type='num',end_value=6,color=TEAL))
ws.conditional_formatting.add(rng('Semantic fit'),ColorScaleRule(start_type='num',start_value=0,start_color='FFFFFF',end_type='num',end_value=100,end_color='63BE7B'))
ws.conditional_formatting.add(rng('Motivation'),ColorScaleRule(start_type='num',start_value=0,start_color='FFFFFF',end_type='num',end_value=100,end_color=BLUE))
ws.conditional_formatting.add(rng('Capacity'),ColorScaleRule(start_type='num',start_value=0,start_color='FFFFFF',end_type='num',end_value=100,end_color=BLUE))
def textrule(h,word,fill,fg):
    ws.conditional_formatting.add(rng(h),FormulaRule(formula=[f'${L[h]}{HR+1}="{word}"'],fill=BG(fill),font=Font(name='Arial',bold=True,color=fg,size=8)))
for w_,fl,fgc in [('Likely',LGREEN,GREEN),('Possible',LAMBER,AMBER),('Unlikely',LRED,RED)]: textrule('Verdict',w_,fl,fgc)
for w_,fl,fgc in [('P1',GREEN,WHITE),('P2',LAMBER,AMBER),('P3',LGRAY,DGRAY)]:
    ws.conditional_formatting.add(rng('Priority'),FormulaRule(formula=[f'${L["Priority"]}{HR+1}="{w_}"'],fill=BG(fl),font=Font(name='Arial',bold=True,color=fgc,size=9)))
ws.conditional_formatting.add(rng('Multi-signal'),FormulaRule(formula=[f'${L["Multi-signal"]}{HR+1}="◆ Multi"'],fill=BG(LGOLD),font=Font(name='Arial',bold=True,color=GOLD,size=8)))
for h in ['Severity','Design-around','DJ/approach risk']:
    textrule(h,'H',LRED,RED); textrule(h,'M',LAMBER,AMBER); textrule(h,'L',LGREEN,GREEN)

# collapsible column groups (detail blocks)
for h0,h1 in [('FTO','Capacity'),('Key headwind','DJ/approach risk'),('Deal structure','Target contact'),('Stage','Notes')]:
    for c in range(hidx[h0],hidx[h1]+1): ws.column_dimensions[gl(c)].outlineLevel=1
ws.sheet_properties.outlinePr.summaryRight=True
ws.freeze_panes=f'D{HR+1}'   # freeze Priority/Buyer/Package + header

# ── Legend & Methodology sheet ─────────────────────────────────────────────────
lg=wb.create_sheet('Legend & Methodology'); lg.sheet_view.showGridLines=False; lg.sheet_properties.tabColor=GOLD
C(lg,1,1,'LEGEND & METHODOLOGY',bold=True,fg=WHITE,bg=NAVY,sz=14); lg.merge_cells('A1:D1')
def sec(r,t,clr=STEEL):
    C(lg,r,1,t,bold=True,fg=WHITE,bg=clr,sz=10); lg.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
def kv(r,a,b):
    C(lg,r,1,a,bold=True,sz=9,bg=LGRAY); C(lg,r,2,b,sz=9,wrap=True); lg.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
for col,w in {'A':30,'B':30,'C':30,'D':30}.items(): lg.column_dimensions[col].width=w
r=3; sec(r,'SOURCING SIGNAL CATEGORIES (how a buyer was identified)',PURPLE); r+=1
for a,b in [('A · In-market','Already acquiring: USPTO assignment-recordation mining, broker buy-side mandates, marketplace inbound, pool-formation calls, auction underbidders.'),
 ('B · Legal-pressure','Conflict-driven: office-action 102/103 citations against their apps, litigation defendants / ITC respondents, PTAB participants, opposition parties, settled/next-in-line defendants.'),
 ('C · Tech-trajectory','Derived: forward-citation mining, filing-velocity, semantic portfolio overlap (IDF tag match), standards-contribution gaps, whitespace inversion (product without patents).'),
 ('D · Commercial-motion','Adjacency triggers: hiring/R&D signals, funding rounds, M&A, product launches & teardowns, import/shipment records, 10-K risk language.'),
 ('E · Financial/structural','Capital seeking assets: defensive aggregators (AST IP3/RPX), IP funds, litigation finance, IP-backed lending, pool net-payers.'),
 ('F · Relationship','Ecosystem-derived: licensees of adjacent portfolios, cross-package adjacency, expiring cross-licences, JV partners, prosecution-counsel channel.'),
 ('G · Self-generated','Your listing as sensor: marketplace analytics, strategic teaser placement, citation-alert subscriptions.')]:
    kv(r,a,b); lg.row_dimensions[r].height=30; r+=1
r+=1; sec(r,'BUY-LIKELIHOOD FACTORS (0–5 each)'); r+=1
for a,b in [('FTO Exposure','Do their products read on the assets? (freedom-to-operate risk)'),
 ('Litigation Pressure','Active disputes / NPE campaigns creating counter-assertion need'),
 ('Citation Linkage','Do they cite / are they blocked by the assets'),
 ('Strategic Fit','Domain & roadmap alignment (tag/semantic fit)'),
 ('Standards / SEP','Standards essentiality / FRAND relevance'),
 ('Financial Capacity','Ability to fund the acquisition'),
 ('Acquisition Appetite','History & posture toward buying IP'),
 ('Urgency','Time pressure to act now')]:
    kv(r,a,b); r+=1
r+=1; sec(r,'DERIVED SCORES & DECISION LOGIC',GREEN); r+=1
for a,b in [('Motivation','avg(FTO, Litigation, Citation, Urgency) × 20  → quadrant x-axis'),
 ('Capacity & Fit','avg(Fit, Standards, Financial, Appetite) × 20  → quadrant y-axis'),
 ('Composite','weighted mean × 20 — FTO & Fit ×2, Financial & Urgency ×1.5, others ×1 (0–100)'),
 ('Verdict','Likely ≥ 75 · Possible 55–74 · Unlikely < 55'),
 ('Multi-signal','◆ flagged when ≥ 2 sourcing channels trip — a priority amplifier'),
 ('Priority','P1 = composite ≥ 75 AND ≥ 2 signals · P2 = composite ≥ 65 OR (≥ 55 with ≥ 2 signals) · P3 = otherwise'),
 ('Expected value','Composite/100 × package asking ($M) — for EV-ranked pursuit')]:
    kv(r,a,b); lg.row_dimensions[r].height=26; r+=1
r+=1; sec(r,'DEAL STRUCTURE · TIME-TO-CLOSE (by archetype)',TEAL); r+=1
for a,b in [('OC-OFF (offensive op-co)','Outright or exclusive licence · Medium (3–5 mo) · hook: product-mapping + EoU + royalty base'),
 ('OC-DEF (defensive op-co)','Outright sale (shield) · Long (6–9 mo) · hook: counter-assertion/FTO, clean title, quiet sale notice'),
 ('OC-EXP (expansion)','Outright or field-of-use · Medium–Long · hook: market-entry blocking + convergence'),
 ('IP-FUND','Outright sale · Medium · hook: citation strength, adjacent re-read, resale optionality / IRR'),
 ('STD-LIC','Non-exclusive / FRAND licence · Long · hook: SEP essentiality + FRAND leverage'),
 ('DEF-AGG','Outright volume lot, fixed price · Short (1–3 mo) · hook: volume/price, clean title, catch-and-release')]:
    kv(r,a,b); lg.row_dimensions[r].height=26; r+=1
r+=1; sec(r,'RISK FLAGS & PIPELINE',RED); r+=1
kv(r,'Severity / Design-around / DJ risk','H / M / L — high is worse. DJ risk = declaratory-judgment exposure from outreach; approach as a sale notice, never an infringement threat.'); lg.row_dimensions[r].height=30; r+=1
kv(r,'Stage','Not started → Contacted → NDA → Diligence → Offer → Closed / Passed'); r+=1
C(lg,r+1,1,'All sample companies are fictional and illustrative. Replace rows on the Master Targeting Table; every formula recomputes.',italic=True,fg=DGRAY,sz=9); lg.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=4)

# ── Data Lineage sheet (engine provenance) ──
lin=wb.create_sheet('Data Lineage'); lin.sheet_view.showGridLines=False; lin.sheet_properties.tabColor=PURPLE
C(lin,1,1,'DATA LINEAGE — how each score is derived',bold=True,fg=WHITE,bg=NAVY,sz=14); lin.merge_cells('A1:D1')
C(lin,2,1,'This table is populated by score_buyers.py from the classification/graph engine + signal feeds. Tiers: T1 verifiable fact · T2 derived from patent data · T3 diligence-grade (manual) · T4 market context. Edit buyer_scores.csv (or cells) to override; formulas recompute.',italic=True,fg=DGRAY,sz=9,wrap=True); lin.merge_cells('A2:D2'); lin.row_dimensions[2].height=40
for j,h in enumerate(['Field','Tier','Source / method','Mode']):
    C(lin,4,j+1,h,bold=True,fg=WHITE,bg=STEEL,sz=9,align='center')
for j,w in enumerate([22,12,46,18]): lin.column_dimensions[gl(j+1)].width=w
import csv as _csv2
for i,_p in enumerate(_csv2.reader(open('scoring_provenance.csv'))):
    if i==0: continue
    r=4+i; bgc=LGRAY if i%2 else WHITE
    for j,v in enumerate(_p): C(lin,r,j+1,v,sz=9,bold=(j==0),wrap=(j==2),bg=bgc)
    lin.row_dimensions[r].height=24
C(lin,4+8+2,1,'Engine-derived (T2): Strategic Fit, Semantic fit, Citation Linkage, FTO proxy.  Feed-derived (T1/T4): Litigation, Appetite, Standards, Financial.  Manual (T3): FTO/EoU confirmation, deal nuance, pipeline.',italic=True,fg=DGRAY,sz=9,wrap=True); lin.merge_cells(start_row=4+8+2,start_column=1,end_row=4+8+2,end_column=4); lin.row_dimensions[4+8+2].height=30

wb.save('Buyer_Targeting_Master_Table.xlsx')
print('saved v2 (engine-populated); cols',NCOL,'rows',len(ROWS),'| P1/P2/P3:',
      sum(1 for x in ROWS if x['_p']=='P1'),sum(1 for x in ROWS if x['_p']=='P2'),sum(1 for x in ROWS if x['_p']=='P3'))
