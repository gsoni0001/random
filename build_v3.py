"""Patent Bundling Template v3 — Pattern A configurability.

Adds:
- Configuration sheet (dropdown to pick preset, Edit Mode toggle, all toggles + thresholds)
- Presets sheet with 6 starter presets + room for 4 custom presets
- Bundle Assignment formulas rewired to honor enabled/disabled toggles and configurable thresholds
- Bundle Quality Scorecard rewired to show DISABLED for off bundles
- Disabled bundles get a [DISABLED] header marker (preserving visibility)
- Instructions panel explaining how to save/load custom presets
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

SRC = '/home/claude/Patent_Bundling_Template_v2.xlsx'
OUT = '/home/claude/Patent_Bundling_Template_v3.xlsx'

wb = load_workbook(SRC)

# -----------------------------------------------------------
# Styles
# -----------------------------------------------------------
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUB_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E78')
SECTION_FILL = PatternFill('solid', start_color='DDEBF7')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', bold=True, size=10)
BLUE_INPUT = Font(name='Arial', size=10, color='0000FF', bold=True)
BLACK_FORMULA = Font(name='Arial', size=10, color='000000')
GREEN_LINK = Font(name='Arial', size=10, color='008000')
DISABLED_FONT = Font(name='Arial', size=10, color='A6A6A6', italic=True)
DISABLED_FILL = PatternFill('solid', start_color='F2F2F2')

# ============================================================
# Bundle metadata — used to build Configuration and Presets sheets
# ============================================================
# (number, short name, default threshold notes, default-on?)
BUNDLES = [
    (1, 'Tech Domain', '', True),
    (2, 'SEP', 'B1 SEP cutoff', True),
    (3, 'Product Architecture', '', True),
    (4, 'Stack Layer', '', True),
    (5, 'Use-Case', '', True),
    (6, 'Manufacturing / Process', '', True),
    (7, 'Materials & Chemistry', '', True),
    (8, 'Algorithm / Software', '', True),
    (9, 'Interoperability', 'B3 Interface cutoff', True),
    (10, 'Generational Roadmap', '', True),
    (11, 'Claim-Type', '', True),
    (12, 'Detectability', 'D1/D2 cutoff', True),
    (13, 'Geographic', '', True),
    (14, 'Family-Tree', 'E1 min family size', True),
    (15, 'Lifecycle / Term', '', True),
    (16, 'Foundational + Improvement', '', True),
    (17, 'Cross-Industry', 'G3 cutoff', True),
    (18, 'Convergent Theme', '', True),
    (19, 'Defensive / Counter-Assert.', 'D3 cutoff', True),
    (20, 'Whitespace / Design-Around', 'C4 cutoff', True),
    (21, 'Prosecution-Status', '', True),
    (22, 'Anchor-and-Halo', 'H1 anchor cutoff', True),
    (23, 'Picket-Fence', '', True),
    (24, 'Strong-Core + Tail', '', True),
    (25, 'Continuation-Live', '', True),
    (26, 'EoU-Backed', '', True),
    (27, 'Battle-Tested', '', True),
    (28, 'Clean-Title', '', True),
    (29, 'High-Citation', 'H5 citation threshold', True),
    (30, 'Adjacent Re-Read', '', True),
    (31, 'Salvage / Volume Lot', '', True),
    (32, 'Pre-Expiry', 'E4 pre-expiry window', True),
    (33, 'Provenance-Coherent', '', True),
]

# Tunable thresholds: (param_name, default_value, description)
THRESHOLDS = [
    ('SEP_B1_cutoff',          2, 'B2 SEP bundle: minimum B1 SEP potential (default 2)'),
    ('Interface_B3_cutoff',    2, 'B9 Interop bundle: minimum B3 interface role (default 2)'),
    ('Detect_D1_cutoff',       2, 'B12 Detectability: D1 external-detect cutoff (default 2)'),
    ('Detect_D2_cutoff',       2, 'B12 Detectability: D2 teardown cutoff (default 2)'),
    ('Family_E1_min',          2, 'B14 Family bundle: minimum E1 family size (default 2)'),
    ('CrossIndustry_G3_cutoff',2, 'B17 Cross-Industry: minimum G3 cross-industry score (default 2)'),
    ('Defensive_D3_cutoff',    2, 'B19 Defensive: minimum D3 reads-on-products (default 2)'),
    ('Whitespace_C4_cutoff',   2, 'B20 Whitespace: minimum C4 design-around difficulty (default 2)'),
    ('Anchor_H1_cutoff',       2, 'B22 Anchor+Halo: minimum H1 claim strength for anchor (default 2)'),
    ('HighCitation_H5_min',   15, 'B29 High-Citation: minimum H5 forward citations (default 15)'),
    ('PreExpiry_min_years',    1, 'B32 Pre-Expiry: minimum remaining term in years (default 1)'),
    ('PreExpiry_max_years',    4, 'B32 Pre-Expiry: maximum remaining term in years (default 4)'),
    ('Salvage_H1_max',         1, 'B31 Salvage: maximum H1 (default 1)'),
    ('Salvage_E4_max',         5, 'B31 Salvage: maximum remaining term in years (default 5)'),
    ('Salvage_H2_max',         1, 'B31 Salvage: maximum H2 prior-art exposure (default 1)'),
    ('Strength_depth_min',     4, 'Scorecard STRONG flag: minimum coverage depth (default 4)'),
    ('Strength_detect_min',    2, 'Scorecard STRONG flag: minimum avg detectability (default 2)'),
    ('Strength_term_min',     10, 'Scorecard STRONG flag: minimum avg remaining term (default 10)'),
]

# Quality gates
GATES = [
    ('Gate_WeakestH1',         True, 'Show "Weakest H1 in bundle" column'),
    ('Gate_InvalidityExposure',True, 'Show "Invalidity exposure %" column'),
    ('Gate_EoUReady',          True, 'Show "EoU-ready %" column'),
    ('Gate_Survived',          True, 'Show "Survived %" column'),
    ('Gate_ContOptionality',   True, 'Show "Continuation-optionality %" column'),
]

# Starter presets — each is a dict of param_name -> value
PRESETS = [
    {
        'name': 'All ON (default)',
        'description': 'Every bundle and gate enabled, all thresholds at default. Baseline view.',
        'bundles': {i: True for i in range(1, 34)},
        'thresholds': {p[0]: p[1] for p in THRESHOLDS},
        'gates': {g[0]: True for g in GATES},
    },
    {
        'name': 'NPE / Counter-Assertion',
        'description': 'Focus on assertion-ready assets: detectability, EoU, survived challenges, defensive reads, salvage lots.',
        'bundles': {
            i: i in {2, 5, 9, 11, 12, 13, 17, 19, 22, 26, 27, 29, 31}
            for i in range(1, 34)
        },
        'thresholds': {
            'Detect_D1_cutoff': 2, 'Detect_D2_cutoff': 2, 'Defensive_D3_cutoff': 2,
            'HighCitation_H5_min': 10,
        },
        'gates': {'Gate_WeakestH1': True, 'Gate_InvalidityExposure': True,
                  'Gate_EoUReady': True, 'Gate_Survived': True,
                  'Gate_ContOptionality': False},
    },
    {
        'name': 'Operating Company / FTO',
        'description': 'Focus on subsystem coverage, interfaces, claim breadth, continuation optionality. Skips litigation-focused bundles.',
        'bundles': {
            i: i in {1, 3, 4, 5, 6, 7, 8, 9, 13, 14, 16, 17, 20, 22, 25, 28, 33}
            for i in range(1, 34)
        },
        'thresholds': {
            'Family_E1_min': 3, 'Whitespace_C4_cutoff': 2,
        },
        'gates': {'Gate_WeakestH1': True, 'Gate_InvalidityExposure': True,
                  'Gate_EoUReady': False, 'Gate_Survived': False,
                  'Gate_ContOptionality': True},
    },
    {
        'name': 'Defensive Aggregator',
        'description': 'Volume-driven: salvage, geographic breadth, claim-type variety, prosecution mix.',
        'bundles': {
            i: i in {1, 4, 10, 11, 13, 15, 17, 21, 24, 28, 31, 32}
            for i in range(1, 34)
        },
        'thresholds': {
            'Salvage_H1_max': 2, 'Salvage_E4_max': 8,
        },
        'gates': {'Gate_WeakestH1': False, 'Gate_InvalidityExposure': True,
                  'Gate_EoUReady': False, 'Gate_Survived': False,
                  'Gate_ContOptionality': False},
    },
    {
        'name': 'Standards Licensee',
        'description': 'Tight focus on standards-essential, interoperability, generational, geographic, battle-tested assets.',
        'bundles': {
            i: i in {2, 9, 10, 12, 13, 22, 26, 27, 28, 29}
            for i in range(1, 34)
        },
        'thresholds': {
            'SEP_B1_cutoff': 2, 'Interface_B3_cutoff': 2,
        },
        'gates': {'Gate_WeakestH1': True, 'Gate_InvalidityExposure': True,
                  'Gate_EoUReady': True, 'Gate_Survived': True,
                  'Gate_ContOptionality': True},
    },
    {
        'name': 'EV Powertrain Sale',
        'description': 'EV-focused: subsystems, materials, manufacturing, thermal, generational. Skips wireless/SEP/AI bundles.',
        'bundles': {
            i: i in {1, 3, 4, 5, 6, 7, 10, 13, 14, 16, 17, 20, 22, 24, 28, 33}
            for i in range(1, 34)
        },
        'thresholds': {
            'CrossIndustry_G3_cutoff': 2,
        },
        'gates': {'Gate_WeakestH1': True, 'Gate_InvalidityExposure': True,
                  'Gate_EoUReady': False, 'Gate_Survived': False,
                  'Gate_ContOptionality': True},
    },
]
# Pad to 10 presets total (6 starter + 4 empty custom slots for users)
for i in range(1, 5):
    PRESETS.append({
        'name': f'Custom Preset {i}',
        'description': '(empty — overwrite this column with your own configuration)',
        'bundles': {i_: True for i_ in range(1, 34)},
        'thresholds': {p[0]: p[1] for p in THRESHOLDS},
        'gates': {g[0]: True for g in GATES},
    })

# ============================================================
# BUILD "Presets" SHEET
# Layout:
#   Col A = parameter name (key)
#   Col B = description
#   Cols C..L = 10 preset columns (6 starter + 4 custom)
# Rows:
#   1 = header
#   2 = preset name row
#   3 = preset description row
#   4 = "BUNDLE TOGGLES" section header
#   5..37 = bundle toggles 1..33
#   38 = "THRESHOLDS" section header
#   39..(39+len(THRESHOLDS)-1) = thresholds
#   next = "QUALITY GATES" section header
#   next+1..+5 = gates
# ============================================================
ws_p = wb.create_sheet('Presets', 1)  # insert after README (index 1)
ws_p.sheet_view.showGridLines = False

# Header row
headers = ['Parameter', 'Description'] + [p['name'] for p in PRESETS]
for i, h in enumerate(headers, 1):
    c = ws_p.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# Preset name row (already in headers row 1)
# Preset description row
ws_p.cell(row=2, column=1, value='_description').font = ARIAL_BOLD
ws_p.cell(row=2, column=2, value='Short description of the preset').font = ARIAL
for i, preset in enumerate(PRESETS, 3):
    c = ws_p.cell(row=2, column=i, value=preset['description'])
    c.font = ARIAL
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = BORDER

# Section header: BUNDLE TOGGLES
ws_p.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
c = ws_p.cell(row=3, column=1, value='BUNDLE TOGGLES (Yes = enabled, No = disabled)')
c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
c.fill = PatternFill('solid', start_color='2E75B6')
c.alignment = LEFT
c.border = BORDER

# Bundle toggle rows
for i, (bnum, bname, _, default_on) in enumerate(BUNDLES):
    r = 4 + i
    param_key = f'B{bnum}_enabled'
    ws_p.cell(row=r, column=1, value=param_key).font = ARIAL_BOLD
    ws_p.cell(row=r, column=2, value=f'Bundle {bnum} — {bname}').font = ARIAL
    ws_p.cell(row=r, column=2).alignment = LEFT
    for j, preset in enumerate(PRESETS, 3):
        val = 'Yes' if preset['bundles'].get(bnum, False) else 'No'
        c = ws_p.cell(row=r, column=j, value=val)
        c.font = ARIAL
        c.alignment = CENTER
        c.border = BORDER

# Section header: THRESHOLDS
thresh_section_row = 4 + len(BUNDLES)
ws_p.merge_cells(start_row=thresh_section_row, start_column=1,
                 end_row=thresh_section_row, end_column=len(headers))
c = ws_p.cell(row=thresh_section_row, column=1,
              value='THRESHOLDS (numeric values used in routing rules)')
c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
c.fill = PatternFill('solid', start_color='2E75B6')
c.alignment = LEFT
c.border = BORDER

# Threshold rows
for i, (tname, tdefault, tdesc) in enumerate(THRESHOLDS):
    r = thresh_section_row + 1 + i
    ws_p.cell(row=r, column=1, value=tname).font = ARIAL_BOLD
    ws_p.cell(row=r, column=2, value=tdesc).font = ARIAL
    ws_p.cell(row=r, column=2).alignment = LEFT
    for j, preset in enumerate(PRESETS, 3):
        val = preset['thresholds'].get(tname, tdefault)
        c = ws_p.cell(row=r, column=j, value=val)
        c.font = ARIAL
        c.alignment = CENTER
        c.border = BORDER

# Section header: QUALITY GATES
gates_section_row = thresh_section_row + 1 + len(THRESHOLDS)
ws_p.merge_cells(start_row=gates_section_row, start_column=1,
                 end_row=gates_section_row, end_column=len(headers))
c = ws_p.cell(row=gates_section_row, column=1,
              value='QUALITY GATES (Yes = column shown on Scorecard, No = column blanked)')
c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
c.fill = PatternFill('solid', start_color='2E75B6')
c.alignment = LEFT
c.border = BORDER

for i, (gname, gdefault, gdesc) in enumerate(GATES):
    r = gates_section_row + 1 + i
    ws_p.cell(row=r, column=1, value=gname).font = ARIAL_BOLD
    ws_p.cell(row=r, column=2, value=gdesc).font = ARIAL
    ws_p.cell(row=r, column=2).alignment = LEFT
    for j, preset in enumerate(PRESETS, 3):
        val = 'Yes' if preset['gates'].get(gname, True) else 'No'
        c = ws_p.cell(row=r, column=j, value=val)
        c.font = ARIAL
        c.alignment = CENTER
        c.border = BORDER

# Borders and widths
last_p_row = gates_section_row + len(GATES)
for r in range(1, last_p_row + 1):
    for col in range(1, len(headers) + 1):
        cell = ws_p.cell(row=r, column=col)
        if cell.border.left.style is None:
            cell.border = BORDER

ws_p.column_dimensions['A'].width = 28
ws_p.column_dimensions['B'].width = 48
for j in range(3, 3 + len(PRESETS)):
    ws_p.column_dimensions[get_column_letter(j)].width = 18
ws_p.row_dimensions[1].height = 32
ws_p.row_dimensions[2].height = 48
ws_p.freeze_panes = 'C4'

# Add an instructions panel below the data
inst_row = last_p_row + 3
instructions = [
    ('HOW TO USE THE PRESETS SHEET', 'h2'),
    ('• The "Active Preset" dropdown on the Configuration sheet picks one column from this sheet.', 'p'),
    ('• Columns C-H are starter presets — useful as-is or as templates.', 'p'),
    ('• Columns I-L are empty "Custom Preset" slots — overwrite their values to define your own.', 'p'),
    ('• To rename a preset: edit row 1 of its column. The dropdown on the Configuration sheet updates automatically.', 'p'),
    ('• To save your current Configuration as a new preset:', 'p'),
    ('  1. On the Configuration sheet, set Edit Mode = Yes and configure all toggles/thresholds the way you want.', 'p'),
    ('  2. Come back to this sheet. In the Custom Preset column you want to use, paste-special each value section by section, or type values directly.', 'p'),
    ('  3. Rename the column header (row 1) to your preset name.', 'p'),
    ('  4. Back on Configuration: set Edit Mode = No and select your new preset from the dropdown.', 'p'),
    ('• You can add more preset columns by extending to columns M onward — just update the dropdown range on the Configuration sheet.', 'p'),
]
for offset, (text, kind) in enumerate(instructions):
    r = inst_row + offset
    c = ws_p.cell(row=r, column=1, value=text)
    if kind == 'h2':
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
    else:
        c.font = ARIAL
        c.alignment = Alignment(wrap_text=True, vertical='top')
    ws_p.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))

# Data validations on toggle rows (Yes/No)
dv_yn = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
ws_p.add_data_validation(dv_yn)
# Bundle toggle rows: rows 4..(4+33-1)=36 ; preset columns C..L (3..12)
for r in range(4, 4 + len(BUNDLES)):
    dv_yn.add(f'C{r}:{get_column_letter(2 + len(PRESETS))}{r}')
# Gate toggle rows
for i in range(len(GATES)):
    r = gates_section_row + 1 + i
    dv_yn.add(f'C{r}:{get_column_letter(2 + len(PRESETS))}{r}')

# ============================================================
# BUILD "Configuration" SHEET
# Layout:
#   Row 1-3: title + sub
#   Row 5: "Active Preset:" label + dropdown cell (B5)
#   Row 6: "Edit Mode:" label + Yes/No dropdown (B6)
#   Row 7: Summary line (counts of active bundles/gates)
#   Row 9+: Bundle toggles section
#   Then: Thresholds section
#   Then: Quality gates section
# Each parameter cell's formula:
#   IF(EditMode="Yes", manual_value_cell, INDEX(Presets!..., MATCH(ActivePreset, Presets!row1, 0)))
# But Excel doesn't let one cell have both a formula AND a manual value.
# Solution: each parameter has a "manual override" column on the right (col E), and the "active value"
# column (col C) shows either the manual override (if Edit Mode = Yes) or the preset value (if No).
# So col C is always formula-driven; col E is the user's hand-entered override.
# ============================================================
ws_c = wb.create_sheet('Configuration', 1)  # immediately after README, before Presets
wb.move_sheet('Configuration', offset=-1)   # ensure Configuration is before Presets

# Actually move sheets to enforce order: README, Configuration, Presets, Attribute Dictionary, ...
desired_order = ['README', 'Configuration', 'Presets', 'Attribute Dictionary',
                 'Patent Portfolio', 'Bundle Rules', 'Bundle Assignment',
                 'Bundle Quality Scorecard', 'Sample Bundles',
                 'Bundle Composition Strategy']
# Get current order
current_order = wb.sheetnames
# Build new ordering by moving sheets as needed
for target_idx, name in enumerate(desired_order):
    if name in wb.sheetnames:
        cur_idx = wb.sheetnames.index(name)
        if cur_idx != target_idx:
            wb.move_sheet(name, offset=target_idx - cur_idx)

ws_c.sheet_view.showGridLines = False

# Title
c = ws_c.cell(row=1, column=1, value='CONFIGURATION')
c.font = Font(name='Arial', bold=True, size=18, color='1F4E78')
ws_c.merge_cells('A1:G1')

c = ws_c.cell(row=2, column=1,
              value='Pick a preset, or switch to Edit Mode to override individual values. All Bundle Assignment and Scorecard formulas read from this sheet.')
c.font = Font(name='Arial', italic=True, size=11, color='595959')
c.alignment = Alignment(wrap_text=True, vertical='top')
ws_c.merge_cells('A2:G2')

# Active Preset dropdown
c = ws_c.cell(row=4, column=1, value='Active Preset:')
c.font = ARIAL_BOLD
c.alignment = Alignment(horizontal='right')

c = ws_c.cell(row=4, column=2, value=PRESETS[0]['name'])
c.font = BLUE_INPUT
c.fill = PatternFill('solid', start_color='FFF2CC')
c.alignment = CENTER
c.border = BORDER
ws_c.merge_cells('B4:D4')

# Validation list = preset names from Presets!C1:L1
dv_preset = DataValidation(type='list',
                           formula1=f"=Presets!$C$1:${get_column_letter(2 + len(PRESETS))}$1",
                           allow_blank=False)
ws_c.add_data_validation(dv_preset)
dv_preset.add('B4')

# Edit Mode toggle
c = ws_c.cell(row=5, column=1, value='Edit Mode:')
c.font = ARIAL_BOLD
c.alignment = Alignment(horizontal='right')

c = ws_c.cell(row=5, column=2, value='No')
c.font = BLUE_INPUT
c.fill = PatternFill('solid', start_color='FFF2CC')
c.alignment = CENTER
c.border = BORDER
ws_c.merge_cells('B5:D5')

dv_em = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
ws_c.add_data_validation(dv_em)
dv_em.add('B5')

# Notes
c = ws_c.cell(row=4, column=5,
              value='← Pick from Presets sheet. Changing this updates all Active Values below.')
c.font = Font(name='Arial', italic=True, size=9, color='595959')
c.alignment = LEFT
ws_c.merge_cells('E4:H4')

c = ws_c.cell(row=5, column=5,
              value='← Yes = use Manual Override column. No = use selected preset.')
c.font = Font(name='Arial', italic=True, size=9, color='595959')
c.alignment = LEFT
ws_c.merge_cells('E5:H5')

# Summary row (will fill with counts via formula after we know the ranges)
c = ws_c.cell(row=6, column=1, value='Summary:')
c.font = ARIAL_BOLD
c.alignment = Alignment(horizontal='right')

# Column headers for the parameter table
table_header_row = 8
ws_c.cell(row=table_header_row, column=1, value='Parameter').font = SUB_HEADER_FONT
ws_c.cell(row=table_header_row, column=1).fill = PatternFill('solid', start_color='404040')
ws_c.cell(row=table_header_row, column=1).alignment = CENTER
ws_c.cell(row=table_header_row, column=1).border = BORDER

ws_c.cell(row=table_header_row, column=2, value='Description').font = SUB_HEADER_FONT
ws_c.cell(row=table_header_row, column=2).fill = PatternFill('solid', start_color='404040')
ws_c.cell(row=table_header_row, column=2).alignment = CENTER
ws_c.cell(row=table_header_row, column=2).border = BORDER

ws_c.cell(row=table_header_row, column=3, value='Active Value (use this in formulas)').font = SUB_HEADER_FONT
ws_c.cell(row=table_header_row, column=3).fill = PatternFill('solid', start_color='1F4E78')
ws_c.cell(row=table_header_row, column=3).alignment = CENTER
ws_c.cell(row=table_header_row, column=3).border = BORDER

ws_c.cell(row=table_header_row, column=4, value='From Preset').font = SUB_HEADER_FONT
ws_c.cell(row=table_header_row, column=4).fill = PatternFill('solid', start_color='808080')
ws_c.cell(row=table_header_row, column=4).alignment = CENTER
ws_c.cell(row=table_header_row, column=4).border = BORDER

ws_c.cell(row=table_header_row, column=5, value='Manual Override').font = SUB_HEADER_FONT
ws_c.cell(row=table_header_row, column=5).fill = PatternFill('solid', start_color='808080')
ws_c.cell(row=table_header_row, column=5).alignment = CENTER
ws_c.cell(row=table_header_row, column=5).border = BORDER

# We need named ranges or formula references to:
#   - ActivePreset (= Configuration!B4)
#   - EditMode (= Configuration!B5)
#   - Presets parameter rows (each parameter is on a known row in Presets sheet)
#   - We will use INDEX/MATCH on Presets!$C$1:$L$1 to find the preset column

PRESETS_FIRST_COL = 3
PRESETS_LAST_COL = 2 + len(PRESETS)
PRESETS_HEADER_ROW = 1
PRESETS_COL_RANGE = f"Presets!$C$1:${get_column_letter(PRESETS_LAST_COL)}$1"

def preset_lookup_formula(presets_row: int) -> str:
    """Return an INDEX/MATCH formula that returns the value at presets_row for the active preset."""
    return (f"INDEX(Presets!$C${presets_row}:${get_column_letter(PRESETS_LAST_COL)}${presets_row},"
            f"MATCH($B$4,{PRESETS_COL_RANGE},0))")

# Row offsets in Presets sheet:
PRESETS_BUNDLE_ROWS = {bnum: 4 + idx for idx, (bnum, *_) in enumerate(BUNDLES)}
PRESETS_THRESHOLD_ROWS = {tname: thresh_section_row + 1 + idx
                          for idx, (tname, *_) in enumerate(THRESHOLDS)}
PRESETS_GATE_ROWS = {gname: gates_section_row + 1 + idx
                     for idx, (gname, *_) in enumerate(GATES)}

# Map each Configuration row to (param_key, description, presets_row, type)
# type: 'YN' (Yes/No) or 'NUM'
config_rows = []
# Section header
section_label_rows = {}

# Section: BUNDLE TOGGLES
config_rows.append(('__SECTION__', 'BUNDLE TOGGLES (Yes = bundle is active in this analysis)', None, None))
for bnum, bname, _, _ in BUNDLES:
    key = f'B{bnum}_enabled'
    desc = f'Bundle {bnum} — {bname}'
    config_rows.append((key, desc, PRESETS_BUNDLE_ROWS[bnum], 'YN'))

config_rows.append(('__SECTION__', 'THRESHOLDS (numeric values referenced in routing rules)', None, None))
for tname, tdefault, tdesc in THRESHOLDS:
    config_rows.append((tname, tdesc, PRESETS_THRESHOLD_ROWS[tname], 'NUM'))

config_rows.append(('__SECTION__', 'QUALITY GATES (Yes = gate column shown on Scorecard)', None, None))
for gname, gdefault, gdesc in GATES:
    config_rows.append((gname, gdesc, PRESETS_GATE_ROWS[gname], 'YN'))

# Write configuration rows
config_first_data_row = table_header_row + 1
config_key_to_row = {}  # param_key -> Configuration row index

current_row = config_first_data_row
for (key, desc, presets_row, ptype) in config_rows:
    if key == '__SECTION__':
        # Section header row
        ws_c.merge_cells(start_row=current_row, start_column=1,
                         end_row=current_row, end_column=5)
        c = ws_c.cell(row=current_row, column=1, value=desc)
        c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='2E75B6')
        c.alignment = LEFT
        c.border = BORDER
        section_label_rows[desc] = current_row
        current_row += 1
        continue

    # Param key
    ws_c.cell(row=current_row, column=1, value=key).font = ARIAL_BOLD
    ws_c.cell(row=current_row, column=1).alignment = LEFT
    ws_c.cell(row=current_row, column=1).border = BORDER

    # Description
    ws_c.cell(row=current_row, column=2, value=desc).font = ARIAL
    ws_c.cell(row=current_row, column=2).alignment = LEFT
    ws_c.cell(row=current_row, column=2).border = BORDER

    # Manual Override (col E) — user-editable; default value = the preset's "All ON (default)" value
    if ptype == 'YN':
        manual_default = 'Yes'
    else:
        # For numeric, pull the default from THRESHOLDS dict
        manual_default = next((t[1] for t in THRESHOLDS if t[0] == key), 1)
    c = ws_c.cell(row=current_row, column=5, value=manual_default)
    c.font = BLUE_INPUT
    c.fill = PatternFill('solid', start_color='FFF2CC')
    c.alignment = CENTER
    c.border = BORDER

    # From Preset (col D) — formula = INDEX/MATCH on Presets sheet
    preset_formula = '=' + preset_lookup_formula(presets_row)
    c = ws_c.cell(row=current_row, column=4, value=preset_formula)
    c.font = GREEN_LINK
    c.alignment = CENTER
    c.border = BORDER

    # Active Value (col C) — IF(EditMode="Yes", ManualOverride, FromPreset)
    active_formula = f'=IF($B$5="Yes",E{current_row},D{current_row})'
    c = ws_c.cell(row=current_row, column=3, value=active_formula)
    c.font = ARIAL_BOLD
    c.alignment = CENTER
    c.border = BORDER
    if ptype == 'YN':
        c.fill = PatternFill('solid', start_color='E2EFDA')
    else:
        c.fill = PatternFill('solid', start_color='FFF2CC')

    config_key_to_row[key] = current_row
    current_row += 1

# Apply data validations on the manual override column
dv_yn_c = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
ws_c.add_data_validation(dv_yn_c)
for (key, _, _, ptype) in config_rows:
    if key in ('__SECTION__',):
        continue
    if ptype == 'YN':
        dv_yn_c.add(f'E{config_key_to_row[key]}')

# Column widths
ws_c.column_dimensions['A'].width = 28
ws_c.column_dimensions['B'].width = 52
ws_c.column_dimensions['C'].width = 26
ws_c.column_dimensions['D'].width = 16
ws_c.column_dimensions['E'].width = 18

ws_c.row_dimensions[1].height = 24
ws_c.row_dimensions[2].height = 38
ws_c.row_dimensions[table_header_row].height = 32
ws_c.freeze_panes = f'A{table_header_row + 1}'

# Summary formulas — count active bundles and gates
# Bundles enabled count = COUNTIF of "Yes" on Active Value column for bundle rows
bundle_active_rows = [config_key_to_row[f'B{b}_enabled'] for b, *_ in BUNDLES]
gate_active_rows = [config_key_to_row[g[0]] for g in GATES]

# We can't easily COUNTIF on a non-contiguous range, but bundle rows ARE contiguous.
first_bundle_row = min(bundle_active_rows)
last_bundle_row = max(bundle_active_rows)
first_gate_row = min(gate_active_rows)
last_gate_row = max(gate_active_rows)

summary_formula = (f'="Bundles active: "&COUNTIF(C{first_bundle_row}:C{last_bundle_row},"Yes")&'
                   f'" of {len(BUNDLES)}    |    Gates active: "&'
                   f'COUNTIF(C{first_gate_row}:C{last_gate_row},"Yes")&" of {len(GATES)}    |    "&'
                   f'IF($B$5="Yes","EDIT MODE — using Manual Override values","Using preset: "&$B$4)')
c = ws_c.cell(row=6, column=2, value=summary_formula)
c.font = Font(name='Arial', bold=True, size=11, color='1F4E78')
c.fill = PatternFill('solid', start_color='DDEBF7')
c.alignment = LEFT
c.border = BORDER
ws_c.merge_cells('B6:H6')

# Conditional formatting for Active Value column — green if Yes, red if No (for YN params)
yn_param_rows = [config_key_to_row[k] for (k, _, _, t) in config_rows
                 if t == 'YN']
for r in yn_param_rows:
    ws_c.conditional_formatting.add(f'C{r}',
        CellIsRule(operator='equal', formula=['"Yes"'],
                   fill=PatternFill('solid', start_color='C6EFCE')))
    ws_c.conditional_formatting.add(f'C{r}',
        CellIsRule(operator='equal', formula=['"No"'],
                   fill=PatternFill('solid', start_color='FFC7CE')))

# Build named references (use direct formula references rather than defined names for clarity)
# We'll build a Python dict that maps each parameter to its address like 'Configuration!$C$10'
CFG = {}
for (key, _, _, _) in config_rows:
    if key == '__SECTION__':
        continue
    CFG[key] = f"Configuration!$C${config_key_to_row[key]}"

# Convenience helpers
def b_enabled(bnum):
    return CFG[f'B{bnum}_enabled']

def thr(name):
    return CFG[name]

def gate_enabled(name):
    return CFG[name]

# ============================================================
# REWIRE "Bundle Assignment" formulas
# - Wrap each bundle's per-patent formula in:
#     =IF(<bundle_enabled>="No","",<existing_qualification_formula>)
# - Replace hardcoded thresholds with cell references to Configuration
# - Update header row 2 to show "[DISABLED]" when bundle is off
# - Total formula counts only TRUE values (blanks won't count, which is what we want)
# ============================================================
ws_ba = wb['Bundle Assignment']
PP = "'Patent Portfolio'"

# Find data row range
n_rows = 0
for r in range(3, 200):
    if ws_ba.cell(row=r, column=1).value or ws_ba.cell(row=r, column=2).value:
        n_rows = r - 2
    else:
        break

start_data_row = 3  # row in Patent Portfolio where data begins

# For each bundle column, rewrite the per-patent formula
# Bundle columns: 3..35 = bundles 1..33
# Pre-compute the qualification body (the OLD formula's inner condition) for each bundle.
# We will use threshold references from Configuration for the bundles that have thresholds.

def qual_body(bundle_num: int, src_row: int) -> str:
    """Return the inner qualification expression (without leading '=') for a bundle and patent row.
    References Patent Portfolio columns by letter."""
    r = src_row
    pp = PP

    if bundle_num == 1:  # Tech Domain — A1 populated
        return f'IFERROR(IF({pp}!C{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 2:  # SEP — B1 >= cutoff AND B2 populated
        return (f'IFERROR(IF(AND(N({pp}!H{r})>={thr("SEP_B1_cutoff")},'
                f'{pp}!I{r}<>""),TRUE,FALSE),FALSE)')
    if bundle_num == 3:  # Product Arch
        return f'IFERROR(IF({pp}!F{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 4:  # Stack Layer
        return f'IFERROR(IF({pp}!E{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 5:  # Use-Case
        return f'IFERROR(IF({pp}!G{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 6:  # Mfg/Process — A1 contains process/fab/manufac AND C1=Method
        return (f'IFERROR(IF(AND(OR(ISNUMBER(SEARCH("process",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("fab",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("manufac",{pp}!C{r}))),{pp}!K{r}="Method"),'
                f'TRUE,FALSE),FALSE)')
    if bundle_num == 7:  # Materials
        return (f'IFERROR(IF(AND(OR(ISNUMBER(SEARCH("material",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("chem",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("battery",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("electrolyte",{pp}!C{r})),'
                f'ISNUMBER(SEARCH("polymer",{pp}!C{r}))),'
                f'OR({pp}!K{r}="Apparatus",{pp}!K{r}="Method")),TRUE,FALSE),FALSE)')
    if bundle_num == 8:  # Algorithm
        return (f'IFERROR(IF(AND(OR({pp}!E{r}="App",{pp}!E{r}="Middleware",{pp}!E{r}="Cloud"),'
                f'OR({pp}!K{r}="Method",{pp}!K{r}="CRM")),TRUE,FALSE),FALSE)')
    if bundle_num == 9:  # Interop — B3 >= cutoff
        return f'IFERROR(IF(N({pp}!J{r})>={thr("Interface_B3_cutoff")},TRUE,FALSE),FALSE)'
    if bundle_num == 10:  # Generational
        return f'IFERROR(IF({pp}!AA{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 11:  # Claim-Type
        return f'IFERROR(IF({pp}!K{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 12:  # Detectability — D1>=cutoff OR D2>=cutoff
        return (f'IFERROR(IF(OR(N({pp}!O{r})>={thr("Detect_D1_cutoff")},'
                f'N({pp}!P{r})>={thr("Detect_D2_cutoff")}),TRUE,FALSE),FALSE)')
    if bundle_num == 13:  # Geographic — F2=Yes
        return f'IFERROR(IF({pp}!X{r}="Yes",TRUE,FALSE),FALSE)'
    if bundle_num == 14:  # Family — E1>=min
        return f'IFERROR(IF(N({pp}!R{r})>={thr("Family_E1_min")},TRUE,FALSE),FALSE)'
    if bundle_num == 15:  # Lifecycle — E4>0
        return f'IFERROR(IF(N({pp}!U{r})>0,TRUE,FALSE),FALSE)'
    if bundle_num == 16:  # Foundational+Improvement — C2=3 OR C2<=1
        return f'IFERROR(IF(OR(N({pp}!L{r})=3,N({pp}!L{r})<=1),TRUE,FALSE),FALSE)'
    if bundle_num == 17:  # Cross-Industry — G3>=cutoff
        return f'IFERROR(IF(N({pp}!AB{r})>={thr("CrossIndustry_G3_cutoff")},TRUE,FALSE),FALSE)'
    if bundle_num == 18:  # Convergent Theme — G1 populated
        return f'IFERROR(IF({pp}!Z{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 19:  # Defensive — D3>=cutoff
        return f'IFERROR(IF(N({pp}!Q{r})>={thr("Defensive_D3_cutoff")},TRUE,FALSE),FALSE)'
    if bundle_num == 20:  # Whitespace — C4>=cutoff AND A1 populated
        return (f'IFERROR(IF(AND(N({pp}!N{r})>={thr("Whitespace_C4_cutoff")},'
                f'{pp}!C{r}<>""),TRUE,FALSE),FALSE)')
    if bundle_num == 21:  # Prosecution-Status — E2=Pending OR E3=Yes
        return f'IFERROR(IF(OR({pp}!S{r}="Pending",{pp}!T{r}="Yes"),TRUE,FALSE),FALSE)'
    if bundle_num == 22:  # Anchor+Halo — H1>=cutoff AND C2>=1
        return (f'IFERROR(IF(AND(N({pp}!AC{r})>={thr("Anchor_H1_cutoff")},'
                f'N({pp}!L{r})>=1),TRUE,FALSE),FALSE)')
    if bundle_num == 23:  # Picket-Fence — C4>=1 AND (A1 OR B2 populated)
        return (f'IFERROR(IF(AND(N({pp}!N{r})>=1,'
                f'OR({pp}!C{r}<>"",{pp}!I{r}<>"")),TRUE,FALSE),FALSE)')
    if bundle_num == 24:  # Strong+Tail — A1 populated
        return f'IFERROR(IF({pp}!C{r}<>"",TRUE,FALSE),FALSE)'
    if bundle_num == 25:  # Continuation-Live — E3=Yes
        return f'IFERROR(IF({pp}!T{r}="Yes",TRUE,FALSE),FALSE)'
    if bundle_num == 26:  # EoU-Backed — H9 in (Partial, Full)
        return f'IFERROR(IF(OR({pp}!AK{r}="Partial",{pp}!AK{r}="Full"),TRUE,FALSE),FALSE)'
    if bundle_num == 27:  # Battle-Tested — H7=Survived
        return f'IFERROR(IF({pp}!AI{r}="Survived",TRUE,FALSE),FALSE)'
    if bundle_num == 28:  # Clean-Title — H8=Clean AND H10=None
        return f'IFERROR(IF(AND({pp}!AJ{r}="Clean",{pp}!AL{r}="None"),TRUE,FALSE),FALSE)'
    if bundle_num == 29:  # High-Citation — H5>=cutoff
        return f'IFERROR(IF(N({pp}!AG{r})>={thr("HighCitation_H5_min")},TRUE,FALSE),FALSE)'
    if bundle_num == 30:  # Adjacent Re-Read — I3>=2 AND G3>=2  (we'll keep these literal since they're already stricter than CrossIndustry cutoff and you can refine later)
        return f'IFERROR(IF(AND(N({pp}!AO{r})>=2,N({pp}!AB{r})>=2),TRUE,FALSE),FALSE)'
    if bundle_num == 31:  # Salvage — H1<=max OR E4<max OR H2<=max
        return (f'IFERROR(IF(OR(N({pp}!AC{r})<={thr("Salvage_H1_max")},'
                f'N({pp}!U{r})<{thr("Salvage_E4_max")},'
                f'N({pp}!AD{r})<={thr("Salvage_H2_max")}),TRUE,FALSE),FALSE)')
    if bundle_num == 32:  # Pre-Expiry — between min and max
        return (f'IFERROR(IF(AND(N({pp}!U{r})>={thr("PreExpiry_min_years")},'
                f'N({pp}!U{r})<={thr("PreExpiry_max_years")}),TRUE,FALSE),FALSE)')
    if bundle_num == 33:  # Provenance — A1 AND A4 populated
        return f'IFERROR(IF(AND({pp}!C{r}<>"",{pp}!F{r}<>""),TRUE,FALSE),FALSE)'
    raise ValueError(f"Unknown bundle {bundle_num}")

# Rewrite each bundle column
for i in range(n_rows):
    src_row = start_data_row + i
    out_row = 3 + i
    for bnum in range(1, 34):
        ba_col = 2 + bnum  # bundle 1 -> col 3, ..., bundle 33 -> col 35
        body = qual_body(bnum, src_row)
        wrapped = f'=IF({b_enabled(bnum)}="No","",{body})'
        c = ws_ba.cell(row=out_row, column=ba_col, value=wrapped)
        c.font = BLACK_FORMULA
        c.alignment = CENTER
        c.border = BORDER

# Update bundle column headers in row 2 to show "[DISABLED]" marker dynamically
# We need a header that changes based on the toggle.
# Approach: header is a formula that returns either "B{n} {name}" or "[DISABLED] B{n} {name}"
bundle_short_names_short = [
    'Tech Domain', 'SEP', 'Product Arch', 'Stack Layer', 'Use-Case',
    'Mfg/Process', 'Materials', 'Algorithm', 'Interop', 'Generational',
    'Claim-Type', 'Detectable', 'Geographic', 'Family', 'Lifecycle',
    'Foundational+Improv.', 'Cross-Industry', 'Convergent Theme',
    'Defensive', 'Whitespace', 'Prosecution-Status',
    'Anchor+Halo', 'Picket-Fence', 'Strong+Tail', 'Continuation-Live',
    'EoU-Backed', 'Battle-Tested', 'Clean-Title', 'High-Citation',
    'Adjacent Re-Read', 'Salvage-Volume', 'Pre-Expiry', 'Provenance',
]
for bnum in range(1, 34):
    ba_col = 2 + bnum
    short = bundle_short_names_short[bnum - 1]
    header_formula = (f'=IF({b_enabled(bnum)}="No","[DISABLED] B{bnum} {short}",'
                      f'"B{bnum} {short}")')
    c = ws_ba.cell(row=2, column=ba_col, value=header_formula)
    c.font = SUB_HEADER_FONT
    c.fill = PatternFill('solid', start_color='404040')
    c.alignment = CENTER
    c.border = BORDER

# Conditional formatting on bundle columns: gray out disabled columns
# Apply formula-based rule: if the header in row 2 starts with "[DISABLED]" -> gray fill
last_assign_row = 2 + n_rows
for bnum in range(1, 34):
    ba_col_letter = get_column_letter(2 + bnum)
    rng = f'{ba_col_letter}3:{ba_col_letter}{last_assign_row}'
    # Rule using formula referring to row 2 of this column
    ws_ba.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("[DISABLED]",{ba_col_letter}$2))'],
                    fill=DISABLED_FILL, font=DISABLED_FONT, stopIfTrue=True))

# Re-add the green/red TRUE/FALSE conditional formatting after the disabled rule
green_fill = PatternFill('solid', start_color='C6EFCE')
red_fill = PatternFill('solid', start_color='FFC7CE')
ws_ba.conditional_formatting.add(f'C3:AI{last_assign_row}',
    CellIsRule(operator='equal', formula=['TRUE'], fill=green_fill))
ws_ba.conditional_formatting.add(f'C3:AI{last_assign_row}',
    CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill))

# Update Total column formula (col 36 / AJ) — still COUNTIF TRUE; blanks won't count
for i in range(n_rows):
    out_row = 3 + i
    last_col_letter = get_column_letter(35)
    c = ws_ba.cell(row=out_row, column=36,
                   value=f'=COUNTIF(C{out_row}:{last_col_letter}{out_row},TRUE)')
    c.font = ARIAL_BOLD
    c.alignment = CENTER
    c.border = BORDER

# ============================================================
# REWIRE "Bundle Quality Scorecard"
# - Bundle name column (col B): show "[DISABLED] <name>" when bundle is off
# - All metric columns: return "" when bundle is disabled
# - Strength flag uses configurable thresholds
# - Gate columns (K..O): return "" when their gate toggle is off
# - Bundle name and metric cells get gray-out conditional formatting when disabled
# ============================================================
ws_sc = wb['Bundle Quality Scorecard']

# Patent Portfolio range constants
pp_first = 3
n_pp_rows = 20  # samples + blanks
pp_last = pp_first + n_pp_rows - 1
BA = "'Bundle Assignment'"

# Patent Portfolio column references
d1 = f"{PP}!O{pp_first}:O{pp_last}"
d2 = f"{PP}!P{pp_first}:P{pp_last}"
e4 = f"{PP}!U{pp_first}:U{pp_last}"
f2 = f"{PP}!X{pp_first}:X{pp_last}"
e3 = f"{PP}!T{pp_first}:T{pp_last}"
b1 = f"{PP}!H{pp_first}:H{pp_last}"
c2 = f"{PP}!L{pp_first}:L{pp_last}"
h1_rng = f"{PP}!AC{pp_first}:AC{pp_last}"
h2_rng = f"{PP}!AD{pp_first}:AD{pp_last}"
h7_rng = f"{PP}!AI{pp_first}:AI{pp_last}"
h9_rng = f"{PP}!AK{pp_first}:AK{pp_last}"

bundle_long_names = [b[1] for b in BUNDLES]

for bundle_idx in range(1, 34):
    r = 1 + bundle_idx  # scorecard row 2 = bundle 1, etc.
    ba_col_letter = get_column_letter(3 + (bundle_idx - 1))
    ba_range = f"{BA}!{ba_col_letter}3:{ba_col_letter}{2 + n_pp_rows}"
    enabled_cell = b_enabled(bundle_idx)
    bname = bundle_long_names[bundle_idx - 1]

    # # (col 1) — always shown
    ws_sc.cell(row=r, column=1, value=bundle_idx).font = ARIAL_BOLD
    ws_sc.cell(row=r, column=1).alignment = CENTER
    ws_sc.cell(row=r, column=1).border = BORDER

    # Bundle name (col 2) — formula switches between name and [DISABLED] name
    name_formula = f'=IF({enabled_cell}="No","[DISABLED] {bname}","{bname}")'
    c = ws_sc.cell(row=r, column=2, value=name_formula)
    c.font = ARIAL
    c.alignment = LEFT
    c.border = BORDER

    # Helper to wrap any metric: if bundle disabled -> ""; else compute
    def wrap_metric(metric_expr):
        return f'=IF({enabled_cell}="No","",{metric_expr})'

    # Coverage depth
    ws_sc.cell(row=r, column=3,
               value=wrap_metric(f'COUNTIF({ba_range},TRUE)'))

    # Avg detectability
    ws_sc.cell(row=r, column=4,
        value=wrap_metric(
            f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),({d1}+{d2})/2)/COUNTIF({ba_range},TRUE),0)'))

    # Avg remaining term
    ws_sc.cell(row=r, column=5,
        value=wrap_metric(
            f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),{e4})/COUNTIF({ba_range},TRUE),0)'))

    # Trilateral %
    ws_sc.cell(row=r, column=6,
        value=wrap_metric(
            f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({f2}="Yes"))/COUNTIF({ba_range},TRUE),0)'))

    # Continuation %
    ws_sc.cell(row=r, column=7,
        value=wrap_metric(
            f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({e3}="Yes"))/COUNTIF({ba_range},TRUE),0)'))

    # SEP %
    ws_sc.cell(row=r, column=8,
        value=wrap_metric(
            f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({b1}>=2))/COUNTIF({ba_range},TRUE),0)'))

    # Pioneer count
    ws_sc.cell(row=r, column=9,
        value=wrap_metric(
            f'SUMPRODUCT(--({ba_range}=TRUE),--({c2}=3))'))

    # Strength flag — uses Configuration thresholds
    flag_expr = (f'IF(C{r}<2,"WEAK",IF(AND(C{r}>={thr("Strength_depth_min")},'
                 f'D{r}>={thr("Strength_detect_min")},E{r}>={thr("Strength_term_min")}),'
                 f'"STRONG","MODERATE"))')
    ws_sc.cell(row=r, column=10, value=wrap_metric(flag_expr))

    # Gate columns K..O
    # K = Weakest H1 -> gated by Gate_WeakestH1
    weakest_expr = f'IFERROR(AGGREGATE(5,6,{h1_rng}/({ba_range}=TRUE)),"")'
    k_formula = (f'=IF({enabled_cell}="No","",'
                 f'IF({gate_enabled("Gate_WeakestH1")}="No","",{weakest_expr}))')
    ws_sc.cell(row=r, column=11, value=k_formula)

    # L = Invalidity exposure % -> gated by Gate_InvalidityExposure
    inval_expr = (f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({h2_rng}<=1))/'
                  f'COUNTIF({ba_range},TRUE),0)')
    l_formula = (f'=IF({enabled_cell}="No","",'
                 f'IF({gate_enabled("Gate_InvalidityExposure")}="No","",{inval_expr}))')
    ws_sc.cell(row=r, column=12, value=l_formula)

    # M = EoU-ready %
    eou_expr = (f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),'
                f'--({h9_rng}<>"None"),--({h9_rng}<>""))/COUNTIF({ba_range},TRUE),0)')
    m_formula = (f'=IF({enabled_cell}="No","",'
                 f'IF({gate_enabled("Gate_EoUReady")}="No","",{eou_expr}))')
    ws_sc.cell(row=r, column=13, value=m_formula)

    # N = Survived %
    surv_expr = (f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({h7_rng}="Survived"))/'
                 f'COUNTIF({ba_range},TRUE),0)')
    n_formula = (f'=IF({enabled_cell}="No","",'
                 f'IF({gate_enabled("Gate_Survived")}="No","",{surv_expr}))')
    ws_sc.cell(row=r, column=14, value=n_formula)

    # O = Continuation-optionality %
    cont_expr = (f'IFERROR(SUMPRODUCT(--({ba_range}=TRUE),--({e3}="Yes"))/'
                 f'COUNTIF({ba_range},TRUE),0)')
    o_formula = (f'=IF({enabled_cell}="No","",'
                 f'IF({gate_enabled("Gate_ContOptionality")}="No","",{cont_expr}))')
    ws_sc.cell(row=r, column=15, value=o_formula)

    # Formatting for all metric cells
    for col in range(3, 16):
        cell = ws_sc.cell(row=r, column=col)
        cell.border = BORDER
        cell.font = ARIAL
        cell.alignment = CENTER

    ws_sc.cell(row=r, column=4).number_format = '0.00'
    ws_sc.cell(row=r, column=5).number_format = '0.0'
    for col in (6, 7, 8):
        ws_sc.cell(row=r, column=col).number_format = '0.0%'
    ws_sc.cell(row=r, column=11).number_format = '0'
    for col in (12, 13, 14, 15):
        ws_sc.cell(row=r, column=col).number_format = '0.0%'

# Conditional formatting on Scorecard:
# 1) Gray-out rows where bundle name (col B) starts with [DISABLED]
for r in range(2, 35):
    rng = f'A{r}:O{r}'
    ws_sc.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("[DISABLED]",$B{r}))'],
                    fill=DISABLED_FILL, font=DISABLED_FONT, stopIfTrue=True))

# 2) Gray-out gate columns when gate is OFF
# Header cells K1..O1 should show DISABLED marker too. Replace the header text with a formula.
gate_columns = {
    11: ('Weakest H1 (min in bundle)', 'Gate_WeakestH1'),
    12: ('Invalidity exposure % (H2≤1)', 'Gate_InvalidityExposure'),
    13: ('EoU-ready % (H9≠None)', 'Gate_EoUReady'),
    14: ('Survived % (H7=Survived)', 'Gate_Survived'),
    15: ('Continuation-optionality %', 'Gate_ContOptionality'),
}
for col, (label, gkey) in gate_columns.items():
    header_formula = f'=IF({gate_enabled(gkey)}="No","[DISABLED] {label}","{label}")'
    c = ws_sc.cell(row=1, column=col, value=header_formula)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER
    # Gray-out rule for the column when gate is off
    col_letter = get_column_letter(col)
    rng = f'{col_letter}2:{col_letter}34'
    ws_sc.conditional_formatting.add(rng,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("[DISABLED]",${col_letter}$1))'],
                    fill=DISABLED_FILL, font=DISABLED_FONT, stopIfTrue=True))

# 3) Strength flag conditional fills (re-applied so they still work for enabled rows)
strong_fill = PatternFill('solid', start_color='C6EFCE')
mod_fill = PatternFill('solid', start_color='FFEB9C')
weak_fill = PatternFill('solid', start_color='FFC7CE')
ws_sc.conditional_formatting.add('J2:J34',
    CellIsRule(operator='equal', formula=['"STRONG"'], fill=strong_fill))
ws_sc.conditional_formatting.add('J2:J34',
    CellIsRule(operator='equal', formula=['"MODERATE"'], fill=mod_fill))
ws_sc.conditional_formatting.add('J2:J34',
    CellIsRule(operator='equal', formula=['"WEAK"'], fill=weak_fill))

# ============================================================
# UPDATE README to mention v3
# ============================================================
ws_readme = wb['README']
existing_max = ws_readme.max_row
addendum_v3 = [
    ('', ''),
    ('v3 ADDITIONS — Pattern A Configurability', 'h2'),
    ('Configuration sheet — dropdown at the top picks a preset (Active Preset cell); Edit Mode toggle lets you override individual values from the Manual Override column. All routing rules and Scorecard formulas read from this sheet.', 'p'),
    ('Presets sheet — 6 starter presets (All ON, NPE / Counter-Assertion, Operating Company / FTO, Defensive Aggregator, Standards Licensee, EV Powertrain Sale) plus 4 empty "Custom Preset" slots you can fill in and rename.', 'p'),
    ('Disabled bundles — Bundle Assignment columns and Scorecard rows show "[DISABLED]" in the header/name and are visibly grayed out. Cells return blanks instead of TRUE/FALSE so they don\'t inflate the per-patent Total. Re-enable by setting the toggle back to Yes.', 'p'),
    ('Disabled gates — Scorecard gate columns work the same way: header gets [DISABLED] marker, cells go blank, column is grayed.', 'p'),
    ('Configurable thresholds — SEP cutoff, detectability cutoffs, family size minimum, citation threshold, pre-expiry window, salvage cutoffs, and STRONG-flag thresholds are all on the Configuration sheet and can be tuned per analysis.', 'p'),
    ('', ''),
    ('HOW TO USE THE CONFIGURATION', 'h2'),
    ('1. Go to the Configuration sheet.', 'p'),
    ('2. Pick a preset from the "Active Preset" dropdown at the top (cell B4). Everything recomputes immediately.', 'p'),
    ('3. To tweak individual values: set "Edit Mode" (B5) to Yes. Now the Active Value column reads from the Manual Override column. Edit the override cells as needed.', 'p'),
    ('4. To save your custom configuration as a preset: with the values you want set on the Configuration sheet, go to the Presets sheet and paste them into a "Custom Preset" column. Rename the column header. Done.', 'p'),
    ('5. Set Edit Mode back to No and pick your new preset from the dropdown.', 'p'),
]
for offset, (text, kind) in enumerate(addendum_v3, 1):
    r = existing_max + offset
    c = ws_readme.cell(row=r, column=1, value=text)
    if kind == 'h2':
        c.font = Font(name='Arial', bold=True, size=12, color='1F4E78')
    elif kind == 'p':
        c.font = ARIAL
        c.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(OUT)
print(f'Saved: {OUT}')
