"""Patent Bundling Template v4 — add Attribute Procedure sheet.

Option B (per user choice):
- New "Attribute Procedure" sheet, sits between Attribute Dictionary and Patent Portfolio
- Top section: 10-step recommended workflow table
- Bottom section: full per-attribute master procedure table (one row per attribute, 7 columns)
- No hover comments on Patent Portfolio (per user choice)
- Existing Attribute Dictionary sheet left unchanged
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/home/claude/Patent_Bundling_Template_v3.xlsx'
OUT = '/home/claude/Patent_Bundling_Template_v4.xlsx'

wb = load_workbook(SRC)

# -----------------------------------------------------------
# Styles
# -----------------------------------------------------------
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUB_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E78')
SECTION_FILL = PatternFill('solid', start_color='DDEBF7')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', bold=True, size=10)
ARIAL_SMALL = Font(name='Arial', size=9)
ARIAL_ITALIC = Font(name='Arial', size=9, italic=True, color='595959')

# Group fills matching Patent Portfolio sheet
GROUP_FILLS = {
    'A': PatternFill('solid', start_color='2E75B6'),
    'B': PatternFill('solid', start_color='5B9BD5'),
    'C': PatternFill('solid', start_color='9DC3E6'),
    'D': PatternFill('solid', start_color='BDD7EE'),
    'E': PatternFill('solid', start_color='A9D08E'),
    'F': PatternFill('solid', start_color='C6E0B4'),
    'G': PatternFill('solid', start_color='E2EFDA'),
    'H': PatternFill('solid', start_color='F4B084'),
    'I': PatternFill('solid', start_color='FFD966'),
}
GROUP_TEXT_COLORS = {
    'A': 'FFFFFF', 'B': 'FFFFFF', 'C': '1A1A1A', 'D': '1A1A1A',
    'E': 'FFFFFF', 'F': '1A1A1A', 'G': '1A1A1A', 'H': 'FFFFFF', 'I': '1A1A1A',
}

TYPE_COLORS = {
    'Direct lookup':           '548235',
    'Direct lookup (count)':   '548235',
    'Direct lookup (list)':    '548235',
    'Direct lookup (boolean)': '548235',
    'Direct mapping (research)': 'C65911',
    'Computed':                '1F4E78',
    'Derived (rubric)':        'C65911',
    'Derived (rubric + SME)':  'C65911',
    'Derived (tag)':           'C65911',
    'Derived (boolean from F1)': '1F4E78',
    'Derived (rubric from F1)':  '1F4E78',
    'Derived (claim analysis)':  'C65911',
    'Derived (file-wrapper review)': 'C65911',
    'Derived (search-based)':    'C65911',
    'Derived (rubric + engineering SME)': 'C65911',
    'Derived (specification + history)':  'C65911',
    'Internal artifact check':   'C65911',
}

# ============================================================
# Data
# ============================================================
WORKFLOW = [
    (1,  'Pull bibliographic & family data',
     'For each patent: get the front page, claims, abstract, family tree, citations, assignments, and prosecution status. Use Google Patents + Espacenet + USPTO Patent Center.',
     'A1, A2, C1, C3, E1–E5, F1, H5, H6, H8'),
    (2,  'Read claims and specification',
     'Read the first independent claim of each patent + the Background / Field-of-Invention. This is the irreplaceable manual step.',
     'A3, A4, A5, C2, C4, B3'),
    (3,  'Check SEP declarations',
     'For wireless / video / connectivity patents: check ETSI IPR, IEEE-SA, ITU declarations against the family.',
     'B1, B2'),
    (4,  'Score detectability',
     'For each claim, judge external vs. teardown detectability with the rubric. Cross-check teardowns where available.',
     'D1, D2'),
    (5,  'Map to products',
     'Identify which commercial products read on each patent. Build a quick read-on list. This is the highest-leverage step for buyer value.',
     'D3, I1'),
    (6,  'Compute geography',
     'From the family list, derive trilateral flag and major-market score.',
     'F2, F3 (auto from F1)'),
    (7,  'Tag themes and generations',
     'Apply your convergence-theme dictionary and generation labels.',
     'G1, G2, G3'),
    (8,  'Quality & vulnerability review',
     'Patent-counsel-grade review: claim strength, prior-art exposure, prosecution estoppel, divided-infringement, litigation history. Most labor-intensive but most decisive for buyer due diligence.',
     'H1, H2, H3, H4, H7'),
    (9,  'EoU & encumbrance audit',
     'Check internal claim-chart library; check encumbrance records.',
     'H9, H10'),
    (10, 'Market signals review',
     'Implementation maturity (was it productized?), adjacent re-read potential, workaround complexity. SME-driven.',
     'I2, I3, I4'),
]

ATTRS = [
    # GROUP A
    ('A', 'A1', 'Primary technology domain', 'Direct lookup',
     'Front page of the patent: CPC main classification (most specific subgroup), or IPC. Also check the title and abstract for confirmation.',
     'Take the CPC main code and translate it to a human-readable domain tag using a fixed taxonomy (e.g., CPC H04W → "Wireless networking"; CPC H01M → "Battery cells"). Maintain a CPC→tag mapping spreadsheet so values stay consistent across the portfolio.',
     'Google Patents, USPTO PatentsView, Espacenet, EPO Open Patent Services, Lens.org. CPC scheme reference: cpc.uspto.gov'),
    ('A', 'A2', 'Secondary domains', 'Direct lookup',
     'CPC classifications other than the main one (front page lists multiple). Also pull keywords from the abstract and the first independent claim.',
     'List every additional CPC subgroup beyond A1. Optionally enrich with 2-3 keyword tags drawn from the abstract (e.g., "MIMO", "beamforming", "edge inference").',
     'Google Patents, Espacenet, Lens.org. For keyword extraction at scale: use a simple TF-IDF over the abstract against the rest of your portfolio.'),
    ('A', 'A3', 'Stack layer', 'Derived (rubric)',
     'Read the first independent claim and the "Field of the Invention" / "Background" section.',
     'Apply a 7-bucket rubric. Hardware: physical device, circuit, or material composition. Firmware: low-level code embedded in hardware. OS: kernel/driver level. Middleware: protocol stacks, libraries, frameworks. App: end-user application logic. Cloud: server-side service or backend. UI: visual interface or interaction pattern. Pick the dominant layer; if the claim mixes layers, pick where the novelty lives.',
     'Patent specification. No external database — judgment-based against the rubric.'),
    ('A', 'A4', 'Product subsystem', 'Derived (rubric)',
     'Read the first independent claim, the "Detailed Description" preamble, and any "Embodiments" section listing real-world products.',
     'Map the claim subject matter to a subsystem of an identifiable end product. Example: for an EV — "battery cell", "battery pack", "BMS", "traction inverter", "thermal mgmt", "charging port", "motor", "regen brake controller". Maintain a per-industry subsystem dictionary so tags are consistent.',
     'Patent specification + product architecture knowledge (engineering team / SME input).'),
    ('A', 'A5', 'Use-case', 'Derived (rubric)',
     '"Field of the Invention", "Background", and the "Detailed Description" — look for problem statements ("there is a need to...", "addresses the problem of...").',
     'Phrase the use-case as a customer-facing problem (e.g., "indoor positioning", "battery fast-charging", "fraud detection in payments", "low-latency video streaming"). Maintain a portfolio-wide use-case dictionary; reuse tags rather than coining new ones for each patent.',
     'Patent specification. Use natural-language extraction from the Background section. Cross-check against market reports if needed.'),
    # GROUP B
    ('B', 'B1', 'SEP potential', 'Derived (rubric)',
     'Check (a) SEP databases for any declaration on this family, (b) the claim language for standards-spec phrasing, (c) any prior 3GPP / IEEE / ITU letters of assurance.',
     'Score 0-3. 3: declared SEP in an SSO database, OR claim language directly mirrors a normative spec section. 2: claim reads on a specific clause of a known standard but no formal declaration. 1: standard-adjacent but normative essentiality is debatable. 0: no standard tie. For 5G/Wi-Fi/Bluetooth/USB-C/H.26x patents, always run a claim-vs-spec mapping before assigning ≥2.',
     'ETSI IPR Online Database (search.ipr.etsi.org), IEEE-SA, ITU, IETF declarations, Unified Patents SEP database, IPlytics, Amplified.ai. Pair with claim-charts where available.'),
    ('B', 'B2', 'Standard tagged', 'Direct lookup',
     'Same sources as B1. Read the actual SSO declaration if one exists.',
     'Record the exact standard identifier (e.g., "3GPP TS 38.211 (5G NR PHY)", "IEEE 802.11be (Wi-Fi 7)", "USB PD 3.1", "ITU-T H.266/VVC"). If multiple standards are implicated, list all. Blank if B1 = 0.',
     'ETSI IPR database (full declaration text usually names the spec). Otherwise infer from claim text + standards expert review.'),
    ('B', 'B3', 'Interface role', 'Derived (rubric)',
     'Read the first independent claim. Look for protocol/handshake/format/connector language.',
     'Score 0-3. 3: claim sits at a mandatory interface between two systems (e.g., wire protocol, connector pinout, RPC format). 2: claim covers a widely-used interoperability mechanism that is not technically mandatory but is de-facto standard. 1: claim involves an interface but the interface is internal to one product. 0: no interface dimension.',
     'Patent specification + SME interpretation. Cross-check whether the interface is named in any public spec.'),
    # GROUP C
    ('C', 'C1', 'Claim type', 'Direct lookup',
     'Read the preamble of each independent claim.',
     'Classify the dominant independent claim type. Apparatus: "an apparatus comprising…", "a device comprising…". Method: "a method comprising the steps of…". System: "a system for X comprising…". CRM: "a non-transitory computer-readable medium storing instructions that, when executed…". Design: design patents (D-numbered, drawings only). If mixed, pick the most enforceable type given the product target.',
     'Patent claims section (always front-of-patent in granted patents).'),
    ('C', 'C2', 'Claim breadth', 'Derived (rubric)',
     'Read each independent claim. Count words, count specific limitations, look for "wherein" clauses.',
     'Score 0-3. 3 (pioneer/foundational): few elements, abstract/general language, no narrowing wherein clauses, broad genus. 2 (broad): moderate element count, some structural narrowing. 1 (narrow): many specific element constraints, specific value ranges, named structural features. 0 (very narrow/picture claim): tied to one specific implementation. Rule of thumb: shorter independent claims with fewer limitations are usually broader.',
     'Patent claims + judgment. Tools like ClaimMaster, LexisNexis PatentAdvisor, or simple word-count heuristics can prefilter.'),
    ('C', 'C3', 'Independent claim count', 'Direct lookup (count)',
     'Claims section. Independent claims are those that do not refer to another claim ("a method comprising…" rather than "the method of claim 1, wherein…").',
     'Simple count. Look for claims whose first word is not "The" referring back to a prior claim.',
     'Patent claims section. Most patent APIs (PatentsView, EPO OPS) return claim text in structured form.'),
    ('C', 'C4', 'Design-around difficulty', 'Derived (rubric + SME)',
     'Read all independent claims. Look at the size of the genus claimed, the number of dependent claims covering variants, and the breadth of the disclosure.',
     'Score 0-3. 3 (hard): independent claim covers a broad genus AND dependent claims systematically cover variants AND the specification suggests further unclaimed variants enabling future continuations. 2: covers main routes but leaves specific alternatives. 1: claim is one of several obvious alternatives. 0: trivial workaround exists. Engineering review strongly recommended for any patent scoring ≥2.',
     'Patent claims + specification + engineering SME review. There is no automated way to do this well — always sanity-check with a domain expert.'),
    # GROUP D
    ('D', 'D1', 'External detectability', 'Derived (rubric)',
     'Read the claim and ask: can I confirm infringement by observing the product without opening it?',
     'Score 0-3. 3: visible in UI, marketed feature, public spec sheet, or measurable from outputs (e.g., a specific protocol behavior on the wire). 2: detectable from network traffic capture or external instrumentation. 1: detectable only with significant external testing. 0: requires teardown or internal access.',
     'Patent claims + product knowledge. Use the SSO spec or product manuals to confirm whether a feature is externally observable.'),
    ('D', 'D2', 'Teardown detectability', 'Derived (rubric)',
     'Read the claim and ask: if I tear down the product, can I see the claimed element?',
     'Score 0-3. 3: visible in standard teardown (circuit traces, chip die markings, mechanical components). 2: detectable with electrical probing or chip-level reverse engineering. 1: detectable only with deep RE (decapping, firmware dump). 0: process-internal — no teardown reveals it (e.g., a manufacturing process step). Apparatus and chip-layout claims usually score 2-3; pure-software-method claims usually 0-1.',
     'Patent claims + RE feasibility judgment. TechInsights and Chipworks reports can validate teardown observability for chips.'),
    ('D', 'D3', 'Reads-on-known-products', 'Direct mapping (research)',
     'Match claim elements against features of specific commercial products. Build a claim chart.',
     'Score 0-3 based on number of confirmed reads. 3: 4+ commercial products confirmed to infringe. 2: 2-3 confirmed. 1: 1 confirmed or strong suspicion. 0: no known target. Always document the products in a separate EoU/Claim-Chart artifact (see H9).',
     'Product teardowns (TechInsights, iFixit, Chipworks), product manuals, marketing collateral, prior litigation complaints, IPlytics product-mapping, expert SME review.'),
    # GROUP E
    ('E', 'E1', 'Family size', 'Direct lookup (count)',
     'Family view on Espacenet, Google Patents "Family" tab, or PatBase family report. Includes parent + continuations + divisionals + CIPs + foreign counterparts.',
     'Count all family members across jurisdictions. Use DOCDB simple family or INPADOC extended family — pick one definition and apply it consistently. INPADOC is broader.',
     'Espacenet (free), PatBase, Derwent Innovation, Google Patents (Family tab), Patsnap.'),
    ('E', 'E2', 'Prosecution status', 'Direct lookup',
     'USPTO PAIR for US, EPO Register for EP, equivalent national registers for other jurisdictions. Or aggregated view in Espacenet / Patsnap.',
     'If all family members are granted → "Granted". If at least one is still pending → "Pending" or "Mixed". If a continuation is pending alongside granted parents → "Mixed". Choose "Mixed" if both granted and pending exist.',
     'USPTO Patent Center (formerly PAIR), EPO Register, JPO J-PlatPat, CNIPA, KIPRIS.'),
    ('E', 'E3', 'Continuation available', 'Direct lookup (boolean)',
     'Check the family tree on USPTO Patent Center / Espacenet for any pending child application linked to this patent or its parent.',
     'Yes if at least one continuation, divisional, or CIP application is currently pending and has not been abandoned. No otherwise. Note: a recently filed continuation is the most valuable signal — flag patents with continuations filed within the last 12 months.',
     'USPTO Patent Center, EPO Register, family trees in Espacenet/Patsnap.'),
    ('E', 'E4', 'Remaining term (years)', 'Computed',
     'Earliest patent expiry date in the family. For US: priority date + 20 years, adjusted for PTA (Patent Term Adjustment) and PTE (Patent Term Extension).',
     'Remaining_term = (expiry_date − today) in years, rounded. Use the EARLIEST expiry in the family for conservative scoring. For US patents granted from applications filed after June 8, 1995: expiry = filing_date + 20 years + PTA. Cross-check PAIR for PTA disclaimers and terminal disclaimers.',
     'USPTO PAIR, EPO Register, INPADOC. Patsnap and Derwent auto-compute expiry with adjustments.'),
    ('E', 'E5', 'Maintenance status', 'Direct lookup',
     'USPTO maintenance fee status (free lookup), EPO renewals register, national equivalents.',
     'Active: all maintenance fees paid up to current period. At-risk: a maintenance fee window is open and unpaid. Lapsed: maintenance fee deadline missed and grace period expired. Track this PER family member, not just the lead patent — a lapsed counterpart in a key jurisdiction is critical.',
     'USPTO Maintenance Fee Status (patentcenter.uspto.gov), EPO renewals, national patent office registers.'),
    # GROUP F
    ('F', 'F1', 'Jurisdictions', 'Direct lookup (list)',
     'Family view on Espacenet / Patsnap. Lists all countries where a family member is granted (not just filed).',
     'List all jurisdictions where at least one family member is currently GRANTED and active (not lapsed). Use 2-letter country codes: US, EP, CN, JP, KR, IN, DE, GB, FR, BR, CA, AU, etc. Note: EP coverage is composite — list EP and also which national validations exist if known.',
     'Espacenet, Patsnap, Derwent, Google Patents Family tab.'),
    ('F', 'F2', 'Trilateral coverage flag', 'Derived (boolean from F1)',
     'Derived from F1.',
     'Yes if patent is granted in US AND EP AND at least one of {CN, JP, KR}. No otherwise. This is the standard "trilateral+" definition; adjust if your buyer pool cares about specific markets (e.g., add IN for South Asia-focused buyers).',
     'Computed from F1.'),
    ('F', 'F3', 'Major-market coverage score', 'Derived (rubric from F1)',
     'Derived from F1.',
     'Score 0-3 by counting major-market grants. 3: granted in 4+ of {US, EP, CN, JP, KR}. 2: granted in 3. 1: granted in 2. 0: granted in 0-1. Adjust major-market list based on buyer (e.g., add IN, BR for emerging-market focused buyers).',
     'Computed from F1.'),
    # GROUP G
    ('G', 'G1', 'Convergence theme', 'Derived (tag)',
     'Read the abstract, claim, and detailed description. Match against a maintained list of current convergence themes.',
     'Tag the patent with one or more themes from a fixed dictionary you maintain. Example dictionary: "AI+healthcare", "edge AI", "AR/VR", "AI+chip-design", "Robotics+CV", "Quantum", "Sustainability+materials", "AI+industrial", "Web3", "Spatial computing". Keep the dictionary stable — coin a new tag only when several patents converge on a new theme.',
     'Patent specification + your maintained convergence-theme dictionary. Cross-check with market intelligence sources (Gartner Hype Cycles, McKinsey emerging tech reports).'),
    ('G', 'G2', 'Generation', 'Derived (rubric)',
     'Read the abstract and detailed description; check whether the patent maps to a specific technology generation (e.g., 4G vs 5G vs 6G, Wi-Fi 5 vs Wi-Fi 7, USB 2.0 vs USB 3.x vs USB4, 28nm vs 5nm).',
     'Legacy: tied to a superseded generation still in service but on decline. Current: tied to the dominant deployed generation. Next-gen: tied to the emerging or pre-deployment generation. For non-generation-tagged tech (materials, basic algorithms), leave blank.',
     'Patent specification + standards roadmap knowledge. Use generation transition timelines from analyst reports (Counterpoint, Omdia, IDC).'),
    ('G', 'G3', 'Cross-industry applicability', 'Derived (rubric)',
     'Read the independent claim. Ask: in how many distinct industries could this claim plausibly read on a real product?',
     'Score 0-3 by count of plausible target industries. 0: single industry only. 1: 2 industries. 2: 3 industries. 3: 4+ industries. Use a fixed industry list (Consumer Electronics, Automotive, Healthcare, Industrial/Manufacturing, Telecom, Energy, Aerospace, FinTech, AgTech, Defense). Cross-check by searching forward citations: if citing patents come from diverse CPC areas, the patent likely has cross-industry pull.',
     'Patent claims + forward-citation CPC distribution + SME review.'),
    # GROUP H
    ('H', 'H1', 'Claim strength rating', 'Derived (rubric + SME)',
     'Read all independent claims for language clarity; look at dependent claim depth; check the specification for support.',
     'Score 0-3. 3: claim language clean and definite, key terms have clear antecedent basis, dependent claims provide layered fallbacks, specification fully supports. 2: minor ambiguity but enforceable. 1: vague terms, weak antecedent basis, thin dependent layer. 0: ambiguous, indefinite, or claims poorly supported by specification. Engage patent counsel for any patent scoring ≥2 if it is going into a Premium tier bundle.',
     'Patent claims + specification + patent counsel review. Tools: ClaimMaster, LexisNexis PatentAdvisor for prefilter; final score requires legal review.'),
    ('H', 'H2', 'Prior-art exposure', 'Derived (search-based)',
     'Run a targeted prior-art search on the independent claim using key claim terms + publication date constraints.',
     'Score 0-3. 3: clean — focused PA search returns no closely matching references predating priority date. 2: minor prior art found but distinguishable on at least one independent-claim element. 1: significant prior art on most elements; novelty under threat. 0: prior art reads on the claim — high invalidity risk. Always check IPR/PTAB history (H7) for known challenges. For valuable patents, commission a formal invalidity search.',
     'Google Patents PA search, Patsnap, PatBase, LexisNexis IP Analytics, Innography. For high-stakes: commission a formal Patentability / Invalidity search from a search firm (Cardinal, Landon IP, Questel).'),
    ('H', 'H3', 'Prosecution history risk', 'Derived (file-wrapper review)',
     'Pull the full file wrapper from USPTO PAIR (US) or EPO Register. Read office actions and applicant responses.',
     'Score 0-3. 3: clean prosecution — allowance without significant claim narrowing, no §103 rejections argued around with narrowing amendments. 2: claims narrowed once to overcome prior art, but narrowing is peripheral. 1: significant narrowing arguments tied to claim terms still in the granted claims (heavy estoppel). 0: pattern of arguments creating broad estoppel; equivalents doctrine likely barred. This requires a patent attorney review.',
     'USPTO Patent Center file wrapper, EPO Register file inspection. Patent counsel review recommended.'),
    ('H', 'H4', 'Divided infringement risk', 'Derived (claim analysis)',
     'Read each independent method claim. Identify who performs each step.',
     'Score 0-3. 3: all claim steps performed by a single actor (apparatus claims usually score 3 by default). 2: claim steps performed by a single actor with minor user interaction. 1: claim steps require two parties (e.g., client + server with no clear "single mastermind"). 0: explicit multi-party performance with no joint enterprise. Method claims that span client + server + third-party services often score 0-1. Post-Akamai v. Limelight US law allows joint-infringement liability in narrow circumstances — patent counsel review for any 0-1 score.',
     'Patent claims + patent counsel analysis. Federal Circuit case law on joint infringement evolves; require legal review.'),
    ('H', 'H5', 'Forward citation count', 'Direct lookup (count)',
     'Google Patents "Cited by" count, USPTO PatentsView API, Lens.org, PatBase.',
     'Pull total forward citations across all jurisdictions for the family. Normalize for age: a 3-year-old patent with 30 citations is more influential than a 15-year-old patent with 50. Track both raw count and normalized "citations per year since publication".',
     'Google Patents (free), PatentsView API (free), Lens.org (free), Patsnap, Derwent.'),
    ('H', 'H6', 'Backward citation density', 'Direct lookup (count)',
     'Front page of the patent: "References Cited" — both applicant-cited and examiner-cited.',
     'Count total cited references on the front page. Higher count usually indicates a crowded art area (potential novelty pressure). Useful as a sanity-check against H2 — a patent with 50+ backward citations in a narrow domain has elevated prior-art exposure.',
     'Patent front page, Google Patents, PatentsView.'),
    ('H', 'H7', 'Litigation/PTAB history', 'Direct lookup',
     'Lex Machina (US litigation), Darts-IP (global litigation), USPTO PTAB Trials database, RPX, Unified Patents.',
     'None: no litigation or post-grant challenge. Survived: at least one IPR/PGR/EPO opposition resulted in final written decision with claims maintained. Pending: an active IPR/litigation. Lost-claims: at least one independent claim was invalidated or canceled by a final decision. Always check the actual outcome status — "instituted" is not "lost".',
     'Lex Machina (paid), Darts-IP (paid), USPTO PTAB (free at uspto.gov/ptab), RPX Insight, Unified Patents Portal, PACER for district court records.'),
    ('H', 'H8', 'Chain-of-title cleanliness', 'Direct lookup',
     'USPTO Assignment Search (free), EPO Register assignments, national patent office assignment records. Pull the full assignment chain from inventor(s) to current owner.',
     'Clean: complete assignment chain from all inventors to current owner, all properly recorded, no gaps. Minor gaps: chain is reconstructible but missing recorded assignments (e.g., an employment-agreement assignment that was never recorded). Encumbered: prior co-owners not joined, unresolved inventor disputes, prior security interests not released. Engage IP counsel for any non-Clean rating before bundling.',
     'USPTO Assignment Search (assignment.uspto.gov), EPO Register, national assignment databases. Real estate/UCC searches for security interests.'),
    ('H', 'H9', 'EoU / claim-chart availability', 'Internal artifact check',
     'Your internal claim-chart library / matter management. Check whether an Evidence-of-Use document exists mapping the claim to a named product or standard.',
     'None: no claim chart exists. Partial: chart maps the independent claim elements to a product but lacks secondary references / element-level evidence. Full: complete HD claim chart with every claim element tied to product documentation, source code (where available), or spec section. For SEP patents the EoU usually maps to a standard clause rather than a product.',
     'Internal claim-chart files. If none exist, commission claim charts for anchor patents before going to market — this is a major value driver.'),
    ('H', 'H10', 'Encumbrance status', 'Direct lookup',
     'Internal license register, prior settlement agreements, UCC filings, prior security agreements, RAND/FRAND commitments to SSOs.',
     'None: no licenses, liens, commitments, or carve-outs. Non-exclusive licensed: one or more non-exclusive licenses exist but rights to assert against unlicensed parties remain. Exclusive licensed: exclusive license exists — only the licensee can assert. Encumbered: liens, security interests, court orders, or covenants-not-to-sue limiting assertion. Always disclose to buyers; failure to disclose can void the sale.',
     'Internal license register, internal counsel review, UCC search at the secretary of state, IP collateral records, SSO declarations.'),
    # GROUP I
    ('I', 'I1', 'Product-mapping confidence', 'Derived (rubric)',
     'Same artifacts as D3 plus any internal claim charts. Ask: how confidently can we name commercial products that infringe?',
     'Score 0-3. 3: high — claim chart exists naming specific commercial products with element-level mapping. 2: moderate — known to map to product category but no specific SKU. 1: speculative — claim language suggests product fit but no confirmed mapping. 0: no current product fit identified.',
     'Internal claim-chart artifacts + product teardown research + SME review.'),
    ('I', 'I2', 'Implementation maturity', 'Derived (specification + history)',
     'Specification: does it describe working examples? File history: were prototype demos referenced? Inventor publications and product history.',
     'Idea-only: specification describes a concept with no working example or experimental data. Prototyped: specification or external sources reference a working prototype, lab demo, or research publication. Productized: a commercial product implementing the claim exists (either from the patentee or a third party). Cite the evidence in a note column.',
     'Patent specification, inventor publications (Google Scholar, IEEE Xplore), product announcements, press releases.'),
    ('I', 'I3', 'Adjacent-market re-read', 'Derived (rubric + SME)',
     'Read the independent claim; identify the named industry / use-case it was drafted for; ask whether the claim can be re-mapped to products in a different industry.',
     'Score 0-3. 3: claim language is industry-agnostic and clearly reads on 2+ adjacent industries beyond the patentee\'s. 2: re-read plausible in one adjacent industry with no claim contortion. 1: re-read possible but requires aggressive interpretation. 0: claim language ties it to one industry. Adjacent re-reads typically come from horizontal technologies (sensing, networking, AI methods).',
     'Patent claims + market knowledge + SME review across industries. Cross-check by examining whether forward citations come from different industry CPCs.'),
    ('I', 'I4', 'Workaround complexity', 'Derived (rubric + engineering SME)',
     'Read the claim. Ask an engineer: "If a competitor needed to ship a competing product without infringing, how much redesign is needed?"',
     'Score 0-3. 3: deep redesign — claim covers the natural/efficient approach; alternatives are expensive or technically inferior. 2: significant redesign — alternatives exist but are commercially undesirable. 1: minor redesign — straightforward alternative is available. 0: trivial workaround — designers can easily substitute. Strongly correlates with C4 design-around difficulty but considered separately because C4 is about claim language whereas I4 is about commercial impact.',
     'Patent claims + engineering SME review. The most reliable answer comes from an engineer who has actually shipped products in the space.'),
]

# ============================================================
# Build "Attribute Procedure" sheet
# ============================================================
# Insert it after Attribute Dictionary (which sits after Presets in v3).
# After insertion, move it to position immediately after Attribute Dictionary.
ws = wb.create_sheet('Attribute Procedure')
ws.sheet_view.showGridLines = False

# Move into the right position: after Attribute Dictionary, before Patent Portfolio
sheetnames = wb.sheetnames
# Find indices
ad_idx = sheetnames.index('Attribute Dictionary')
ap_idx = sheetnames.index('Attribute Procedure')
# Move Attribute Procedure to position ad_idx + 1
wb.move_sheet('Attribute Procedure', offset=(ad_idx + 1) - ap_idx)

# -----------------------------------------------------------
# Section 1: Title + intro
# -----------------------------------------------------------
ws.merge_cells('A1:G1')
c = ws.cell(row=1, column=1, value='ATTRIBUTE DERIVATION PROCEDURE')
c.font = Font(name='Arial', bold=True, size=18, color='1F4E78')
c.alignment = LEFT_CENTER

ws.merge_cells('A2:G2')
c = ws.cell(row=2, column=1,
            value='A field-by-field procedure for populating the 40-attribute scoring model from raw patent data. Sits alongside the Attribute Dictionary as the long-form reference.')
c.font = Font(name='Arial', italic=True, size=11, color='595959')
c.alignment = LEFT_CENTER

# Legend for the Type column
ws.merge_cells('A3:G3')
c = ws.cell(row=3, column=1, value='TYPE LEGEND')
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = LEFT_CENTER

LEGEND = [
    ('Direct lookup', 'Pull the value straight from a patent document field or database; no judgment required.'),
    ('Computed', 'Apply an arithmetic formula or boolean rule to other already-collected fields.'),
    ('Derived (rubric)', 'Apply a scoring rubric to claim text or specification language. Requires reading and judgment.'),
    ('Derived (rubric + SME)', 'Rubric-based but should be cross-checked with a domain expert or patent counsel.'),
]
for i, (label, desc) in enumerate(LEGEND):
    r = 4 + i
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(name='Arial', bold=True, size=10,
                  color=TYPE_COLORS.get(label, '595959'))
    c.alignment = LEFT_CENTER
    c.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    c = ws.cell(row=r, column=2, value=desc)
    c.font = ARIAL
    c.alignment = LEFT_CENTER
    c.border = BORDER

# -----------------------------------------------------------
# Section 2: 10-step workflow table
# -----------------------------------------------------------
wf_section_row = 4 + len(LEGEND) + 1
ws.merge_cells(start_row=wf_section_row, start_column=1, end_row=wf_section_row, end_column=7)
c = ws.cell(row=wf_section_row, column=1,
            value='RECOMMENDED WORKFLOW — 10 steps (run in order; steps 1-3 mostly automatable, steps 4-10 require reading & judgment)')
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = LEFT_CENTER

# Workflow table headers
wf_header_row = wf_section_row + 1
wf_headers = [('#', 1), ('Step', 1), ('What to do', 3), ('Fields populated', 2)]
col_cursor = 1
header_col_map = []
for label, span in wf_headers:
    end_col = col_cursor + span - 1
    if span > 1:
        ws.merge_cells(start_row=wf_header_row, start_column=col_cursor,
                       end_row=wf_header_row, end_column=end_col)
    c = ws.cell(row=wf_header_row, column=col_cursor, value=label)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER
    header_col_map.append((col_cursor, end_col))
    col_cursor = end_col + 1

# Workflow rows
for i, (n, step, what, fields) in enumerate(WORKFLOW):
    r = wf_header_row + 1 + i

    c = ws.cell(row=r, column=1, value=n)
    c.font = ARIAL_BOLD
    c.alignment = CENTER
    c.fill = PatternFill('solid', start_color='DDEBF7')
    c.border = BORDER

    c = ws.cell(row=r, column=2, value=step)
    c.font = ARIAL_BOLD
    c.alignment = LEFT
    c.border = BORDER

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c = ws.cell(row=r, column=3, value=what)
    c.font = ARIAL
    c.alignment = LEFT
    c.border = BORDER

    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    c = ws.cell(row=r, column=6, value=fields)
    c.font = Font(name='Arial', size=10, italic=True, color='1F4E78')
    c.alignment = LEFT
    c.border = BORDER

# Borders on merged cells
for i, (n, step, what, fields) in enumerate(WORKFLOW):
    r = wf_header_row + 1 + i
    for col in range(3, 8):
        ws.cell(row=r, column=col).border = BORDER

# -----------------------------------------------------------
# Section 3: Master per-attribute procedure table
# -----------------------------------------------------------
mt_section_row = wf_header_row + 1 + len(WORKFLOW) + 2  # leave a gap
ws.merge_cells(start_row=mt_section_row, start_column=1, end_row=mt_section_row, end_column=7)
c = ws.cell(row=mt_section_row, column=1,
            value='MASTER ATTRIBUTE PROCEDURE TABLE — one row per attribute (40 attributes total)')
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = LEFT_CENTER

# Table headers
mt_header_row = mt_section_row + 1
mt_headers = ['Grp', 'Code', 'Attribute', 'Type', 'Where to find', 'How to derive / compute', 'Best data source']
for i, h in enumerate(mt_headers, 1):
    c = ws.cell(row=mt_header_row, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# Attribute rows
for i, attr in enumerate(ATTRS):
    grp, code, name, typ, where, how, src = attr
    r = mt_header_row + 1 + i

    # Group letter (color-coded)
    c = ws.cell(row=r, column=1, value=grp)
    c.font = Font(name='Arial', bold=True, size=12,
                  color=GROUP_TEXT_COLORS.get(grp, 'FFFFFF'))
    c.fill = GROUP_FILLS.get(grp, HEADER_FILL)
    c.alignment = CENTER
    c.border = BORDER

    # Code
    c = ws.cell(row=r, column=2, value=code)
    c.font = Font(name='Arial', bold=True, size=10, color='1F4E78')
    c.alignment = CENTER
    c.border = BORDER

    # Attribute name
    c = ws.cell(row=r, column=3, value=name)
    c.font = ARIAL_BOLD
    c.alignment = LEFT
    c.border = BORDER

    # Type
    c = ws.cell(row=r, column=4, value=typ)
    c.font = Font(name='Arial', size=9, italic=True,
                  color=TYPE_COLORS.get(typ, '595959'))
    c.alignment = LEFT
    c.border = BORDER

    # Where to find
    c = ws.cell(row=r, column=5, value=where)
    c.font = ARIAL_SMALL
    c.alignment = LEFT
    c.border = BORDER

    # How to derive
    c = ws.cell(row=r, column=6, value=how)
    c.font = ARIAL_SMALL
    c.alignment = LEFT
    c.border = BORDER

    # Best data source
    c = ws.cell(row=r, column=7, value=src)
    c.font = ARIAL_ITALIC
    c.alignment = LEFT
    c.border = BORDER

# -----------------------------------------------------------
# Section 4: Data source reference (free vs paid)
# -----------------------------------------------------------
src_section_row = mt_header_row + 1 + len(ATTRS) + 2
ws.merge_cells(start_row=src_section_row, start_column=1, end_row=src_section_row, end_column=7)
c = ws.cell(row=src_section_row, column=1,
            value='DATA SOURCE REFERENCE — free and paid sources used across the procedure')
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = LEFT_CENTER

SOURCES = [
    ('Free', 'Google Patents', 'patents.google.com', 'Bibliographic, family, citations, claims, full text. Best free starting point.'),
    ('Free', 'Espacenet', 'worldwide.espacenet.com', 'EPO\'s global patent search; best for family view (INPADOC).'),
    ('Free', 'USPTO Patent Center', 'patentcenter.uspto.gov', 'US file wrapper, prosecution history, maintenance fees, assignments.'),
    ('Free', 'PatentsView', 'patentsview.org', 'USPTO bulk data API. Excellent for programmatic scoring across a portfolio.'),
    ('Free', 'Lens.org', 'lens.org', 'Free patent + scholarly search; good for forward-citation analysis.'),
    ('Free', 'EPO Register', 'register.epo.org', 'EP file wrapper, oppositions, renewals.'),
    ('Free', 'USPTO PTAB Database', 'developer.uspto.gov/ptab-api', 'IPR/PGR/CBM filings and outcomes.'),
    ('Free', 'ETSI IPR Database', 'ipr.etsi.org', 'SEP declarations for 3GPP / ETSI standards.'),
    ('Free', 'USPTO Assignment Search', 'assignment.uspto.gov', 'Chain of title for US patents.'),
    ('Paid', 'Patsnap', 'patsnap.com', 'Integrated platform: family, citations, litigation, valuation. Strong for portfolio-scale analysis.'),
    ('Paid', 'Derwent Innovation (Clarivate)', 'clarivate.com', 'Premium curated abstracts and family data; strong for prior art.'),
    ('Paid', 'PatBase', 'patbase.com', 'Family-centric search; gold standard for accurate family definitions.'),
    ('Paid', 'IPlytics', 'iplytics.com', 'SEP analytics, product-mapping intelligence.'),
    ('Paid', 'Lex Machina', 'lexmachina.com', 'US litigation analytics and outcomes.'),
    ('Paid', 'Darts-IP', 'darts-ip.com', 'Global patent litigation database.'),
    ('Paid', 'RPX Insight', 'rpxcorp.com', 'NPE litigation, market intelligence.'),
    ('Paid', 'Unified Patents Portal', 'unifiedpatents.com', 'PTAB activity, NPE assertion trends.'),
    ('Paid', 'TechInsights / Chipworks', 'techinsights.com', 'Authoritative chip teardowns — vital for D2/D3/I1.'),
    ('Paid', 'iFixit', 'ifixit.com', 'Consumer product teardowns (free + paid tiers).'),
]

# Source-table header
src_header_row = src_section_row + 1
src_headers = [('Tier', 1), ('Source', 1), ('URL', 2), ('What it gives you', 3)]
col_cursor = 1
for label, span in src_headers:
    end_col = col_cursor + span - 1
    if span > 1:
        ws.merge_cells(start_row=src_header_row, start_column=col_cursor,
                       end_row=src_header_row, end_column=end_col)
    c = ws.cell(row=src_header_row, column=col_cursor, value=label)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER
    col_cursor = end_col + 1

# Source rows
for i, (tier, src, url, what) in enumerate(SOURCES):
    r = src_header_row + 1 + i
    tier_color = '548235' if tier == 'Free' else 'C65911'

    c = ws.cell(row=r, column=1, value=tier)
    c.font = Font(name='Arial', bold=True, size=10, color=tier_color)
    c.alignment = CENTER
    c.border = BORDER

    c = ws.cell(row=r, column=2, value=src)
    c.font = ARIAL_BOLD
    c.alignment = LEFT_CENTER
    c.border = BORDER

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c = ws.cell(row=r, column=3, value=url)
    c.font = Font(name='Arial', size=9, color='595959')
    c.alignment = LEFT_CENTER
    c.border = BORDER

    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    c = ws.cell(row=r, column=5, value=what)
    c.font = ARIAL_SMALL
    c.alignment = LEFT
    c.border = BORDER

# Borders on merged sub-cells (URL and What columns)
for i in range(len(SOURCES)):
    r = src_header_row + 1 + i
    for col in (3, 4, 5, 6, 7):
        ws.cell(row=r, column=col).border = BORDER

# -----------------------------------------------------------
# Section 5: Practical notes
# -----------------------------------------------------------
notes_section_row = src_header_row + 1 + len(SOURCES) + 2
ws.merge_cells(start_row=notes_section_row, start_column=1, end_row=notes_section_row, end_column=7)
c = ws.cell(row=notes_section_row, column=1, value='PRACTICAL NOTES')
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = LEFT_CENTER

NOTES = [
    ('Scoring consistency',
     'For every 0-3 rubric attribute, calibrate by scoring 5-10 sample patents first as a team, then locking the calibration before processing the rest. Otherwise scores drift across reviewers. Maintain a portfolio-specific data dictionary for A1, A4, A5, G1 so tags are reused rather than re-coined for each patent. Inconsistent tags break the routing rules.'),
    ('Time and effort allocation',
     'On a 50-patent portfolio: Steps 1-3 (data pull) take ~1 hour total when automated. Steps 4-7 (claim reading, rubric scoring) take ~30 min per patent. Step 8 (quality and vulnerability) takes ~45 min per patent and benefits from patent-counsel involvement. Step 9 (EoU and encumbrance) varies wildly — anywhere from 5 min to several hours depending on internal artifact availability.'),
    ('Where to invest in claim charts',
     'Don\'t build EoU claim charts for the whole portfolio. Build them for: anchor patents (H1=3), patents likely to land in EoU-Backed (Bundle 26) or Battle-Tested (Bundle 27), and anything destined for the Premium tier. Charts are expensive — invest them where they shift bundle value the most.'),
    ('Automation pointers',
     'PatentsView API + Espacenet OPS API together cover ~80% of the Direct Lookup fields programmatically. A simple Python script can populate A1, A2, B2, C1, C3, E1-E5, F1, H5, H6, H8 across the entire Patent Portfolio sheet in one batch. Reserve manual work for the rubric-based and SME-based attributes.'),
    ('Threshold calibration warning',
     'The numeric thresholds on the Configuration sheet (SEP cutoff, citation minimum, etc.) only make sense if your 0-3 scores are calibrated. If two reviewers would give the same patent different H1 scores, the threshold-based routing produces noise. Run a calibration session before bulk scoring.'),
]

for i, (heading, body) in enumerate(NOTES):
    r = notes_section_row + 1 + (i * 2)

    # Heading
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(row=r, column=1, value=heading)
    c.font = Font(name='Arial', bold=True, size=11, color='1F4E78')
    c.alignment = LEFT_CENTER
    c.border = BORDER

    # Body
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)
    c = ws.cell(row=r + 1, column=1, value=body)
    c.font = ARIAL
    c.alignment = LEFT
    c.border = BORDER

# -----------------------------------------------------------
# Column widths and freeze
# -----------------------------------------------------------
ws.column_dimensions['A'].width = 5    # Grp
ws.column_dimensions['B'].width = 9    # Code / Step #
ws.column_dimensions['C'].width = 30   # Attribute name / Step
ws.column_dimensions['D'].width = 22   # Type
ws.column_dimensions['E'].width = 42   # Where to find
ws.column_dimensions['F'].width = 54   # How to derive
ws.column_dimensions['G'].width = 38   # Best data source

# Row heights — make rubric rows tall enough
# Title rows
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 36
# Master table rows — variable based on content length; default 95
for i in range(len(ATTRS)):
    r = mt_header_row + 1 + i
    body_len = max(len(ATTRS[i][4]), len(ATTRS[i][5]), len(ATTRS[i][6]))
    # Approx row height — 14 px per ~70 chars in a column of width 54
    ws.row_dimensions[r].height = max(70, min(220, body_len / 4))
# Workflow rows
for i in range(len(WORKFLOW)):
    r = wf_header_row + 1 + i
    ws.row_dimensions[r].height = 50
# Source rows
for i in range(len(SOURCES)):
    r = src_header_row + 1 + i
    ws.row_dimensions[r].height = 32
# Notes body rows
for i in range(len(NOTES)):
    r = notes_section_row + 1 + (i * 2) + 1  # body row
    ws.row_dimensions[r].height = 60

# Freeze panes at the master table header so it stays visible when scrolling
ws.freeze_panes = f'A{mt_header_row + 1}'

# ============================================================
# Update README to mention v4 addition
# ============================================================
ws_readme = wb['README']
existing_max = ws_readme.max_row
addendum_v4 = [
    ('', ''),
    ('v4 ADDITION — Attribute Derivation Procedure', 'h2'),
    ('New "Attribute Procedure" sheet sits between Attribute Dictionary and Patent Portfolio. It contains the field-by-field procedure for populating the 40 attributes from raw patent data — what to look up, where to look it up, and how to score the rubric-based fields.', 'p'),
    ('The sheet has four sections: (1) Type legend, (2) Recommended 10-step workflow, (3) Master per-attribute procedure table (one row per attribute with where to find / how to derive / data source), (4) Free vs Paid data source reference, (5) Practical notes on calibration and effort allocation.', 'p'),
    ('Use this sheet as the long-form reference. Keep using the Attribute Dictionary for quick scale/values lookups while entering data on the Patent Portfolio sheet.', 'p'),
]
for offset, (text, kind) in enumerate(addendum_v4, 1):
    r = existing_max + offset
    c = ws_readme.cell(row=r, column=1, value=text)
    if kind == 'h2':
        c.font = Font(name='Arial', bold=True, size=12, color='1F4E78')
    elif kind == 'p':
        c.font = ARIAL
        c.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheet order: {wb.sheetnames}')
