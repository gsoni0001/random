"""
build_v5.py — Extends Patent_Bundling_Template_v4.xlsx to v5
Adds:
  1. Three new lifecycle presets in the Presets sheet (cols 13-15)
     - "Short Lifecycle — Critical (1-2yr)"
     - "Short Lifecycle — Monetization (3-5yr)"
     - "Short Lifecycle — Strategic (6-8yr)"
  2. 30 new sample patents (rows 15-44 in Patent Portfolio)
     10 per band, spanning 5 tech domains, with realistic attributes
  3. Portfolio Context Guide sheet — reference sheet explaining each band,
     buyer fit, signals, and how to use the presets
  4. Updates Configuration dropdown to include the 3 new presets
  5. Updates README with v5 notes
"""

from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill, PatternFill)
from openpyxl.styles.fills import PatternFill as PF
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import copy

# ─── Load workbook ─────────────────────────────────────────────────────────
wb = load_workbook('Patent_Bundling_Template_v4.xlsx')

# ─── Style helpers ──────────────────────────────────────────────────────────
def font(bold=False, color='000000', size=10, name='Arial'):
    return Font(name=name, bold=bold, color=color, size=size)

def fill(hex_color):
    return PF('solid', start_color=hex_color, end_color=hex_color)

def border_thin():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style='medium', color='2E75B6')
    return Border(left=s, right=s, top=s, bottom=s)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

CENTER = align('center')
LEFT   = align('left')
WRAP   = align('left', wrap=True)

# Color palette matching v4
NAVY   = '1F3864'
BLUE   = '2E75B6'
LBLUE  = 'D6E4F0'
DBLUE  = '17375E'
WHITE  = 'FFFFFF'
LGRAY  = 'F5F5F5'
DGRAY  = '595959'
GREEN  = '375623'
LGREEN = 'E2EFD8'
ORANGE = 'C55A11'
LORANG = 'FCE4D6'
RED    = '7B0000'
LRED   = 'FCE4D6'
GOLD   = 'C09000'
LGOLD  = 'FFF2CC'
TEAL   = '005B7F'
LTEAL  = 'DDEEF5'

HDR_FONT  = Font(name='Arial', bold=True, color=WHITE, size=10)
HDR_FILL  = PF('solid', start_color=NAVY, end_color=NAVY)
BAND_A_FILL = PF('solid', start_color='FFF0F0', end_color='FFF0F0')  # critical - light red
BAND_B_FILL = PF('solid', start_color='FFF8EE', end_color='FFF8EE')  # window - light orange
BAND_C_FILL = PF('solid', start_color='F0F8F0', end_color='F0F8F0')  # strategic - light green

THIN = Side(style='thin', color='CCCCCC')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── 1. NEW PRESETS (columns 13, 14, 15) ────────────────────────────────────
ws_p = wb['Presets']

PRESET_CRITICAL = {
    'name': 'Short Lifecycle — Critical (1-2yr)',
    'description': 'Last-window enforcement: EoU-backed, battle-tested, detectable. Pre-Expiry window set 1-2yr. Disables continuation, prosecution-status, and long-horizon bundles.',
    'bundles': {
        1:'No', 2:'No', 3:'No', 4:'No', 5:'No', 6:'No', 7:'No', 8:'No', 9:'No', 10:'No',
        11:'Yes', 12:'Yes', 13:'Yes', 14:'No', 15:'No', 16:'No', 17:'No', 18:'No',
        19:'Yes', 20:'No', 21:'No', 22:'Yes', 23:'No', 24:'No', 25:'No',
        26:'Yes', 27:'Yes', 28:'Yes', 29:'No', 30:'No', 31:'Yes', 32:'Yes', 33:'No',
    },
    'thresholds': {
        'SEP_B1_cutoff': 2, 'Interface_B3_cutoff': 2,
        'Detect_D1_cutoff': 2, 'Detect_D2_cutoff': 2,
        'Family_E1_min': 2, 'CrossIndustry_G3_cutoff': 2,
        'Defensive_D3_cutoff': 2, 'Whitespace_C4_cutoff': 2,
        'Anchor_H1_cutoff': 3,       # Raise: only pristine anchors at 1-2yr
        'HighCitation_H5_min': 10,
        'PreExpiry_min_years': 1,    # Band floor
        'PreExpiry_max_years': 2,    # Band ceiling
        'Salvage_H1_max': 1, 'Salvage_E4_max': 2, 'Salvage_H2_max': 1,
        'Strength_depth_min': 3,     # Lower depth threshold for small bundles
        'Strength_detect_min': 2,
        'Strength_term_min': 1,      # Adjusted: min avg term is just 1yr for this band
    },
    'gates': {
        'Gate_WeakestH1': 'Yes',
        'Gate_InvalidityExposure': 'Yes',
        'Gate_EoUReady': 'Yes',
        'Gate_Survived': 'Yes',
        'Gate_ContOptionality': 'No',  # Not relevant at 1-2yr
    }
}

PRESET_WINDOW = {
    'name': 'Short Lifecycle — Monetization (3-5yr)',
    'description': 'Primary enforcement & licensing window. EoU-backed, detectable, anchor+halo core. Pre-Expiry 3-5yr. Adds claim-type and use-case bundles. Defensive aggregator secondary.',
    'bundles': {
        1:'No', 2:'No', 3:'No', 4:'No', 5:'Yes', 6:'No', 7:'No', 8:'No', 9:'No', 10:'No',
        11:'Yes', 12:'Yes', 13:'Yes', 14:'No', 15:'Yes', 16:'Yes', 17:'Yes', 18:'No',
        19:'Yes', 20:'No', 21:'No', 22:'Yes', 23:'No', 24:'Yes', 25:'No',
        26:'Yes', 27:'Yes', 28:'Yes', 29:'No', 30:'No', 31:'Yes', 32:'Yes', 33:'No',
    },
    'thresholds': {
        'SEP_B1_cutoff': 2, 'Interface_B3_cutoff': 2,
        'Detect_D1_cutoff': 2, 'Detect_D2_cutoff': 2,
        'Family_E1_min': 2, 'CrossIndustry_G3_cutoff': 2,
        'Defensive_D3_cutoff': 2, 'Whitespace_C4_cutoff': 2,
        'Anchor_H1_cutoff': 2,
        'HighCitation_H5_min': 10,
        'PreExpiry_min_years': 3,    # Band floor
        'PreExpiry_max_years': 5,    # Band ceiling
        'Salvage_H1_max': 1, 'Salvage_E4_max': 3, 'Salvage_H2_max': 1,
        'Strength_depth_min': 3,
        'Strength_detect_min': 2,
        'Strength_term_min': 3,      # Adjusted for 3-5yr band
    },
    'gates': {
        'Gate_WeakestH1': 'Yes',
        'Gate_InvalidityExposure': 'Yes',
        'Gate_EoUReady': 'Yes',
        'Gate_Survived': 'Yes',
        'Gate_ContOptionality': 'No',
    }
}

PRESET_STRATEGIC = {
    'name': 'Short Lifecycle — Strategic (6-8yr)',
    'description': 'Broadest buyer pool for short-lifecycle portfolios. OC-DEF viable. Adds tech-domain, defensive, cross-industry, provenance bundles. Pre-Expiry 6-8yr. Strength_term_min reduced to 6.',
    'bundles': {
        1:'Yes', 2:'No', 3:'Yes', 4:'No', 5:'Yes', 6:'No', 7:'No', 8:'No', 9:'No', 10:'No',
        11:'Yes', 12:'Yes', 13:'Yes', 14:'No', 15:'Yes', 16:'Yes', 17:'Yes', 18:'No',
        19:'Yes', 20:'Yes', 21:'No', 22:'Yes', 23:'No', 24:'Yes', 25:'No',
        26:'Yes', 27:'Yes', 28:'Yes', 29:'Yes', 30:'Yes', 31:'No', 32:'Yes', 33:'Yes',
    },
    'thresholds': {
        'SEP_B1_cutoff': 2, 'Interface_B3_cutoff': 2,
        'Detect_D1_cutoff': 2, 'Detect_D2_cutoff': 2,
        'Family_E1_min': 2, 'CrossIndustry_G3_cutoff': 2,
        'Defensive_D3_cutoff': 2, 'Whitespace_C4_cutoff': 2,
        'Anchor_H1_cutoff': 2,
        'HighCitation_H5_min': 10,
        'PreExpiry_min_years': 6,    # Band floor
        'PreExpiry_max_years': 8,    # Band ceiling
        'Salvage_H1_max': 1, 'Salvage_E4_max': 5, 'Salvage_H2_max': 1,
        'Strength_depth_min': 4,
        'Strength_detect_min': 2,
        'Strength_term_min': 6,      # Adjusted: minimum viable avg term for STRONG flag
    },
    'gates': {
        'Gate_WeakestH1': 'Yes',
        'Gate_InvalidityExposure': 'Yes',
        'Gate_EoUReady': 'Yes',
        'Gate_Survived': 'Yes',
        'Gate_ContOptionality': 'Yes',
    }
}

NEW_PRESETS = [PRESET_CRITICAL, PRESET_WINDOW, PRESET_STRATEGIC]
NEW_PRESET_COLS = [13, 14, 15]

# Row mapping in Presets sheet
BUNDLE_ROW_START = 4   # Row 4 = B1_enabled
THRESH_ROW_START = 38  # Row 38 = SEP_B1_cutoff
GATE_ROW_START   = 57  # Row 57 = Gate_WeakestH1

THRESH_KEYS = [
    'SEP_B1_cutoff','Interface_B3_cutoff','Detect_D1_cutoff','Detect_D2_cutoff',
    'Family_E1_min','CrossIndustry_G3_cutoff','Defensive_D3_cutoff','Whitespace_C4_cutoff',
    'Anchor_H1_cutoff','HighCitation_H5_min','PreExpiry_min_years','PreExpiry_max_years',
    'Salvage_H1_max','Salvage_E4_max','Salvage_H2_max',
    'Strength_depth_min','Strength_detect_min','Strength_term_min',
]
GATE_KEYS = ['Gate_WeakestH1','Gate_InvalidityExposure','Gate_EoUReady','Gate_Survived','Gate_ContOptionality']

# Preset header colors per band
PRESET_HDR_COLORS = ['C00000', 'C55A11', '375623']  # red, orange, green
PRESET_DESC_COLORS = ['FFF0F0', 'FFF8EE', 'F0F8F0']

for i, (preset, col, hdr_color, desc_color) in enumerate(zip(NEW_PRESETS, NEW_PRESET_COLS, PRESET_HDR_COLORS, PRESET_DESC_COLORS)):
    # Row 1: Preset name
    c = ws_p.cell(1, col, preset['name'])
    c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill = PF('solid', start_color=hdr_color, end_color=hdr_color)
    c.alignment = CENTER
    ws_p.column_dimensions[get_column_letter(col)].width = 22

    # Row 2: Description
    c = ws_p.cell(2, col, preset['description'])
    c.font = Font(name='Arial', size=9, color='333333')
    c.fill = PF('solid', start_color=desc_color, end_color=desc_color)
    c.alignment = align('left', wrap=True)

    # Row 3: Section header (bundle toggles)
    ws_p.cell(3, col, None)

    # Bundle toggles (rows 4-36)
    for bnum in range(1, 34):
        row = BUNDLE_ROW_START + bnum - 1
        val = preset['bundles'].get(bnum, 'Yes')
        c = ws_p.cell(row, col, val)
        c.font = Font(name='Arial', size=10,
                      color='1D6F42' if val == 'Yes' else '7B0000',
                      bold=True)
        c.alignment = CENTER

    # Row 37: Section header (thresholds)
    ws_p.cell(37, col, None)

    # Thresholds (rows 38-55)
    for j, key in enumerate(THRESH_KEYS):
        row = THRESH_ROW_START + j
        val = preset['thresholds'].get(key, 2)
        c = ws_p.cell(row, col, val)
        c.font = Font(name='Arial', size=10, color='000000')
        c.alignment = CENTER

    # Row 56: Section header (gates)
    ws_p.cell(56, col, None)

    # Gates (rows 57-61)
    for j, key in enumerate(GATE_KEYS):
        row = GATE_ROW_START + j
        val = preset['gates'].get(key, 'Yes')
        c = ws_p.cell(row, col, val)
        c.font = Font(name='Arial', size=10,
                      color='1D6F42' if val == 'Yes' else '7B0000',
                      bold=True)
        c.alignment = CENTER

# ─── Update Configuration sheet dropdown to include new presets ──────────────
ws_c = wb['Configuration']
# Update data validation for B4 to cover cols C through O (col 15)
for dv in ws_c.data_validations.dataValidation:
    if 'B4' in str(dv.sqref):
        dv.formula1 = "Presets!$C$1:$O$1"
        break

# ─── 2. NEW PATENTS (30 patents, 10 per band) ────────────────────────────────
# Format: (patent_id, title, A1, A2, A3, A4, A5, B1, B2, B3, C1, C2, C3, C4,
#           D1, D2, D3, E1, E2, E3, E4, E5, F1, F2, F3,
#           G1, G2, G3, H1, H2, H3, H4, H5, H6, H7, H8, H9, H10,
#           I1, I2, I3, I4)

PATENTS = [
    # ── BAND A: Critical Window (1-2 years) ────────────────────────────────
    # Strong EoU, survived challenges, tight term — NPE-LIT / LIT-FIN targets
    ('SL-001', 'LTE handover parameter optimization method',
     'Wireless PHY','Handover, mobility','Firmware','Baseband processor','Network handover',
     3,'3GPP LTE',3,'Method',2,3,2,
     3,2,3,4,'Granted','No',2,'Active','US, EP, CN','Yes',3,
     '','Legacy',1,3,3,3,3,31,18,'Survived','Clean','Full','None',
     3,'Productized',1,3),

    ('SL-002', 'Touch haptic feedback encoding for mobile UI',
     'UI/UX','Haptics, touch','App','Touch controller','Mobile UX',
     0,'',1,'Method',2,2,2,
     3,1,3,3,'Granted','No',2,'Active','US, EP','No',2,
     '','Current',2,2,2,3,2,12,9,'Survived','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-003', 'Low-power wake-word detection on microcontrollers',
     'Edge AI','TinyML, audio','Firmware','MCU core','Voice interface',
     0,'',0,'CRM',2,2,1,
     2,1,3,2,'Granted','No',1,'Active','US','No',1,
     'Edge AI','Current',2,2,2,2,2,8,14,'None','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-004', 'Fingerprint sensor anti-spoofing via impedance spectroscopy',
     'Biometrics','Sensor, security','Hardware','Fingerprint sensor','Identity auth',
     0,'',1,'Apparatus',3,2,3,
     3,2,3,3,'Granted','No',2,'Active','US, EP, CN','Yes',2,
     '','Current',3,3,3,3,3,22,11,'Survived','Clean','Full','None',
     3,'Productized',2,3),

    ('SL-005', 'DRAM refresh power optimization algorithm',
     'Memory','DRAM, power','Firmware','Memory controller','Low-power memory',
     0,'',0,'Method',2,3,2,
     2,3,3,4,'Granted','No',1,'Active','US, KR','No',2,
     '','Legacy',2,2,2,3,2,16,20,'Survived','Clean','Partial','None',
     3,'Productized',1,2),

    ('SL-006', 'Packet classification with ternary content-addressable memory',
     'Networking','TCAM, SDN','Hardware','Network ASIC','Traffic routing',
     0,'',2,'Apparatus',2,2,2,
     3,2,2,3,'Granted','No',2,'Active','US, EP','No',2,
     '','Current',2,2,3,2,3,14,17,'None','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-007', 'EV charging session authentication via PLC',
     'EV charging','PLC, V2G','Middleware','Charging port','EV charging auth',
     0,'',2,'Method',2,2,2,
     3,1,3,3,'Granted','No',2,'Active','US, EP, CN','Yes',2,
     'EV/Energy','Current',2,2,3,2,3,9,12,'None','Clean','Full','None',
     2,'Productized',2,2),

    ('SL-008', 'Predictive battery cell balancing using ML',
     'Battery mgmt','BMS, ML','App','Battery pack','Battery longevity',
     0,'',0,'Method',2,2,2,
     2,1,3,3,'Granted','No',1,'Active','US, EP','No',2,
     'Edge AI','Current',2,2,2,3,2,10,15,'Survived','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-009', 'Ultra-wideband indoor localization ranging method',
     'Wireless PHY','UWB, ranging','Hardware','RF module','Indoor positioning',
     2,'IEEE 802.15.4a',2,'Method',2,2,3,
     3,2,3,3,'Granted','No',2,'Active','US, EP','No',2,
     '','Current',3,2,2,2,3,11,8,'None','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-010', 'Adaptive video codec rate control for lossy networks',
     'Video compression','AVC, codec','App','Video pipeline','Streaming QoE',
     1,'H.264/AVC',2,'Method',3,3,3,
     3,0,2,3,'Granted','No',2,'Active','US, EP, CN','Yes',2,
     '','Legacy',2,3,3,3,3,18,21,'None','Minor gaps','Full','None',
     2,'Productized',2,3),

    # ── BAND B: Monetization Window (3-5 years) ────────────────────────────
    # Broader signals — NPE-LIC, NPE-LIT, LIT-FIN targets
    ('SL-011', '5G NR downlink HARQ-ACK timing control',
     'Wireless PHY','HARQ, 5G NR','Firmware','Baseband processor','Cellular comms',
     3,'3GPP 5G NR',3,'Method',2,3,2,
     2,3,3,5,'Granted','No',4,'Active','US, EP, CN, JP, KR','Yes',3,
     '','Current',1,2,2,3,2,14,16,'None','Clean','Partial','None',
     3,'Productized',1,2),

    ('SL-012', 'OLED pixel compensation circuit for burn-in prevention',
     'Display','OLED, pixel','Hardware','Display panel','Display quality',
     0,'',1,'Apparatus',2,2,3,
     2,2,3,4,'Granted','No',3,'Active','US, EP, KR','No',3,
     '','Current',2,2,2,2,2,16,11,'Survived','Clean','Partial','None',
     3,'Productized',1,2),

    ('SL-013', 'Object detection acceleration with sparse convolution',
     'CV algorithm','Sparse conv, ML','App','Inference engine','Autonomous perception',
     0,'',0,'CRM',2,3,2,
     2,1,2,4,'Granted','Yes',4,'Active','US, EP, CN','Yes',3,
     'Edge AI','Current',3,2,2,2,3,21,18,'None','Clean','Partial','None',
     2,'Prototyped',3,2),

    ('SL-014', 'Lithium-sulfur battery electrolyte stabilization',
     'Battery materials','Li-S, electrolyte','Hardware','Battery cell','Energy density',
     0,'',0,'Apparatus',3,3,3,
     0,2,2,3,'Granted','Yes',4,'Active','US, EP, CN, JP','Yes',3,
     'EV/Energy','Next-gen',3,3,3,3,3,24,29,'None','Clean','Partial','None',
     2,'Prototyped',3,3),

    ('SL-015', 'Private 5G network slice orchestration method',
     'Networking','Network slicing, 5G','Cloud','RAN controller','Private 5G',
     2,'3GPP 5G Core',2,'Method',2,2,2,
     2,0,2,3,'Granted','No',3,'Active','US, EP, CN','Yes',2,
     '','Current',2,2,2,3,2,12,14,'None','Clean','Partial','None',
     3,'Productized',2,2),

    ('SL-016', 'Secure enclave key derivation for IoT devices',
     'Security','HSM, TEE','Hardware','Secure element','Device security',
     0,'',1,'Apparatus',2,2,2,
     2,2,3,3,'Granted','No',3,'Active','US, EP','No',2,
     '','Current',2,3,3,3,2,9,13,'Survived','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-017', 'Multilayer ceramic capacitor dielectric formulation',
     'Electronic materials','MLCC, ceramics','Hardware','Passive component','Signal filtering',
     0,'',0,'Apparatus',3,2,3,
     0,3,2,5,'Granted','Yes',5,'Active','US, EP, JP, KR, CN','Yes',3,
     '','Current',3,2,2,3,2,19,38,'None','Clean','None','None',
     2,'Productized',2,3),

    ('SL-018', 'Cooperative MIMO precoding in distributed antenna systems',
     'Wireless PHY','C-RAN, MIMO','Hardware','Radio frontend','Cellular comms',
     2,'3GPP 5G NR',3,'Method',2,2,2,
     3,2,3,4,'Granted','No',5,'Active','US, EP, CN, KR','Yes',3,
     '','Current',1,3,3,3,3,13,19,'None','Clean','Partial','None',
     3,'Productized',1,2),

    ('SL-019', 'Real-time LiDAR point cloud compression for autonomous driving',
     'Sensors','LiDAR, point cloud','Middleware','Sensor fusion','Autonomous driving',
     0,'',0,'Method',2,2,2,
     2,2,3,3,'Granted','Yes',4,'Active','US, EP, CN','Yes',2,
     '','Current',3,2,3,2,3,17,14,'None','Clean','Partial','None',
     2,'Prototyped',3,2),

    ('SL-020', 'Thermal interface material with anisotropic conductivity',
     'Thermal mgmt','TIM, materials','Hardware','Thermal module','Electronics cooling',
     0,'',0,'Apparatus',2,2,3,
     0,3,2,3,'Granted','No',3,'Active','US, EP, CN','Yes',2,
     'EV/Energy','Current',2,3,3,3,3,11,22,'None','Clean','None','None',
     2,'Productized',3,3),

    # ── BAND C: Strategic Window (6-8 years) ───────────────────────────────
    # OC-DEF viable, NPE-LIC, broadest buyer pool
    ('SL-021', 'Wi-Fi 6E multi-BSS OFDMA resource scheduling',
     'Wireless PHY','OFDMA, Wi-Fi','Firmware','WLAN module','Indoor connectivity',
     2,'IEEE 802.11ax',3,'Method',2,3,2,
     2,2,3,5,'Granted','Yes',7,'Active','US, EP, CN, JP, KR','Yes',3,
     '','Current',2,2,2,3,2,11,14,'Survived','Clean','Partial','None',
     3,'Productized',2,2),

    ('SL-022', 'Solid-state LiDAR MEMS mirror resonance control',
     'Sensors','MEMS, LiDAR','Hardware','LiDAR scanner','Autonomous sensing',
     0,'',1,'Apparatus',3,2,3,
     2,2,3,4,'Granted','Yes',7,'Active','US, EP, CN','Yes',3,
     '','Current',3,3,3,3,3,28,17,'None','Clean','Partial','None',
     2,'Prototyped',3,3),

    ('SL-023', 'In-memory computing architecture for DNN inference',
     'Memory','CIM, neuromorphic','Hardware','Memory controller','AI acceleration',
     0,'',0,'Apparatus',3,3,2,
     1,2,2,4,'Granted','Yes',8,'Active','US, EP, CN','Yes',3,
     'Edge AI','Next-gen',3,3,2,3,3,33,22,'None','Clean','None','None',
     2,'Prototyped',3,2),

    ('SL-024', 'Perovskite solar cell stability enhancement method',
     'Photovoltaics','Perovskite, PV','Hardware','Solar cell','Energy generation',
     0,'',0,'Method',2,3,2,
     0,2,2,4,'Granted','Yes',7,'Active','US, EP','No',2,
     'EV/Energy','Next-gen',2,2,3,3,3,19,31,'None','Clean','None','None',
     1,'Prototyped',2,3),

    ('SL-025', 'Network traffic anomaly detection via federated learning',
     'Security','FL, anomaly','Cloud','Security platform','Threat detection',
     0,'',0,'Method',2,2,2,
     2,0,2,3,'Granted','No',6,'Active','US, EP','No',2,
     'Edge AI','Current',2,2,2,3,2,14,11,'None','Clean','None','None',
     2,'Productized',3,2),

    ('SL-026', 'Carbon nanotube field-effect transistor fabrication process',
     'Semiconductors','CNT, nano','Hardware','Logic gate','Post-silicon computing',
     0,'',0,'Method',3,2,3,
     0,2,1,3,'Granted','Yes',8,'Active','US, EP, JP','No',3,
     '','Next-gen',2,3,3,3,3,22,44,'None','Clean','None','None',
     1,'Prototyped',2,3),

    ('SL-027', 'Dynamic spectrum sharing between 5G NR and LTE coexistence',
     'Wireless PHY','DSS, coexistence','Firmware','Baseband processor','Spectrum efficiency',
     2,'3GPP Rel-16',3,'Method',2,2,2,
     3,2,3,5,'Granted','No',7,'Active','US, EP, CN, KR','Yes',3,
     '','Current',1,2,2,3,2,13,17,'None','Clean','Partial','None',
     3,'Productized',2,2),

    ('SL-028', 'Adaptive noise cancellation for over-ear headphones',
     'Audio','ANC, signal proc','Hardware','Audio SoC','Active noise cancellation',
     0,'',1,'System',2,2,2,
     3,1,3,4,'Granted','No',6,'Active','US, EP, CN','Yes',2,
     '','Current',3,2,2,3,2,16,12,'Survived','Clean','Full','None',
     3,'Productized',2,2),

    ('SL-029', 'Blockchain-based supply chain provenance verification',
     'Blockchain','DLT, provenance','Cloud','Smart contract','Supply chain trust',
     0,'',1,'Method',2,2,1,
     1,0,2,3,'Granted','No',7,'Active','US, EP','No',1,
     '','Current',2,1,1,3,1,6,9,'None','Clean','None','None',
     2,'Prototyped',3,1),

    ('SL-030', 'High-efficiency GaN power amplifier for 5G base stations',
     'Power electronics','GaN, RF power','Hardware','RF frontend','Cellular base station',
     1,'3GPP 5G NR',2,'Apparatus',2,2,2,
     2,2,3,4,'Granted','No',8,'Active','US, EP, CN, JP','Yes',3,
     '','Current',1,2,2,3,2,15,19,'None','Clean','Partial','None',
     3,'Productized',2,2),
]

# ─── Write patents to Patent Portfolio ──────────────────────────────────────
ws_pp = wb['Patent Portfolio']

# Band fill mapping by E4 value
def get_band_fill(e4):
    if e4 <= 2:
        return BAND_A_FILL
    elif e4 <= 5:
        return BAND_B_FILL
    else:
        return BAND_C_FILL

def get_band_font_color(e4):
    if e4 <= 2: return RED
    elif e4 <= 5: return ORANGE
    else: return GREEN

# Find next available row (after existing patents)
next_row = ws_pp.max_row + 1

# Write a band separator row before each group
band_info = [
    (PATENTS[0:10],  'BAND A — Critical Window (1–2 years remaining)  ·  Buyer targets: NPE-LIT, LIT-FIN  ·  Preset: Short Lifecycle — Critical (1-2yr)', 'C00000', 'FFE8E8'),
    (PATENTS[10:20], 'BAND B — Monetization Window (3–5 years remaining)  ·  Buyer targets: NPE-LIT, LIT-FIN, NPE-LIC  ·  Preset: Short Lifecycle — Monetization (3-5yr)', 'C55A11', 'FFF3E8'),
    (PATENTS[20:30], 'BAND C — Strategic Window (6–8 years remaining)  ·  Buyer targets: OC-DEF, NPE-LIC, IP-FUND  ·  Preset: Short Lifecycle — Strategic (6-8yr)', '375623', 'EEF5E8'),
]

for band_patents, band_label, band_color, band_bg in band_info:
    # Write separator row
    sep_row = next_row
    ws_pp.row_dimensions[sep_row].height = 18
    c = ws_pp.cell(sep_row, 1, band_label)
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PF('solid', start_color=band_color, end_color=band_color)
    c.alignment = align('left')
    ws_pp.merge_cells(start_row=sep_row, start_column=1,
                      end_row=sep_row, end_column=ws_pp.max_column)
    next_row += 1

    for pat in band_patents:
        row = next_row
        ws_pp.row_dimensions[row].height = 15
        for col, val in enumerate(pat, start=1):
            c = ws_pp.cell(row, col, val)
            e4 = pat[20]  # E4 is index 20 in tuple (0-based)
            c.fill = get_band_fill(e4)
            c.font = Font(name='Arial', size=9, color='1A1A1A')
            c.border = THIN_BORDER
            c.alignment = LEFT
            # Color the E4 column specially
            if col == 21:
                c.font = Font(name='Arial', size=9, bold=True,
                              color=get_band_font_color(e4))
        next_row += 1

# ─── Update Bundle Assignment formulas for new rows ─────────────────────────
ws_ba = wb['Bundle Assignment']

# Find the range of existing formula rows (rows 3-14 are the 12 original patents)
# We need to extend for all new patent rows
# First, get a sample formula from row 3 (first patent)
sample_formulas = {}
for col in range(3, ws_ba.max_column + 1):
    sample_formulas[col] = ws_ba.cell(3, col).value

# For each new patent row in Patent Portfolio, add corresponding BA row
# The new patent rows start at row next_row_start in PP
# We need to map Patent Portfolio rows → Bundle Assignment rows

# Patent Portfolio: original patents at rows 3-14; new ones start after separator rows
# Bundle Assignment: original at rows 3-14; we extend matching

# Build a mapping: PP row → BA row
pp_row_to_ba_row = {}
pp_row = 3
ba_row = 3
for i in range(12):  # Original 12
    pp_row_to_ba_row[pp_row] = ba_row
    pp_row += 1
    ba_row += 1

# Find where new patents actually landed in PP
pp_new_patent_rows = []
for row in range(15, ws_pp.max_row + 1):
    pid = ws_pp.cell(row, 1).value
    if pid and str(pid).startswith('SL-'):
        pp_new_patent_rows.append(row)

# Add Bundle Assignment rows for new patents
ba_next = ws_ba.max_row + 1

# We also need a separator row in BA aligned to each band separator
# Simple approach: write BA rows for each new PP patent row
for pp_row in pp_new_patent_rows:
    ba_row = ba_next
    # Col A: Patent ID (reference to PP)
    c = ws_ba.cell(ba_row, 1, f"='Patent Portfolio'!A{pp_row}")
    c.font = Font(name='Arial', size=9)
    # Col B: Title (reference to PP)
    c = ws_ba.cell(ba_row, 2, f"='Patent Portfolio'!B{pp_row}")
    c.font = Font(name='Arial', size=9)

    # Get the E4 for band fill
    e4_val = ws_pp.cell(pp_row, 21).value
    row_fill = get_band_fill(e4_val) if e4_val else PF('solid', start_color=WHITE, end_color=WHITE)

    ws_ba.cell(ba_row, 1).fill = row_fill
    ws_ba.cell(ba_row, 2).fill = row_fill

    # Cols 3-35: Bundle qualification formulas
    # Take sample formula from original row 3, replace row number
    for col in range(3, 36):
        sample = ws_ba.cell(3, col).value
        if sample and isinstance(sample, str) and 'Patent Portfolio' in sample:
            # Replace row 3 references with the new PP row
            new_formula = sample.replace("'Patent Portfolio'!A3", f"'Patent Portfolio'!A{pp_row}")
            new_formula = new_formula.replace("'Patent Portfolio'!B3", f"'Patent Portfolio'!B{pp_row}")
            # Replace all other PP row references
            import re
            def replace_pp_row(m):
                col_letter = m.group(1)
                return f"'Patent Portfolio'!{col_letter}{pp_row}"
            new_formula = re.sub(r"'Patent Portfolio'!([A-Z]+)3(?!\d)", replace_pp_row, new_formula)
            c = ws_ba.cell(ba_row, col, new_formula)
        elif sample:
            c = ws_ba.cell(ba_row, col, sample)
        else:
            c = ws_ba.cell(ba_row, col, '')
        c.fill = row_fill
        c.font = Font(name='Arial', size=9)
        c.border = THIN_BORDER
        c.alignment = CENTER

    # Col 36: Total formula
    total_col = 36
    c = ws_ba.cell(ba_row, total_col,
                   f"=COUNTIF(C{ba_row}:AH{ba_row},TRUE)")
    c.fill = row_fill
    c.font = Font(name='Arial', bold=True, size=9)
    c.border = THIN_BORDER
    c.alignment = CENTER

    ba_next += 1

# ─── 3. PORTFOLIO CONTEXT GUIDE SHEET ────────────────────────────────────────
if 'Portfolio Context Guide' in wb.sheetnames:
    del wb['Portfolio Context Guide']

ws_g = wb.create_sheet('Portfolio Context Guide', 2)  # Insert after README
ws_g.sheet_view.showGridLines = False
ws_g.sheet_view.showRowColHeaders = True

# Column widths
ws_g.column_dimensions['A'].width = 26
ws_g.column_dimensions['B'].width = 34
ws_g.column_dimensions['C'].width = 34
ws_g.column_dimensions['D'].width = 34
ws_g.column_dimensions['E'].width = 26

def guide_cell(ws, row, col, value, bold=False, fg='000000', bg=None,
               sz=10, wrap=True, align_h='left'):
    c = ws.cell(row, col, value)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    if bg:
        c.fill = PF('solid', start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal=align_h, vertical='top', wrap_text=wrap)
    c.border = THIN_BORDER
    return c

def section_banner(ws, row, text, color, text_color=WHITE, col_span=5):
    c = ws.cell(row, 1, text)
    c.font = Font(name='Arial', bold=True, color=text_color, size=12)
    c.fill = PF('solid', start_color=color, end_color=color)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 22

def sub_banner(ws, row, text, color, text_color=WHITE, col_span=5):
    c = ws.cell(row, 1, text)
    c.font = Font(name='Arial', bold=True, color=text_color, size=10)
    c.fill = PF('solid', start_color=color, end_color=color)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 18

row = 1

# ── Title banner ───────────────────────────────────────────────────────────
ws_g.row_dimensions[row].height = 30
c = ws_g.cell(row, 1, 'PORTFOLIO CONTEXT GUIDE  —  Short-Lifecycle Assets (≤8 Years Remaining Term)')
c.font = Font(name='Arial', bold=True, color=WHITE, size=14)
c.fill = PF('solid', start_color=NAVY, end_color=NAVY)
c.alignment = Alignment(horizontal='left', vertical='center')
ws_g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

ws_g.row_dimensions[row].height = 16
c = ws_g.cell(row, 1, 'v5 Addition  ·  Companion to Presets: Short Lifecycle — Critical (1-2yr) / Monetization (3-5yr) / Strategic (6-8yr)  ·  For outright patent sales')
c.font = Font(name='Arial', italic=True, color=DGRAY, size=9)
c.fill = PF('solid', start_color=LBLUE, end_color=LBLUE)
c.alignment = Alignment(horizontal='left', vertical='center')
ws_g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 2

# ── Overview table ─────────────────────────────────────────────────────────
section_banner(ws_g, row, '1.  TERM BAND OVERVIEW', NAVY)
row += 1

# Header row
for col, hdr in enumerate(['', 'Critical Window', 'Monetization Window', 'Strategic Window'], 1):
    if col == 1: continue
    c = guide_cell(ws_g, row, col, hdr, bold=True, fg=WHITE,
                   bg='C00000' if col==2 else 'C55A11' if col==3 else '375623',
                   align_h='center')
guide_cell(ws_g, row, 1, 'Dimension', bold=True, fg=WHITE, bg=NAVY, align_h='left')
row += 1

overview = [
    ('Remaining term',    '1 – 2 years',               '3 – 5 years',               '6 – 8 years'),
    ('Preset to use',     'Short Lifecycle — Critical', 'Short Lifecycle — Monetization', 'Short Lifecycle — Strategic'),
    ('Primary buyers',    'NPE-LIT, LIT-FIN',           'NPE-LIT, LIT-FIN, NPE-LIC', 'OC-DEF, NPE-LIC, IP-FUND'),
    ('Secondary buyers',  'DEF-AGG (weak assets)',      'DEF-AGG, OC-DEF (borderline)', 'DEF-AGG, STD-LIC'),
    ('Core value driver', 'EoU chart + named defendant', 'Enforcement window + broad claims', 'Coverage depth + product mapping'),
    ('Urgency framing',   '"Last enforcement window"',  '"Campaign horizon: 3–5yr"', '"Full product-gen coverage"'),
    ('Sample portfolios', 'SL-001 to SL-010',           'SL-011 to SL-020',          'SL-021 to SL-030'),
]

for dim, a, b, c_val in overview:
    bg = LGRAY if overview.index((dim,a,b,c_val)) % 2 == 0 else WHITE
    guide_cell(ws_g, row, 1, dim, bold=True, fg=NAVY, bg=bg)
    guide_cell(ws_g, row, 2, a, fg='7B0000', bg='FFF0F0' if bg==LGRAY else 'FFF8F8')
    guide_cell(ws_g, row, 3, b, fg=ORANGE, bg='FFF8EE' if bg==LGRAY else 'FFFAF5')
    guide_cell(ws_g, row, 4, c_val, fg=GREEN, bg='F0F8F0' if bg==LGRAY else 'F5FBF5')
    guide_cell(ws_g, row, 5, '', bg=bg)
    ws_g.row_dimensions[row].height = 28
    row += 1

row += 1

# ── How to use the presets ─────────────────────────────────────────────────
section_banner(ws_g, row, '2.  HOW TO USE THESE PRESETS WITH THE WORKBOOK', BLUE)
row += 1

steps = [
    ('Step 1', 'Go to the Configuration sheet. In cell B4 (Active Preset dropdown), select the preset that matches your portfolio term band.'),
    ('Step 2', 'The Bundle Assignment and Bundle Quality Scorecard sheets recompute automatically. Disabled bundle columns show [DISABLED] and are grayed out.'),
    ('Step 3', 'Review the Scorecard. For Critical and Monetization bands, look for bundles with EoU-ready % > 50% and Survived % > 0. For Strategic, focus on coverage depth and Weakest H1.'),
    ('Step 4', 'To fine-tune: set Edit Mode (B5) to Yes. You can then override individual thresholds in the Manual Override column without affecting the preset.'),
    ('Step 5', 'The 30 sample patents (SL-001 to SL-030) in the Patent Portfolio sheet are pre-scored for this scenario. Review them to understand how the routing logic differs across bands.'),
    ('Step 6', 'When pitching to buyers, match the IAM Market language register to the term band. Use the "Buyer Fit" and "Positioning Strategy" columns on this sheet as a reference.'),
]

for step, desc in steps:
    bg = LBLUE if steps.index((step, desc)) % 2 == 0 else WHITE
    c1 = guide_cell(ws_g, row, 1, step, bold=True, fg=NAVY, bg=bg)
    c2 = ws_g.cell(row, 2, desc)
    c2.font = Font(name='Arial', size=10)
    c2.fill = PF('solid', start_color=bg, end_color=bg)
    c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c2.border = THIN_BORDER
    ws_g.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws_g.row_dimensions[row].height = 32
    row += 1

row += 1

# ── Buy signals by band ────────────────────────────────────────────────────
section_banner(ws_g, row, '3.  CRITICAL SIGNALS BY TERM BAND  (Attributes driving sale success)', NAVY)
row += 1

# Header
for col, hdr in enumerate(['Signal / Attribute', 'Critical (1-2yr)', 'Monetization (3-5yr)', 'Strategic (6-8yr)', 'Notes'], 1):
    bg = 'C00000' if col==2 else 'C55A11' if col==3 else '375623' if col==4 else NAVY
    guide_cell(ws_g, row, col, hdr, bold=True, fg=WHITE, bg=bg, align_h='center')
row += 1

signals = [
    ('H9 EoU available',        'MUST HAVE — Full chart required', 'Must-have for NPE-LIT, nice-to-have for NPE-LIC', 'Nice-to-have; OC-DEF doesn\'t require it', 'Full chart = ready to file; Partial = 60-90d prep needed'),
    ('H7 Survived challenge',   'MUST HAVE — IPR filed day 1',     'Strong premium; validates claim integrity',          'Important for OC-OFF and IP-FUND',                    'Reduces invalidity risk for all buyers'),
    ('D1 External detectability','MUST HAVE — no discovery time',  'Must-have; no time for forensic analysis',           'Required for OC-DEF product mapping',                  'H1+D1 is the enforcement readiness pair'),
    ('H1 Claim strength',       'Min 3 — must survive IPR',        'Min 2 — IPR risk manageable',                       'Min 2 — OC buyers vet every claim',                    'Weakest H1 gate on Scorecard shows bundle floor'),
    ('H4 Divided infringement', '3 = single-actor (essential)',    '3 preferred; 2 acceptable with careful drafting',   '2+ acceptable; OC-DEF less sensitive',                'H4≤1 kills litigation route entirely'),
    ('E4 Remaining term',       '1-2yr: urgency is the asset',    '3-5yr: full campaign window',                       '6-8yr: product lifecycle coverage',                   'Pre-Expiry bundle threshold is set per band'),
    ('C2 Claim breadth',        '2+ helpful for damages scope',   '2+ for multi-target licensing pool',                '2+ for OC blocking and NPE-LIC programs',             'C2=3 (pioneer) is highest premium signal'),
    ('H8 Chain of title',       'Clean: fast close essential',    'Clean: standard requirement',                       'Clean: OC buyers are strictest on title',              'Any gap doubles due diligence timeline'),
    ('G3 Cross-industry',       'Marginal — no time for program', '2+ opens NPE-LIC secondary market',                '2+ opens IP-FUND and DEF-AGG buyers',                'B30 Adjacent Re-Read bundle leverages this'),
    ('H5 Forward citations',    'Low weight at this term',        'Moderate: helps IP-FUND if ≥10',                   'High weight: IP-FUND and CORP-VC key signal',          'B29 High-Citation bundle threshold = 10 (lowered from 15 in short-lifecycle presets)'),
]

for i, (sig, crit, wind, strat, note) in enumerate(signals):
    bg = LGRAY if i % 2 == 0 else WHITE
    r_fill = [('FFF0F0' if bg==LGRAY else 'FFF8F8'),
              ('FFF8EE' if bg==LGRAY else 'FFFAF5'),
              ('F0F8F0' if bg==LGRAY else 'F5FBF5')]
    guide_cell(ws_g, row, 1, sig, bold=True, fg=NAVY, bg=bg)
    guide_cell(ws_g, row, 2, crit, fg='7B0000', bg=r_fill[0])
    guide_cell(ws_g, row, 3, wind, fg=ORANGE, bg=r_fill[1])
    guide_cell(ws_g, row, 4, strat, fg=GREEN, bg=r_fill[2])
    guide_cell(ws_g, row, 5, note, fg=DGRAY, bg=bg, sz=9)
    ws_g.row_dimensions[row].height = 32
    row += 1

row += 1

# ── Deal killers ───────────────────────────────────────────────────────────
section_banner(ws_g, row, '4.  DEAL KILLERS — ROUTE TO SALVAGE OR REMOVE FROM SALE PACKAGE', RED)
row += 1

killers = [
    ('E4 < 1yr at close',       'Unsellable — case cannot be filed and resolved before expiry. Remove from package.',                                              'All bands'),
    ('H4 ≤ 1 (divided infringement)', 'Kills litigation route. OC-DEF still possible but at heavy discount. Salvage lot or remove.',                              'All bands'),
    ('H10 ≠ None (FRAND/encumbrance)', 'Restricts who can be targeted and caps royalty rates. Must disclose; kills many deals outright.',                         'All bands'),
    ('H2 = 0 (very high invalidity)',  'Critical and Monetization bands: IPR filed immediately, terminates campaign. Strategic: OC-DEF will pass.',               'Critical, Monetization'),
    ('H9 = None (no EoU)',             'Critical band: removes from primary sale tier, route to Salvage. Monetization: reduces price significantly.',             'Critical, Monetization'),
    ('H8 = Clouded title',             'Doubles due diligence timeline. OC buyers will walk. Must resolve pre-sale or heavily discount.',                          'All bands'),
    ('H7 = Pending (IPR active)',       'Outcome uncertainty is a deal-stopper for NPE-LIT and LIT-FIN. Wait for resolution or route to Salvage.',                'Critical, Monetization'),
    ('E3 = No AND H1 ≤ 1',             'No continuation option + weak claims = no future value. IP-FUND and OC-OFF will pass. Salvage or lot pricing only.',      'Strategic'),
]

guide_cell(ws_g, row, 1, 'Deal Killer (Attribute value)', bold=True, fg=WHITE, bg=RED, align_h='center')
guide_cell(ws_g, row, 2, 'Impact & recommended action', bold=True, fg=WHITE, bg=RED, align_h='center')
ws_g.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
guide_cell(ws_g, row, 5, 'Applies to', bold=True, fg=WHITE, bg=RED, align_h='center')
row += 1

for i, (dk, impact, applies) in enumerate(killers):
    bg = 'FCF0F0' if i % 2 == 0 else WHITE
    guide_cell(ws_g, row, 1, dk, bold=True, fg='7B0000', bg=bg)
    c = ws_g.cell(row, 2, impact)
    c.font = Font(name='Arial', size=10)
    c.fill = PF('solid', start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = THIN_BORDER
    ws_g.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    guide_cell(ws_g, row, 5, applies, fg=DGRAY, bg=bg, sz=9)
    ws_g.row_dimensions[row].height = 30
    row += 1

row += 1

# ── IAM Market language register ──────────────────────────────────────────
section_banner(ws_g, row, '5.  IAM MARKET LANGUAGE REGISTER BY TERM BAND', BLUE)
row += 1

guide_cell(ws_g, row, 1, 'Language element', bold=True, fg=WHITE, bg=NAVY)
guide_cell(ws_g, row, 2, 'Critical (1-2yr)', bold=True, fg=WHITE, bg='C00000', align_h='center')
guide_cell(ws_g, row, 3, 'Monetization (3-5yr)', bold=True, fg=WHITE, bg=ORANGE, align_h='center')
guide_cell(ws_g, row, 4, 'Strategic (6-8yr)', bold=True, fg=WHITE, bg=GREEN, align_h='center')
guide_cell(ws_g, row, 5, '', bold=True, fg=WHITE, bg=NAVY)
row += 1

lang_rows = [
    ('Term framing',     '"Enforcement window open through [date]"',           '"3–5yr campaign horizon; two full licensing rounds"', '"Seven years of enforceable coverage; full product gen"'),
    ('Value hook',       '"Case economics: damages exposure >> acquisition cost"', '"Assertion-ready; EoU charts prepared for [category]"', '"FTO depth across [tech domain]; clean title, immediate transfer"'),
    ('Urgency language', 'Use urgency explicitly — window is the asset',        'Frame as: "act before the window closes"',              'Avoid urgency; use coverage and blocking language'),
    ('Buyer language',   '"Assertion vehicle", "defendant-mapped", "case file"', '"Licensing program", "royalty base", "detectable infringement"', '"FTO coverage", "design-around blocking", "portfolio depth"'),
    ('Avoid',            '"Long-term strategic value" — overclaims',            '"Expiring soon" — sounds apologetic; flip to "last window"', '"Enforcement ready" without EoU support — needs evidence'),
]

for i, (elem, a, b, c_val) in enumerate(lang_rows):
    bg = LGRAY if i % 2 == 0 else WHITE
    guide_cell(ws_g, row, 1, elem, bold=True, fg=NAVY, bg=bg)
    guide_cell(ws_g, row, 2, a, fg='7B0000', bg='FFF0F0' if bg==LGRAY else 'FFF8F8', sz=9)
    guide_cell(ws_g, row, 3, b, fg=ORANGE, bg='FFF8EE' if bg==LGRAY else 'FFFAF5', sz=9)
    guide_cell(ws_g, row, 4, c_val, fg=GREEN, bg='F0F8F0' if bg==LGRAY else 'F5FBF5', sz=9)
    guide_cell(ws_g, row, 5, '', bg=bg)
    ws_g.row_dimensions[row].height = 36
    row += 1

row += 1

# ── 3-tier composition strategy ───────────────────────────────────────────
section_banner(ws_g, row, '6.  3-TIER COMPOSITION WITHIN ANY TERM BAND', NAVY)
row += 1

tiers = [
    ('Tier A  (Premium)', 'EoU-Backed + Survived-Challenge anchors', 'Lead with these. Price to reflect assertion readiness. NPE-LIT and LIT-FIN buyers will pay a premium. Bundle using B26 + B27 + B22.', BLUE, LBLUE),
    ('Tier B  (Standard)', 'Detectability ≥ 2, no full EoU chart yet', 'Position as "program-starter" lots: 60-90 days of claim charting needed before campaign launch. Discount from Tier A. Bundle using B12 + B22 + B28.', GOLD, LGOLD),
    ('Tier C  (Volume / Salvage)', 'Weak H1 or high invalidity exposure', 'Route to Salvage/Volume Lot bundle (B31). Price to clear. Target DEF-AGG buyers only. Never mix with Tier A or B in the same IAM listing.', RED, LRED),
]

guide_cell(ws_g, row, 1, 'Tier', bold=True, fg=WHITE, bg=NAVY, align_h='center')
guide_cell(ws_g, row, 2, 'Qualifying criteria', bold=True, fg=WHITE, bg=NAVY)
guide_cell(ws_g, row, 3, 'Positioning & recommended bundles', bold=True, fg=WHITE, bg=NAVY)
ws_g.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
row += 1

for tier_name, criteria, positioning, txt_color, bg_color in tiers:
    guide_cell(ws_g, row, 1, tier_name, bold=True, fg=txt_color, bg=bg_color, align_h='center')
    guide_cell(ws_g, row, 2, criteria, fg='222222', bg=bg_color)
    c = ws_g.cell(row, 3, positioning)
    c.font = Font(name='Arial', size=10)
    c.fill = PF('solid', start_color=bg_color, end_color=bg_color)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border = THIN_BORDER
    ws_g.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws_g.row_dimensions[row].height = 40
    row += 1

row += 1

# ── Sample patent index ────────────────────────────────────────────────────
section_banner(ws_g, row, '7.  SAMPLE PATENT INDEX  (SL-001 to SL-030 on Patent Portfolio sheet)', NAVY)
row += 1

guide_cell(ws_g, row, 1, 'Patent ID', bold=True, fg=WHITE, bg=NAVY, align_h='center')
guide_cell(ws_g, row, 2, 'Title', bold=True, fg=WHITE, bg=NAVY)
guide_cell(ws_g, row, 3, 'Term Band', bold=True, fg=WHITE, bg=NAVY, align_h='center')
guide_cell(ws_g, row, 4, 'Key buyer signal(s)', bold=True, fg=WHITE, bg=NAVY)
guide_cell(ws_g, row, 5, 'Tier', bold=True, fg=WHITE, bg=NAVY, align_h='center')
row += 1

index_data = [
    ('SL-001','LTE handover parameter optimization','Critical (1-2yr)','Full EoU + Survived IPR + H4=3','A'),
    ('SL-002','Touch haptic feedback encoding','Critical (1-2yr)','Full EoU + Survived + H4=3','A'),
    ('SL-003','Low-power wake-word detection MCU','Critical (1-2yr)','Full EoU; H1=2 — tight case','A'),
    ('SL-004','Fingerprint anti-spoofing impedance','Critical (1-2yr)','Full EoU + Survived + Broad claims','A'),
    ('SL-005','DRAM refresh power optimization','Critical (1-2yr)','Partial EoU + Survived + H4=3','A'),
    ('SL-006','Packet classification TCAM','Critical (1-2yr)','Full EoU + D1=3; no survived','B'),
    ('SL-007','EV charging session auth PLC','Critical (1-2yr)','Full EoU; H2=3 low invalidity risk','B'),
    ('SL-008','Predictive BMS cell balancing ML','Critical (1-2yr)','Full EoU + Survived; narrow term','A'),
    ('SL-009','UWB indoor localization ranging','Critical (1-2yr)','Full EoU + D1=3 + H1=3','A'),
    ('SL-010','Adaptive video codec rate control','Critical (1-2yr)','Full EoU + broad claims; minor title gap → B','B'),
    ('SL-011','5G NR HARQ-ACK timing control','Monetization (3-5yr)','Partial EoU + 5G NR SEP potential','B'),
    ('SL-012','OLED pixel compensation circuit','Monetization (3-5yr)','Partial EoU + Survived + high D3','A'),
    ('SL-013','Sparse convolution object detection','Monetization (3-5yr)','Partial EoU + cross-industry; no survived','B'),
    ('SL-014','Li-S battery electrolyte stabilization','Monetization (3-5yr)','Pioneer claims (C2=3) + broad + no EoU','B'),
    ('SL-015','Private 5G network slice orchestration','Monetization (3-5yr)','Partial EoU + SEP-adjacent + D1=2','B'),
    ('SL-016','Secure enclave key derivation IoT','Monetization (3-5yr)','Full EoU + Survived + clean title','A'),
    ('SL-017','MLCC dielectric formulation','Monetization (3-5yr)','Pioneer claims + trilateral + D2=3','B'),
    ('SL-018','Cooperative MIMO C-RAN precoding','Monetization (3-5yr)','Partial EoU + D1=3; no survived — urgent','B'),
    ('SL-019','LiDAR point cloud compression AD','Monetization (3-5yr)','Partial EoU + cross-industry (auto+robotics)','B'),
    ('SL-020','Anisotropic thermal interface material','Monetization (3-5yr)','Pioneer claims + wide geo coverage','B'),
    ('SL-021','Wi-Fi 6E OFDMA scheduling','Strategic (6-8yr)','Survived IPR + SEP + trilateral','A'),
    ('SL-022','Solid-state LiDAR MEMS mirror control','Strategic (6-8yr)','Pioneer claims + D3=3 + OC-DEF viable','A'),
    ('SL-023','In-memory computing DNN inference','Strategic (6-8yr)','Pioneer claims + 33 fwd citations + H4=3','A'),
    ('SL-024','Perovskite solar stability enhancement','Strategic (6-8yr)','Pioneer claims + wide geo; early stage','B'),
    ('SL-025','Federated learning traffic anomaly','Strategic (6-8yr)','Cross-industry (security+telecom+IoT)','B'),
    ('SL-026','CNT FET fabrication process','Strategic (6-8yr)','Pioneer + 22 fwd cit + next-gen; no EoU','B'),
    ('SL-027','5G NR / LTE dynamic spectrum sharing','Strategic (6-8yr)','SEP potential + D1=3 + detectable','A'),
    ('SL-028','Adaptive ANC for headphones','Strategic (6-8yr)','Full EoU + Survived + D1=3 + OC-DEF prime','A'),
    ('SL-029','Blockchain supply chain provenance','Strategic (6-8yr)','H1=1; Salvage — route to B31','C'),
    ('SL-030','GaN power amplifier 5G base stations','Strategic (6-8yr)','Partial EoU + trilateral + OC-DEF viable','B'),
]

TIER_COLORS = {'A': (GREEN, LGREEN), 'B': (GOLD, LGOLD), 'C': (RED, LRED)}
BAND_TXT_COLORS = {'Critical (1-2yr)': '7B0000', 'Monetization (3-5yr)': ORANGE, 'Strategic (6-8yr)': GREEN}
BAND_BG_COLORS = {'Critical (1-2yr)': 'FFF0F0', 'Monetization (3-5yr)': 'FFF8EE', 'Strategic (6-8yr)': 'F0F8F0'}

for i, (pid, title, band, signals_txt, tier) in enumerate(index_data):
    bg = LGRAY if i % 2 == 0 else WHITE
    tc, bc = TIER_COLORS[tier]
    guide_cell(ws_g, row, 1, pid, bold=True, fg=TEAL, bg=bg, align_h='center')
    guide_cell(ws_g, row, 2, title, fg='222222', bg=bg, sz=9)
    guide_cell(ws_g, row, 3, band, bold=True,
               fg=BAND_TXT_COLORS[band],
               bg=BAND_BG_COLORS[band], align_h='center', sz=9)
    guide_cell(ws_g, row, 4, signals_txt, fg=DGRAY, bg=bg, sz=9)
    guide_cell(ws_g, row, 5, f'Tier {tier}', bold=True, fg=tc, bg=bc, align_h='center')
    ws_g.row_dimensions[row].height = 18
    row += 1

row += 2

# ── Footer ─────────────────────────────────────────────────────────────────
ws_g.row_dimensions[row].height = 16
c = ws_g.cell(row, 1, 'Patent Portfolio Bundling Template v5  ·  Portfolio Context Guide  ·  Companion: Patent_Buyer_Profiles_v1.docx & Value_Proposition_Framework_v3.md')
c.font = Font(name='Arial', italic=True, size=9, color=DGRAY)
c.fill = PF('solid', start_color=LGRAY, end_color=LGRAY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws_g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

# ─── 4. UPDATE README with v5 notes ─────────────────────────────────────────
ws_r = wb['README']
# Find the last populated row and append v5 addition notes
last_row = ws_r.max_row + 2
c = ws_r.cell(last_row, 1,
    'v5 ADDITION — Short-Lifecycle Portfolio Context\n'
    '• Three new presets in Presets sheet (cols 13-15): '
    '"Short Lifecycle — Critical (1-2yr)", "Short Lifecycle — Monetization (3-5yr)", "Short Lifecycle — Strategic (6-8yr)".\n'
    '• 30 new sample patents (SL-001 to SL-030) in Patent Portfolio, 10 per term band, spanning wireless, battery, AI, semiconductors, sensors.\n'
    '• New "Portfolio Context Guide" sheet: term band overview, buyer fit per band, critical signals, deal killers, '
    'IAM Market language register, 3-tier composition strategy, and sample patent index.\n'
    '• PreExpiry thresholds in each new preset are set to match the band window; Strength_term_min is adjusted per band; '
    'Anchor_H1_cutoff raised to 3 for Critical band.\n'
    '• Bundle Assignment extended for all 30 new patents with formulas matching the existing pattern.'
)
c.font = Font(name='Arial', size=10)
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_r.row_dimensions[last_row].height = 100

# ─── 5. UPDATE Configuration dropdown DV range ──────────────────────────────
# The Active Preset DV already updated above. Also extend Manual Override DV
for dv in ws_c.data_validations.dataValidation:
    # The E column (Manual Override) for bundle toggles
    if 'E10' in str(dv.sqref) and 'Yes' in str(dv.formula1):
        # Already covers rows 10-42 and 63-67 for YN fields; keep as is
        pass

# ─── Save ────────────────────────────────────────────────────────────────────
out_path = '/home/claude/Patent_Bundling_Template_v5.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheets: {wb.sheetnames}')
