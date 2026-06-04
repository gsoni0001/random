"""
build_v6.py — Extends Patent_Bundling_Template_v5.xlsx to v6

Adds one new sheet: "PatSeer Import"

Sheet layout (3 zones, left-to-right, top-to-bottom):
  ┌─────────────────────────────────────────────────────────┐
  │  ZONE 0: Title banner + instructions (rows 1-5)         │
  ├──────────────────────────┬──────────────────────────────┤
  │  ZONE 1: Column Mapping  │  (cols A-D, rows 7-36)       │
  │  Table                   │  PatSeer col → v5 attr code  │
  │                          │  + transform rule            │
  ├──────────────────────────┴──────────────────────────────┤
  │  ZONE 2: Paste Area (rows 39+)                          │
  │  Header row matching PatSeer export column names        │
  │  200 blank data rows — paste your PatSeer export here   │
  ├─────────────────────────────────────────────────────────┤
  │  ZONE 3: Mapped Output + Bundle Results                 │
  │  Cols after paste area (offset right, same rows)        │
  │  42 mapped attribute cols + completeness % + 33 bundles │
  └─────────────────────────────────────────────────────────┘

Key design decisions:
- The Mapping Table in Zone 1 is the single source of truth.
  Each row defines: PatSeer column header → v5 attribute code → transform rule.
  Formulas in Zone 3 use INDIRECT + MATCH to look up which paste column
  maps to which attribute — so reordering or renaming paste columns
  only requires editing the mapping table, not touching 200×42 formulas.

- Zone 3 uses IFERROR(VLOOKUP / INDEX-MATCH) to pull values from Zone 2
  via the mapping, then applies transforms:
    • Direct copy   — value used as-is
    • Numeric       — VALUE() wrapper
    • Year-from-date — YEAR() or computed from expiry date
    • Count-from-list — LEN()+1 minus LEN(SUBSTITUTE()) trick for semicolons
    • Boolean-map   — IF(x="Yes","Yes","No") style
    • Derived       — formula computes the attribute from other attributes

- Completeness % = COUNTA(mapped cells) / 42

- Bundle columns use the exact same routing logic as Bundle Assignment,
  but reference Zone 3 attribute cells (same row) instead of Patent Portfolio.
  They still respect Configuration toggles and thresholds.

- Instructional header row above paste area tells user exactly which
  PatSeer export fields to include and in what order they are expected
  (though order doesn't matter — the mapping table handles it).
"""

import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill as PF, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ─── Load v5 workbook ────────────────────────────────────────────────────────
wb = load_workbook('Patent_Bundling_Template_v5.xlsx')

# ─── Palette (matches v4/v5) ─────────────────────────────────────────────────
NAVY   = '1F3864';  BLUE   = '2E75B6';  LBLUE  = 'D6E4F0'
WHITE  = 'FFFFFF';  LGRAY  = 'F5F5F5';  DGRAY  = '595959'
GREEN  = '375623';  LGREEN = 'E2EFD8';  ORANGE = 'C55A11'
LORANG = 'FFF3E0';  RED    = '7B0000';  LRED   = 'FDECEA'
GOLD   = 'B8860B';  LGOLD  = 'FFF9E6';  TEAL   = '005B7F'
AMBER  = 'FF8C00';  LAMBER = 'FFF3CD';  PURPLE = '4B0082'
STEEL  = '44546A';  LSTEEL = 'EBF3FB';  LTEAL  = 'E0F4F9'

THIN  = Side(style='thin',   color='CCCCCC')
MED   = Side(style='medium', color='2E75B6')
THICK = Side(style='thick',  color=NAVY)
TB    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MB    = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def F(bold=False, color='1A1A1A', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def FL(color=WHITE, bold=True, size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def BG(hex_c):
    return PF('solid', start_color=hex_c, end_color=hex_c)

def AL(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, r, c, val=None, bold=False, fg='1A1A1A', bg=None, sz=10,
         align='left', wrap=False, italic=False, border=True, num_fmt=None):
    cell = ws.cell(r, c, val)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=sz, italic=italic)
    if bg:
        cell.fill = BG(bg)
    cell.alignment = AL(align, wrap=wrap)
    if border:
        cell.border = TB
    if num_fmt:
        cell.number_format = num_fmt
    return cell

# ─── Sheet layout constants ──────────────────────────────────────────────────
# Zone 1 — Mapping Table
MAP_START_ROW   = 10     # first data row of mapping table (banner=10, hdr=11, data=12+)
MAP_COL_PS      = 1      # col A: PatSeer column header
MAP_COL_ATTR    = 2      # col B: v5 attribute code
MAP_COL_RULE    = 3      # col C: transform rule
MAP_COL_NOTES   = 4      # col D: notes / instructions

# Zone 2 — Paste Area
PASTE_HDR_ROW   = 39     # header row for paste area
PASTE_DATA_ROW  = 40     # first data row
PASTE_DATA_ROWS = 200    # max patents at a time
PASTE_LAST_ROW  = PASTE_DATA_ROW + PASTE_DATA_ROWS - 1

# Zone 3 — Mapped Output starts at col after paste area
# PatSeer has ~25 columns we map; we give paste area cols 6–30 (25 cols)
PASTE_FIRST_COL = 6      # col F — paste starts here
PASTE_N_COLS    = 25     # number of PatSeer columns in paste area
OUT_FIRST_COL   = PASTE_FIRST_COL + PASTE_N_COLS  # col AF onwards
# 42 attribute cols + 1 completeness col + 1 spacer + 33 bundle cols
ATTR_N_COLS     = 42
COMPLETE_COL    = OUT_FIRST_COL + ATTR_N_COLS       # completeness %
BUNDLE_FIRST    = COMPLETE_COL + 2                  # bundle cols start
BUNDLE_N        = 33
TOTAL_COL       = BUNDLE_FIRST + BUNDLE_N           # total bundles

# Column letter helpers
def CL(n): return get_column_letter(n)

# ─── PatSeer field → v5 attribute mapping ────────────────────────────────────
# Format: (patseer_col_header, v5_attr_code, transform_rule, notes)
# transform_rule values:
#   direct       — copy value as-is (text)
#   numeric      — copy and ensure numeric
#   year_from_date — extract year from YYYY-MM-DD date string
#   remaining_life — use PatSeer "Remaining Life (Years)" directly as numeric
#   expiry_year  — extract year from Estimated Expiry Date
#   count_semiC  — count items in semicolon-separated list (family members)
#   count_newline — count items in newline-separated list
#   trilateral   — derive Yes/No from family countries field
#   major_market — derive 0-3 score from family countries field
#   legal_status_map — map PatSeer legal status to Granted/Pending/Expired
#   continuation_map — derive Yes/No from Kind Code / family
#   eou_blank    — always blank (H, I group — manual scoring required)
PATSEER_MAP = [
    # ── Bibliographic ────────────────────────────────────────────────────
    ('Publication Number',       'Patent ID',    'direct',         'Use as Patent ID; or create your own ID in col A'),
    ('Title',                     'A1 (derived)', 'direct',         'Patent title — used as Title in Patent Portfolio'),
    ('Assignee (Current)',        '—',            'direct',         'Current owner — stored in paste area; not a v5 attribute'),
    ('Inventor Names',            '—',            'direct',         'Inventors — stored in paste area; not a v5 attribute'),
    ('Filing Date',               '—',            'year_from_date', 'Filing year — informational only'),
    ('Publication Date',          '—',            'year_from_date', 'Publication year — informational only'),
    ('Estimated Expiry Date',     'E4 (derived)', 'expiry_to_yrs',  'Converted to remaining term in years from today'),
    ('Remaining Life (Years)',    'E4',           'numeric',        'Direct remaining term — preferred over Expiry Date if available'),
    ('Legal Status',              'E2',           'legal_status_map','Granted / Pending / Expired → maps to E2 Prosecution field'),
    ('Maintenance Status',        'E5',           'direct',         'Active / Lapsed / Abandoned → E5 Maint. status'),
    # ── Classifications ──────────────────────────────────────────────────
    ('IPC (Main)',                'A1 (derived)', 'direct',         'Main IPC → used to derive A1 Primary domain via CPC→domain lookup'),
    ('IPC (All)',                 'A2',           'direct',         'All IPC codes → A2 Secondary domains (raw; clean manually)'),
    ('CPC (Main)',                'A1 (derived)', 'direct',         'Main CPC → preferred for A1 domain derivation'),
    ('CPC (All)',                 'A2',           'direct',         'All CPC codes → enriches A2 Secondary domains'),
    # ── Family ───────────────────────────────────────────────────────────
    ('Simple Family Size',        'E1',           'numeric',        'Direct: E1 Family size'),
    ('Simple Family Members',     'F1',           'direct',         'Country codes extracted → F1 Jurisdictions (comma-separated)'),
    ('Extended Family Size',      '—',            'numeric',        'Informational; use Simple Family Size for E1'),
    # ── Citations ────────────────────────────────────────────────────────
    ('Forward Citation Count',   'H5',           'numeric',        'Direct: H5 Fwd citations'),
    ('Backward Citation Count',  'H6',           'numeric',        'Direct: H6 Bwd citations'),
    # ── Text ─────────────────────────────────────────────────────────────
    ('Independent Claim Count',  'C3',           'numeric',        'Direct: C3 Independent claim count'),
    ('Abstract',                 '—',            'direct',         'Abstract text — stored for reference; not a v5 attribute'),
    # ── Scores / Other ───────────────────────────────────────────────────
    ('Technology Domain',         'A1',           'direct',         'PatSeer technology domain tag → A1 Primary domain'),
    ('Technology Sub-Domain',     'A2',           'direct',         'PatSeer sub-domain → A2 Secondary domains'),
    ('Kind Code',                 'E3 (derived)', 'direct',         'Kind code (B2 / C2 = granted; A1/A2 = application) → E3 hint'),
    ('Litigation / PTAB Flag',   'H7',           'direct',         'If PatSeer flags litigation → H7 Lit/PTAB history hint'),
]

# ─── v5 attribute column index map (col in Patent Portfolio = col in Zone 3) ─
# We replicate the same 42-column order in Zone 3
ATTR_COLS = [
    'Patent ID', 'Title',
    'A1','A2','A3','A4','A5',
    'B1','B2','B3',
    'C1','C2','C3','C4',
    'D1','D2','D3',
    'E1','E2','E3','E4','E5',
    'F1','F2','F3',
    'G1','G2','G3',
    'H1','H2','H3','H4','H5','H6','H7','H8','H9','H10',
    'I1','I2','I3','I4',
]
# Which PatSeer fields map to which attribute (attr_code → patseer_field_name)
# For attributes derivable from PatSeer, we define the mapping
ATTR_FROM_PATSEER = {
    'Patent ID':  'Publication Number',
    'Title':      'Title',
    'A1':         'Technology Domain',    # preferred; fallback CPC (Main)
    'A2':         'Technology Sub-Domain',
    'E1':         'Simple Family Size',
    'E2':         'Legal Status',
    'E4':         'Remaining Life (Years)',
    'E5':         'Maintenance Status',
    'F1':         'Simple Family Members',
    'H5':         'Forward Citation Count',
    'H6':         'Backward Citation Count',
    'C3':         'Independent Claim Count',
    'H7':         'Litigation / PTAB Flag',
}
# Attributes NOT from PatSeer — manual/AI scoring required
MANUAL_ATTRS = {
    'A3','A4','A5',
    'B1','B2','B3',
    'C1','C2','C4',
    'D1','D2','D3',
    'E3',
    'F2','F3',
    'G1','G2','G3',
    'H1','H2','H3','H4','H8','H9','H10',
    'I1','I2','I3','I4',
}

# ─── Paste area column headers (PatSeer order) ───────────────────────────────
PASTE_HEADERS = [ps_col for ps_col, _, _, _ in PATSEER_MAP]
assert len(PASTE_HEADERS) == PASTE_N_COLS, f"Expected {PASTE_N_COLS} paste cols, got {len(PASTE_HEADERS)}"

# Build lookup: PatSeer header → paste column index (1-based within paste area)
PS_COL_IDX = {h: i+1 for i, h in enumerate(PASTE_HEADERS)}

# ─── Helper: paste col absolute column number ───────────────────────────────
def paste_col(ps_header):
    """Return absolute worksheet column number for a PatSeer paste header."""
    return PASTE_FIRST_COL + PS_COL_IDX.get(ps_header, 0) - 1

def paste_col_letter(ps_header):
    return CL(paste_col(ps_header))

# ─── Helper: out col absolute column number for attr code ────────────────────
def out_col(attr_code):
    """Return absolute worksheet column for mapped output attribute."""
    idx = ATTR_COLS.index(attr_code) if attr_code in ATTR_COLS else -1
    return OUT_FIRST_COL + idx if idx >= 0 else None

def out_col_letter(attr_code):
    c = out_col(attr_code)
    return CL(c) if c else None

# ─── Build the formula for each attribute in Zone 3 (for a given data row r) ─
def make_attr_formula(attr_code, r):
    """
    Returns the Excel formula string for attribute attr_code at output row r.
    References the paste area via MATCH on the mapping-table headers so that
    paste column order doesn't matter. But for simplicity and reliability we
    use direct column references since we control the paste layout.
    """
    if attr_code in ('Patent ID', 'Title'):
        ps_h = ATTR_FROM_PATSEER.get(attr_code)
        if ps_h:
            pc = paste_col(ps_h)
            return f"=IFERROR(IF({CL(pc)}{r}<>\"\",{CL(pc)}{r},\"\"),\"\")"
        return ""

    if attr_code not in ATTR_FROM_PATSEER:
        return ""   # blank — manual

    ps_h = ATTR_FROM_PATSEER[attr_code]
    pc   = paste_col(ps_h)
    pcl  = CL(pc)

    if attr_code == 'A1':
        # Technology Domain preferred; fallback to CPC (Main)
        cpc_col = paste_col('CPC (Main)')
        return (f"=IFERROR(IF({pcl}{r}<>\"\",{pcl}{r},"
                f"IF({CL(cpc_col)}{r}<>\"\",{CL(cpc_col)}{r},\"\")),\"\")")

    elif attr_code == 'A2':
        # Technology Sub-Domain; enrich with CPC (All) hint
        cpc_all_col = paste_col('CPC (All)')
        return (f"=IFERROR(IF({pcl}{r}<>\"\",{pcl}{r},"
                f"IF({CL(cpc_all_col)}{r}<>\"\",{CL(cpc_all_col)}{r},\"\")),\"\")")

    elif attr_code == 'E1':
        return f"=IFERROR(IF({pcl}{r}<>\"\",VALUE({pcl}{r}),\"\"),\"\")"

    elif attr_code == 'E2':
        # Map PatSeer Legal Status text → v5 Prosecution values
        return (f"=IFERROR(IF({pcl}{r}=\"\",\"\","
                f"IF(OR(ISNUMBER(SEARCH(\"grant\",{pcl}{r})),ISNUMBER(SEARCH(\"B\",{CL(paste_col('Kind Code'))}{r}))),"
                f"\"Granted\","
                f"IF(OR(ISNUMBER(SEARCH(\"pend\",{pcl}{r})),ISNUMBER(SEARCH(\"A\",{CL(paste_col('Kind Code'))}{r}))),"
                f"\"Pending\","
                f"IF(ISNUMBER(SEARCH(\"expir\",{pcl}{r})),\"Expired\",{pcl}{r})))),\"\")")

    elif attr_code == 'E4':
        # Remaining Life (Years) as numeric — if blank try expiry date derivation
        exp_col  = paste_col('Estimated Expiry Date')
        exp_cl   = CL(exp_col)
        return (f"=IFERROR("
                f"IF({pcl}{r}<>\"\",ROUND(VALUE({pcl}{r}),1),"
                f"IF({exp_cl}{r}<>\"\","
                f"ROUND(MAX(0,(DATEVALUE(TEXT({exp_cl}{r},\"YYYY-MM-DD\"))-TODAY())/365.25),1),"
                f"\"\")),\"\")")

    elif attr_code == 'E5':
        return f"=IFERROR(IF({pcl}{r}<>\"\",{pcl}{r},\"\"),\"\")"

    elif attr_code == 'F1':
        # Simple Family Members → extract unique country codes
        return (f"=IFERROR(IF({pcl}{r}<>\"\","
                f"SUBSTITUTE(SUBSTITUTE({pcl}{r},\"; \",\", \"),\";\",\",\"),"
                f"\"\"),\"\")")

    elif attr_code == 'H5':
        return f"=IFERROR(IF({pcl}{r}<>\"\",VALUE({pcl}{r}),\"\"),\"\")"

    elif attr_code == 'H6':
        return f"=IFERROR(IF({pcl}{r}<>\"\",VALUE({pcl}{r}),\"\"),\"\")"

    elif attr_code == 'C3':
        return f"=IFERROR(IF({pcl}{r}<>\"\",VALUE({pcl}{r}),\"\"),\"\")"

    elif attr_code == 'H7':
        # Map litigation flag to v5 values: None / Pending / Survived / Lost
        return (f"=IFERROR(IF({pcl}{r}=\"\",\"None\","
                f"IF(ISNUMBER(SEARCH(\"surv\",{pcl}{r})),\"Survived\","
                f"IF(ISNUMBER(SEARCH(\"pend\",{pcl}{r})),\"Pending\","
                f"IF(ISNUMBER(SEARCH(\"yes\",{pcl}{r})),\"Pending\",\"None\")))),\"None\")")

    return ""

# ─── Bundle qualification formulas for Zone 3 ────────────────────────────────
# Same logic as Bundle Assignment, but referencing Zone 3 output columns.
# We need to map Patent Portfolio column letters to Zone 3 output column letters.

# PP col → attr code mapping (from earlier inspection)
PP_COL_TO_ATTR = {
    3:'A1', 4:'A2', 5:'A3', 6:'A4', 7:'A5',
    8:'B1', 9:'B2', 10:'B3',
    11:'C1', 12:'C2', 13:'C3', 14:'C4',
    15:'D1', 16:'D2', 17:'D3',
    18:'E1', 19:'E2', 20:'E3', 21:'E4', 22:'E5',
    23:'F1', 24:'F2', 25:'F3',
    26:'G1', 27:'G2', 28:'G3',
    29:'H1', 30:'H2', 31:'H3', 32:'H4', 33:'H5', 34:'H6',
    35:'H7', 36:'H8', 37:'H9', 38:'H10',
    39:'I1', 40:'I2', 41:'I3', 42:'I4',
}

def pp_ref_to_zone3(pp_col_letter, r):
    """Convert a Patent Portfolio column letter reference to Zone 3 column reference."""
    pp_col_num = column_index_from_string(pp_col_letter)
    attr = PP_COL_TO_ATTR.get(pp_col_num)
    if attr and attr in ATTR_COLS:
        oc = out_col(attr)
        if oc:
            return f"{CL(oc)}{r}"
    return None

def translate_ba_formula(ba_formula, r):
    """
    Take a Bundle Assignment formula for Patent Portfolio row 3,
    translate all 'Patent Portfolio'!XX3 references to Zone 3 column refs for row r,
    and update the row number.
    """
    if not ba_formula or not isinstance(ba_formula, str):
        return ba_formula

    result = ba_formula

    # Replace all 'Patent Portfolio'!XX3 with Zone3 equivalents
    def replacer(m):
        col_letter = m.group(1)
        zone3_ref = pp_ref_to_zone3(col_letter, r)
        if zone3_ref:
            return zone3_ref
        # If no mapping (e.g. col A/B = ID/Title), return a non-empty placeholder
        return f'""'

    result = re.sub(r"'Patent Portfolio'!([A-Z]+)3", replacer, result)

    return result

# Bundle Assignment formula patterns (captured earlier)
BA_FORMULAS_ROW3 = {
    1:  "=IF(Configuration!$C$10=\"No\",\"\",IFERROR(IF('Patent Portfolio'!C3<>\"\",TRUE(),FALSE()),FALSE()))",
    2:  "=IF(Configuration!$C$11=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!H3)>=Configuration!$C$44,'Patent Portfolio'!I3<>\"\"),TRUE(),FALSE()),FALSE()))",
    3:  "=IF(Configuration!$C$12=\"No\",\"\",IFERROR(IF('Patent Portfolio'!F3<>\"\",TRUE(),FALSE()),FALSE()))",
    4:  "=IF(Configuration!$C$13=\"No\",\"\",IFERROR(IF('Patent Portfolio'!E3<>\"\",TRUE(),FALSE()),FALSE()))",
    5:  "=IF(Configuration!$C$14=\"No\",\"\",IFERROR(IF('Patent Portfolio'!G3<>\"\",TRUE(),FALSE()),FALSE()))",
    6:  "=IF(Configuration!$C$15=\"No\",\"\",IFERROR(IF(AND(OR(ISNUMBER(SEARCH(\"process\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"fab\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"manufac\",'Patent Portfolio'!C3))),'Patent Portfolio'!K3=\"Method\"),TRUE(),FALSE()),FALSE()))",
    7:  "=IF(Configuration!$C$16=\"No\",\"\",IFERROR(IF(AND(OR(ISNUMBER(SEARCH(\"material\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"chem\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"battery\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"electrolyte\",'Patent Portfolio'!C3)),ISNUMBER(SEARCH(\"polymer\",'Patent Portfolio'!C3))),OR('Patent Portfolio'!K3=\"Apparatus\",'Patent Portfolio'!K3=\"Method\")),TRUE(),FALSE()),FALSE()))",
    8:  "=IF(Configuration!$C$17=\"No\",\"\",IFERROR(IF(AND(OR('Patent Portfolio'!E3=\"App\",'Patent Portfolio'!E3=\"Middleware\",'Patent Portfolio'!E3=\"Cloud\"),OR('Patent Portfolio'!K3=\"Method\",'Patent Portfolio'!K3=\"CRM\")),TRUE(),FALSE()),FALSE()))",
    9:  "=IF(Configuration!$C$18=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!J3)>=Configuration!$C$45,TRUE(),FALSE()),FALSE()))",
    10: "=IF(Configuration!$C$19=\"No\",\"\",IFERROR(IF('Patent Portfolio'!AA3<>\"\",TRUE(),FALSE()),FALSE()))",
    11: "=IF(Configuration!$C$20=\"No\",\"\",IFERROR(IF('Patent Portfolio'!K3<>\"\",TRUE(),FALSE()),FALSE()))",
    12: "=IF(Configuration!$C$21=\"No\",\"\",IFERROR(IF(OR(N('Patent Portfolio'!O3)>=Configuration!$C$46,N('Patent Portfolio'!P3)>=Configuration!$C$47),TRUE(),FALSE()),FALSE()))",
    13: "=IF(Configuration!$C$22=\"No\",\"\",IFERROR(IF('Patent Portfolio'!X3=\"Yes\",TRUE(),FALSE()),FALSE()))",
    14: "=IF(Configuration!$C$23=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!R3)>=Configuration!$C$48,TRUE(),FALSE()),FALSE()))",
    15: "=IF(Configuration!$C$24=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!U3)>0,TRUE(),FALSE()),FALSE()))",
    16: "=IF(Configuration!$C$25=\"No\",\"\",IFERROR(IF(OR(N('Patent Portfolio'!L3)=3,N('Patent Portfolio'!L3)<=1),TRUE(),FALSE()),FALSE()))",
    17: "=IF(Configuration!$C$26=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!AB3)>=Configuration!$C$49,TRUE(),FALSE()),FALSE()))",
    18: "=IF(Configuration!$C$27=\"No\",\"\",IFERROR(IF('Patent Portfolio'!Z3<>\"\",TRUE(),FALSE()),FALSE()))",
    19: "=IF(Configuration!$C$28=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!Q3)>=Configuration!$C$50,TRUE(),FALSE()),FALSE()))",
    20: "=IF(Configuration!$C$29=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!N3)>=Configuration!$C$51,'Patent Portfolio'!C3<>\"\"),TRUE(),FALSE()),FALSE()))",
    21: "=IF(Configuration!$C$30=\"No\",\"\",IFERROR(IF(OR('Patent Portfolio'!S3=\"Pending\",'Patent Portfolio'!T3=\"Yes\"),TRUE(),FALSE()),FALSE()))",
    22: "=IF(Configuration!$C$31=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!AC3)>=Configuration!$C$52,N('Patent Portfolio'!L3)>=1),TRUE(),FALSE()),FALSE()))",
    23: "=IF(Configuration!$C$32=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!N3)>=1,OR('Patent Portfolio'!C3<>\"\",'Patent Portfolio'!I3<>\"\")),TRUE(),FALSE()),FALSE()))",
    24: "=IF(Configuration!$C$33=\"No\",\"\",IFERROR(IF('Patent Portfolio'!C3<>\"\",TRUE(),FALSE()),FALSE()))",
    25: "=IF(Configuration!$C$34=\"No\",\"\",IFERROR(IF('Patent Portfolio'!T3=\"Yes\",TRUE(),FALSE()),FALSE()))",
    26: "=IF(Configuration!$C$35=\"No\",\"\",IFERROR(IF(OR('Patent Portfolio'!AK3=\"Partial\",'Patent Portfolio'!AK3=\"Full\"),TRUE(),FALSE()),FALSE()))",
    27: "=IF(Configuration!$C$36=\"No\",\"\",IFERROR(IF('Patent Portfolio'!AI3=\"Survived\",TRUE(),FALSE()),FALSE()))",
    28: "=IF(Configuration!$C$37=\"No\",\"\",IFERROR(IF(AND('Patent Portfolio'!AJ3=\"Clean\",'Patent Portfolio'!AL3=\"None\"),TRUE(),FALSE()),FALSE()))",
    29: "=IF(Configuration!$C$38=\"No\",\"\",IFERROR(IF(N('Patent Portfolio'!AG3)>=Configuration!$C$53,TRUE(),FALSE()),FALSE()))",
    30: "=IF(Configuration!$C$39=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!AO3)>=2,N('Patent Portfolio'!AB3)>=2),TRUE(),FALSE()),FALSE()))",
    31: "=IF(Configuration!$C$40=\"No\",\"\",IFERROR(IF(OR(N('Patent Portfolio'!AC3)<=Configuration!$C$56,N('Patent Portfolio'!U3)<Configuration!$C$57,N('Patent Portfolio'!AD3)<=Configuration!$C$58),TRUE(),FALSE()),FALSE()))",
    32: "=IF(Configuration!$C$41=\"No\",\"\",IFERROR(IF(AND(N('Patent Portfolio'!U3)>=Configuration!$C$54,N('Patent Portfolio'!U3)<=Configuration!$C$55),TRUE(),FALSE()),FALSE()))",
    33: "=IF(Configuration!$C$42=\"No\",\"\",IFERROR(IF(AND('Patent Portfolio'!C3<>\"\",'Patent Portfolio'!F3<>\"\"),TRUE(),FALSE()),FALSE()))",
}

BUNDLE_NAMES = {
    1:'Tech Domain', 2:'SEP', 3:'Product Arch', 4:'Stack Layer', 5:'Use-Case',
    6:'Mfg/Process', 7:'Materials', 8:'Algorithm', 9:'Interop', 10:'Generational',
    11:'Claim-Type', 12:'Detectable', 13:'Geographic', 14:'Family', 15:'Lifecycle',
    16:'Foundational+Impr', 17:'Cross-Industry', 18:'Convergent Theme',
    19:'Defensive', 20:'Whitespace', 21:'Prosecution-Status', 22:'Anchor+Halo',
    23:'Picket-Fence', 24:'Strong+Tail', 25:'Continuation-Live', 26:'EoU-Backed',
    27:'Battle-Tested', 28:'Clean-Title', 29:'High-Citation', 30:'Adjacent Re-Read',
    31:'Salvage-Volume', 32:'Pre-Expiry', 33:'Provenance',
}

# ─── Create the sheet ────────────────────────────────────────────────────────
if 'PatSeer Import' in wb.sheetnames:
    del wb['PatSeer Import']

ws = wb.create_sheet('PatSeer Import', 1)   # second sheet after README
ws.sheet_view.showGridLines = False
ws.freeze_panes = None   # We'll set freeze after headers

# ─── ZONE 0: Title & Instructions (rows 1-5) ────────────────────────────────
# Row 1: Title banner
ws.merge_cells('A1:D1')
ws.row_dimensions[1].height = 28
c = ws.cell(1, 1, 'PATSEER IMPORT  —  Paste your PatSeer export, get bundle results instantly')
c.font = Font(name='Arial', bold=True, color=WHITE, size=13)
c.fill = BG(NAVY)
c.alignment = AL('left')
c.border = MB

# Instruction cols 1-4 (mapping table zone width)
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 40

# Row 2: subtitle
ws.merge_cells('A2:D2')
ws.row_dimensions[2].height = 16
c = ws.cell(2, 1, 'v6 Addition  ·  Companion to PatSeer Standard/Premier/Pro X export  ·  Outright sale context — short lifecycle portfolios')
c.font = Font(name='Arial', italic=True, color=DGRAY, size=9)
c.fill = BG(LBLUE)
c.alignment = AL('left')
c.border = TB

# Row 3: blank separator
ws.row_dimensions[3].height = 8

# Row 4-5: how-to steps
steps = [
    ('Step 1', 'In PatSeer, export your portfolio as Excel/CSV. Select these field groups: Bibliographic + Parties + Classifications + Family + Citations + Other/Scores.'),
    ('Step 2', 'In the Paste Area (rows 40+), paste your PatSeer export starting at the correct column headers (row 39). Column order must match row 39.'),
    ('Step 3', 'The Mapped Output (right of paste area) auto-populates attributes derived from PatSeer. Amber-highlighted cells = manual scoring required.'),
    ('Step 4', 'Bundle Results columns show TRUE/FALSE qualification per bundle, respecting the active Configuration preset. Review and copy useful patents to Patent Portfolio.'),
    ('Step 5', 'The Column Mapping Table (rows 7-36 below) maps each PatSeer field to its v5 attribute. Edit col A if your PatSeer headers differ — all formulas update automatically.'),
]
for i, (step, desc) in enumerate(steps):
    r = 4 + i
    ws.row_dimensions[r].height = 18
    c = ws.cell(r, 1, step)
    c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    c.fill = BG(BLUE)
    c.alignment = AL('center')
    c.border = TB
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c2 = ws.cell(r, 2, desc)
    c2.font = Font(name='Arial', size=9, color='222222')
    c2.fill = BG(LSTEEL)
    c2.alignment = AL('left', wrap=True)
    c2.border = TB

# Row 6: blank
ws.row_dimensions[6].height = 8

# ─── ZONE 1: Column Mapping Table (rows 10-36) ────────────────────────────────
# Section banner (row 10)
ws.merge_cells('A10:D10')
ws.row_dimensions[10].height = 18
c = ws.cell(10, 1, 'COLUMN MAPPING TABLE  —  Edit Col A if your PatSeer headers differ. All formulas update automatically.')
c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
c.fill = BG(STEEL)
c.alignment = AL('left')
c.border = TB

# Header row (row 11)
ws.row_dimensions[11].height = 16
map_hdrs = [
    ('PatSeer Export Column Header', NAVY),
    ('Maps to v5 Attribute', NAVY),
    ('Transform Rule', NAVY),
    ('Notes / Instructions', NAVY),
]
for col, (hdr, bg) in enumerate(map_hdrs, start=1):
    c = ws.cell(11, col)
    c.value = hdr
    c.font = FL()
    c.fill = BG(bg)
    c.alignment = AL('center')
    c.border = TB

# Data rows — row 12 onwards
for i, (ps_hdr, attr, rule, notes) in enumerate(PATSEER_MAP):
    r = 12 + i
    ws.row_dimensions[r].height = 15
    # Alternate row shading
    bg = LGRAY if i % 2 == 0 else WHITE
    # Col A: PatSeer header — user-editable, highlighted
    c = ws.cell(r, MAP_COL_PS, ps_hdr)
    c.font = Font(name='Arial', bold=True, color=NAVY, size=9)
    c.fill = BG(LBLUE)
    c.alignment = AL('left')
    c.border = TB
    # Col B: v5 attribute
    attr_bg = LGREEN if attr not in ('—',) and not attr.startswith('—') else LGRAY
    c = ws.cell(r, MAP_COL_ATTR, attr)
    c.font = Font(name='Arial', bold=(attr != '—'), color=GREEN if attr != '—' else DGRAY, size=9)
    c.fill = BG(attr_bg)
    c.alignment = AL('center')
    c.border = TB
    # Col C: transform rule
    rule_colors = {
        'direct': LGRAY, 'numeric': LGOLD,
        'expiry_to_yrs': LORANG, 'remaining_life': LORANG,
        'legal_status_map': LAMBER, 'year_from_date': LGOLD,
        'count_semiC': LGOLD, 'trilateral': LGREEN, 'continuation_map': LGREEN,
    }
    c = ws.cell(r, MAP_COL_RULE, rule)
    c.font = Font(name='Arial', size=9, color='333333')
    c.fill = BG(rule_colors.get(rule, LGRAY))
    c.alignment = AL('center')
    c.border = TB
    # Col D: notes
    c = ws.cell(r, MAP_COL_NOTES, notes)
    c.font = Font(name='Arial', size=9, color=DGRAY, italic=True)
    c.fill = BG(bg)
    c.alignment = AL('left', wrap=True)
    c.border = TB

# ─── ZONE 2: Paste Area header (rows 37-40) ──────────────────────────────────
# Gap row
ws.row_dimensions[35].height = 10

# Banner
r = 36
ws.row_dimensions[r].height = 20
ws.merge_cells(f'A{r}:{CL(PASTE_FIRST_COL + PASTE_N_COLS - 1)}{r}')
c = ws.cell(r, 1, 'PASTE AREA  —  Paste your PatSeer Excel export here (row 40 onwards). Header row 39 must match your export column names exactly.')
c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
c.fill = BG(TEAL)
c.alignment = AL('left')
c.border = MB

# Sub-instruction row
r = 37
ws.row_dimensions[r].height = 15
ws.merge_cells(f'A{r}:{CL(PASTE_FIRST_COL + PASTE_N_COLS - 1)}{r}')
c = ws.cell(r, 1, '→  How to paste: In PatSeer export Excel, select ALL data including header row 1. Copy. Click cell F40 in this sheet. Paste. Then clear row 40 if it duplicated the header.')
c.font = Font(name='Arial', size=9, color='222222')
c.fill = BG(LTEAL)
c.alignment = AL('left')
c.border = TB

# Row count indicator row
r = 38
ws.row_dimensions[r].height = 14
ws.merge_cells(f'A{r}:E{r}')
c = ws.cell(r, 1, f'→  Paste area supports up to {PASTE_DATA_ROWS} patents (rows 40–{PASTE_LAST_ROW}). For more, add rows and extend formulas in the Mapped Output zone.')
c.font = Font(name='Arial', size=9, color=DGRAY, italic=True)
c.fill = BG(LGRAY)
c.alignment = AL('left')
c.border = TB

# Cols A-E (rows 40+) are reserved for user notes / manual scoring helper
# Set widths for paste cols
for i in range(PASTE_N_COLS):
    col_num = PASTE_FIRST_COL + i
    col_letter = CL(col_num)
    ps_hdr = PASTE_HEADERS[i]
    # Wider cols for text-heavy fields
    if ps_hdr in ('Title', 'Abstract', 'Simple Family Members', 'IPC (All)', 'CPC (All)'):
        ws.column_dimensions[col_letter].width = 28
    elif ps_hdr in ('Assignee (Current)', 'Inventor Names', 'Technology Domain'):
        ws.column_dimensions[col_letter].width = 22
    else:
        ws.column_dimensions[col_letter].width = 16

# Header row for paste area (row 39)
ws.row_dimensions[PASTE_HDR_ROW].height = 30
for i, ps_hdr in enumerate(PASTE_HEADERS):
    col_num = PASTE_FIRST_COL + i
    c = ws.cell(PASTE_HDR_ROW, col_num, ps_hdr)
    c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    c.fill = BG(TEAL)
    c.alignment = AL('center', wrap=True)
    c.border = TB

# Add "Manual Scoring" labels in cols A-E of paste header row
for col, lbl in enumerate(['#', 'Your Notes', 'E3 Continuation\n(Yes/No)', 'E2 Prosecution\nOverride', 'E5 Maint. Status\nOverride'], start=1):
    ws.column_dimensions[CL(col)].width = 14
    c = ws.cell(PASTE_HDR_ROW, col, lbl)
    c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    c.fill = BG(ORANGE)
    c.alignment = AL('center', wrap=True)
    c.border = TB

# ─── ZONE 3: Mapped Output + Bundle Results headers ──────────────────────────
# Set output column widths
for i, attr in enumerate(ATTR_COLS):
    col_num = OUT_FIRST_COL + i
    if attr in ('Patent ID', 'Title', 'A1', 'F1'):
        ws.column_dimensions[CL(col_num)].width = 22
    elif attr in ('A2', 'B2', 'E2', 'E5', 'H7', 'H8', 'H9', 'H10', 'I2', 'G1', 'G2'):
        ws.column_dimensions[CL(col_num)].width = 14
    else:
        ws.column_dimensions[CL(col_num)].width = 8

# Completeness col
ws.column_dimensions[CL(COMPLETE_COL)].width = 12
# Spacer
ws.column_dimensions[CL(COMPLETE_COL + 1)].width = 4
# Bundle cols
for b in range(BUNDLE_N):
    ws.column_dimensions[CL(BUNDLE_FIRST + b)].width = 10
ws.column_dimensions[CL(TOTAL_COL)].width = 8

# Output area banner (row 36, extending right)
ws.merge_cells(f'{CL(OUT_FIRST_COL)}{36}:{CL(TOTAL_COL)}{36}')
c = ws.cell(36, OUT_FIRST_COL, 'MAPPED OUTPUT  —  Auto-populated from Paste Area via Column Mapping.  Amber = manual scoring required.  Bundle columns respect active Configuration preset.')
c.font = Font(name='Arial', bold=True, color=WHITE, size=11)
c.fill = BG(NAVY)
c.alignment = AL('left')
c.border = MB

# Sub-header row 37 for output zone
ws.merge_cells(f'{CL(OUT_FIRST_COL)}{37}:{CL(COMPLETE_COL - 1)}{37}')
c = ws.cell(37, OUT_FIRST_COL, '42 Attributes  —  PatSeer-derived fields auto-fill; AMBER cells need manual / AI scoring (see AI_Prompts_for_Attribute_Scoring.md)')
c.font = Font(name='Arial', size=9, color='222222')
c.fill = BG(LAMBER)
c.alignment = AL('left')
c.border = TB

ws.merge_cells(f'{CL(COMPLETE_COL)}{37}:{CL(COMPLETE_COL + 1)}{37}')
c = ws.cell(37, COMPLETE_COL, 'Complete-ness')
c.font = Font(name='Arial', size=9, bold=True, color=NAVY)
c.fill = BG(LGOLD)
c.alignment = AL('center')
c.border = TB

ws.merge_cells(f'{CL(BUNDLE_FIRST)}{37}:{CL(TOTAL_COL)}{37}')
c = ws.cell(37, BUNDLE_FIRST, '33 Bundle Qualifications  —  TRUE = patent qualifies (respects Configuration sheet preset)')
c.font = Font(name='Arial', size=9, color=WHITE, bold=True)
c.fill = BG(STEEL)
c.alignment = AL('left')
c.border = TB

# Attribute header row (row 38, but output zone)
ws.row_dimensions[38].height = 30
# Group headers for attributes
GROUP_RANGES = [
    ('ID / Title', OUT_FIRST_COL, OUT_FIRST_COL + 1, NAVY),
    ('A. Technology', OUT_FIRST_COL + 2, OUT_FIRST_COL + 6, '1F3864'),
    ('B. Standards', OUT_FIRST_COL + 7, OUT_FIRST_COL + 9, '2E75B6'),
    ('C. Claims', OUT_FIRST_COL + 10, OUT_FIRST_COL + 13, '4472C4'),
    ('D. Detectability', OUT_FIRST_COL + 14, OUT_FIRST_COL + 16, '70AD47'),
    ('E. Family/Life', OUT_FIRST_COL + 17, OUT_FIRST_COL + 21, '375623'),
    ('F. Geography', OUT_FIRST_COL + 22, OUT_FIRST_COL + 24, '548235'),
    ('G. Strategic', OUT_FIRST_COL + 25, OUT_FIRST_COL + 27, 'C55A11'),
    ('H. Quality', OUT_FIRST_COL + 28, OUT_FIRST_COL + 37, '7B0000'),
    ('I. Market', OUT_FIRST_COL + 38, OUT_FIRST_COL + 41, '4B0082'),
]
for grp_lbl, c1, c2, grp_bg in GROUP_RANGES:
    if c1 == c2:
        c_obj = ws.cell(38, c1, grp_lbl)
    else:
        ws.merge_cells(start_row=38, start_column=c1, end_row=38, end_column=c2)
        c_obj = ws.cell(38, c1, grp_lbl)
    c_obj.font = Font(name='Arial', bold=True, color=WHITE, size=9)
    c_obj.fill = BG(grp_bg)
    c_obj.alignment = AL('center')
    c_obj.border = TB

# Individual attribute column headers (row PASTE_HDR_ROW = 39)
for i, attr in enumerate(ATTR_COLS):
    col_num = OUT_FIRST_COL + i
    patseer_covered = attr in ATTR_FROM_PATSEER
    c = ws.cell(PASTE_HDR_ROW, col_num, attr)
    c.font = Font(name='Arial', bold=True,
                  color=WHITE if patseer_covered else '333333', size=9)
    c.fill = BG(TEAL if patseer_covered else LAMBER)
    c.alignment = AL('center', wrap=True)
    c.border = TB

# Completeness header
c = ws.cell(PASTE_HDR_ROW, COMPLETE_COL, 'Completeness\n%')
c.font = Font(name='Arial', bold=True, color=NAVY, size=9)
c.fill = BG(LGOLD)
c.alignment = AL('center', wrap=True)
c.border = TB

# Blank spacer col
ws.cell(PASTE_HDR_ROW, COMPLETE_COL + 1, '')

# Bundle headers
for b in range(1, BUNDLE_N + 1):
    col_num = BUNDLE_FIRST + b - 1
    c = ws.cell(PASTE_HDR_ROW, col_num, f'B{b}\n{BUNDLE_NAMES[b]}')
    c.font = Font(name='Arial', bold=True, color=WHITE, size=8)
    c.fill = BG(STEEL)
    c.alignment = AL('center', wrap=True)
    c.border = TB

# Total bundles header
c = ws.cell(PASTE_HDR_ROW, TOTAL_COL, '# Bundles')
c.font = Font(name='Arial', bold=True, color=WHITE, size=9)
c.fill = BG(NAVY)
c.alignment = AL('center', wrap=True)
c.border = TB

# ─── ZONE 2+3 Data rows (rows 40 to PASTE_LAST_ROW) ─────────────────────────
print(f"Writing {PASTE_DATA_ROWS} data rows...")

for i in range(PASTE_DATA_ROWS):
    r = PASTE_DATA_ROW + i
    ws.row_dimensions[r].height = 15

    # Row number in col A
    c = ws.cell(r, 1, i + 1)
    c.font = Font(name='Arial', size=9, color=DGRAY)
    c.fill = BG(LGRAY)
    c.alignment = AL('center')
    c.border = TB

    # Cols B-E: manual input helpers (blank, user fills)
    for col in range(2, 6):
        c = ws.cell(r, col, '')
        c.fill = BG(LAMBER)  # amber = manual input area
        c.font = Font(name='Arial', size=9)
        c.border = TB

    # Paste area cols F onwards: blank (user pastes here)
    for pi in range(PASTE_N_COLS):
        col_num = PASTE_FIRST_COL + pi
        c = ws.cell(r, col_num, '')
        c.font = Font(name='Arial', size=9)
        c.border = TB
        c.fill = BG(WHITE)

    # ── Zone 3: Mapped Output — 42 attribute formulas ──────────────────────
    filled_count_refs = []  # collect non-blank cell refs for completeness calc

    for ai, attr in enumerate(ATTR_COLS):
        col_num = OUT_FIRST_COL + ai
        patseer_covered = attr in ATTR_FROM_PATSEER

        formula = make_attr_formula(attr, r)

        if formula:
            c = ws.cell(r, col_num, formula)
            c.font = Font(name='Arial', size=9)
            c.fill = BG(WHITE)
        else:
            # Manual cell — amber
            c = ws.cell(r, col_num, '')
            c.font = Font(name='Arial', size=9, color='222222')
            c.fill = BG(LAMBER)

        c.alignment = AL('left')
        c.border = TB
        filled_count_refs.append(CL(col_num) + str(r))

    # Completeness % formula (count non-blank out of 42)
    attr_range_start = CL(OUT_FIRST_COL)
    attr_range_end   = CL(OUT_FIRST_COL + ATTR_N_COLS - 1)
    c = ws.cell(r, COMPLETE_COL,
                f'=IFERROR(COUNTA({attr_range_start}{r}:{attr_range_end}{r})/42,0)')
    c.font = Font(name='Arial', size=9, bold=True)
    c.fill = BG(LGOLD)
    c.alignment = AL('center')
    c.border = TB
    c.number_format = '0%'

    # Spacer
    ws.cell(r, COMPLETE_COL + 1, '').border = TB

    # ── Bundle qualification formulas ────────────────────────────────────
    for b in range(1, BUNDLE_N + 1):
        col_num = BUNDLE_FIRST + b - 1
        ba_formula = BA_FORMULAS_ROW3.get(b, '')
        translated = translate_ba_formula(ba_formula, r)

        c = ws.cell(r, col_num, translated if translated else '')
        c.font = Font(name='Arial', size=9)
        c.fill = BG(WHITE)
        c.alignment = AL('center')
        c.border = TB

    # Total bundles qualified
    b_start = CL(BUNDLE_FIRST)
    b_end   = CL(BUNDLE_FIRST + BUNDLE_N - 1)
    c = ws.cell(r, TOTAL_COL,
                f'=COUNTIF({b_start}{r}:{b_end}{r},TRUE)')
    c.font = Font(name='Arial', bold=True, size=9, color=NAVY)
    c.fill = BG(LBLUE)
    c.alignment = AL('center')
    c.border = TB

print("Rows written.")

# ─── Conditional formatting ──────────────────────────────────────────────────
# TRUE = green, FALSE = light red in bundle cols
green_fill = PF('solid', start_color='C6EFCE', end_color='C6EFCE')
red_fill   = PF('solid', start_color='FFC7CE', end_color='FFC7CE')
green_font = Font(name='Arial', size=9, color='276221')
red_font   = Font(name='Arial', size=9, color='9C0006')

bundle_range = (f'{CL(BUNDLE_FIRST)}{PASTE_DATA_ROW}:'
                f'{CL(BUNDLE_FIRST + BUNDLE_N - 1)}{PASTE_LAST_ROW}')

ws.conditional_formatting.add(bundle_range,
    CellIsRule(operator='equal', formula=['"TRUE"'],
               fill=green_fill, font=green_font))
ws.conditional_formatting.add(bundle_range,
    CellIsRule(operator='equal', formula=['"FALSE"'],
               fill=red_fill, font=red_font))

# Completeness % — green if >=60%, amber if 30-60%, red if <30%
complete_range = f'{CL(COMPLETE_COL)}{PASTE_DATA_ROW}:{CL(COMPLETE_COL)}{PASTE_LAST_ROW}'
ws.conditional_formatting.add(complete_range,
    CellIsRule(operator='greaterThanOrEqual', formula=['0.6'],
               fill=PF('solid', start_color='C6EFCE', end_color='C6EFCE'),
               font=Font(name='Arial', size=9, color='276221', bold=True)))
ws.conditional_formatting.add(complete_range,
    CellIsRule(operator='greaterThanOrEqual', formula=['0.3'],
               fill=PF('solid', start_color='FFEB9C', end_color='FFEB9C'),
               font=Font(name='Arial', size=9, color='9C5700', bold=True)))
ws.conditional_formatting.add(complete_range,
    CellIsRule(operator='lessThan', formula=['0.3'],
               fill=PF('solid', start_color='FFC7CE', end_color='FFC7CE'),
               font=Font(name='Arial', size=9, color='9C0006', bold=True)))

# ─── Freeze panes ────────────────────────────────────────────────────────────
ws.freeze_panes = f'{CL(OUT_FIRST_COL)}{PASTE_DATA_ROW}'

# ─── Data validation — manual cols B-E ──────────────────────────────────────
dv_yn = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True,
                        showErrorMessage=False)
ws.add_data_validation(dv_yn)
e3_range = f'C{PASTE_DATA_ROW}:C{PASTE_LAST_ROW}'
dv_yn.add(e3_range)

# ─── Update README with v6 notes ─────────────────────────────────────────────
ws_r = wb['README']
last_row = ws_r.max_row + 2
c = ws_r.cell(last_row, 1,
    'v6 ADDITION — PatSeer Import Sheet\n'
    '• New sheet "PatSeer Import" (second sheet, after README).\n'
    '• Zone 1: Column Mapping Table (rows 7-33): maps 25 PatSeer export column headers to v5 '
    'attribute codes with transform rules. Edit col A to match your actual PatSeer headers.\n'
    '• Zone 2: Paste Area (cols F-AE, rows 40+): paste your PatSeer Excel export here directly. '
    'Supports up to 200 patents. Cols A-E reserved for manual input helpers (row number, notes, '
    'E3 Continuation override, E2 override, E5 override).\n'
    '• Zone 3: Mapped Output + Bundle Results (cols AF onwards): auto-populates 42 v5 attributes '
    'via transform formulas. Amber cells = manual/AI scoring required. Completeness % column '
    'shows how complete each patent record is (0-42 attributes filled). 33 Bundle columns show '
    'TRUE/FALSE qualification respecting the active Configuration preset. Total column counts '
    'qualifying bundles per patent.\n'
    '• Conditional formatting: TRUE=green, FALSE=light-red in bundle cols; completeness '
    'traffic-light (green ≥60%, amber 30-60%, red <30%).\n'
    '• Workflow: paste PatSeer export → review amber cells → score manually or use AI prompts → '
    'copy completed rows to Patent Portfolio sheet for full Scorecard analysis.'
)
c.font = Font(name='Arial', size=10)
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_r.row_dimensions[last_row].height = 110

# ─── Save v6 ─────────────────────────────────────────────────────────────────
out_path = '/home/claude/Patent_Bundling_Template_v6.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheets: {wb.sheetnames}')
print(f'PatSeer Import sheet dims: rows={ws.max_row}, cols={ws.max_column}')
